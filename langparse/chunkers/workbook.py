from __future__ import annotations

from collections.abc import Callable

from openpyxl.utils import get_column_letter, range_boundaries

from langparse.chunkers.profiles import (
    WorkbookChunkPolicy,
    WorkbookChunkProfile,
    resolve_workbook_chunk_policy,
)
from langparse.core.rendering import document_metadata
from langparse.types import Chunk, ParsedDocumentResult
from langparse.workbooks.types import (
    FormBlock,
    FormField,
    LogicalRow,
    LogicalTable,
    MatrixBlock,
    MatrixHeader,
    TableContinuation,
    TextBlock,
    TextLine,
    WorkbookBlock,
    WorkbookIR,
)


class WorkbookStructuralChunker:
    """Pack complete raw-grid rows directly from workbook compatibility facts."""

    def __init__(
        self,
        max_chunk_size: int | None = None,
        length_function: Callable[[str], int] = len,
        *,
        profile: str | WorkbookChunkProfile | None = None,
    ):
        self.policy: WorkbookChunkPolicy = resolve_workbook_chunk_policy(profile)
        resolved_size = (
            self.policy.default_max_chunk_size if max_chunk_size is None else max_chunk_size
        )
        if resolved_size <= 0:
            raise ValueError("max_chunk_size must be positive")
        self.max_chunk_size = resolved_size
        self.length_function = length_function

    def chunk(self, parsed: ParsedDocumentResult) -> list[Chunk]:
        if not isinstance(parsed.structure, WorkbookIR):
            raise TypeError("WorkbookStructuralChunker requires WorkbookIR structure")

        ir_sheets = {sheet.index: sheet for sheet in parsed.structure.sheets}
        chunks: list[Chunk] = []
        for page in parsed.pages:
            sheet_name = page.metadata.get("sheet_name")
            sheet_ir = ir_sheets.get(page.page_number - 1)
            if sheet_ir is not None and sheet_ir.blocks:
                for block in sheet_ir.blocks:
                    chunks.extend(
                        self._chunk_block(
                            parsed,
                            str(sheet_name),
                            page.page_number,
                            block,
                            len(chunks),
                        )
                    )
                continue
            confidence = (
                sheet_ir.blocks[0].confidence if sheet_ir is not None and sheet_ir.blocks else 1.0
            )
            for table in page.tables:
                rows = table.get("rows", [])
                if not rows:
                    continue
                columns = [str(value) for value in table.get("columns") or rows[0]]
                data_rows = [[str(value) for value in row] for row in rows[1:]]
                row_numbers = list(table.get("row_numbers", []))
                if len(row_numbers) != len(data_rows):
                    row_numbers = _row_numbers(table.get("source_range"), len(data_rows))
                chunks.extend(
                    self._pack_table(
                        parsed=parsed,
                        sheet_name=str(sheet_name),
                        sheet_ordinal=page.page_number,
                        columns=columns,
                        rows=data_rows,
                        row_numbers=row_numbers,
                        confidence=confidence,
                        chunk_index_offset=len(chunks),
                    )
                )
        self._finalize_chunks(parsed, chunks)
        self._validate_chunks(parsed, chunks)
        return chunks

    def _finalize_chunks(self, parsed: ParsedDocumentResult, chunks: list[Chunk]) -> None:
        workbook_ir = parsed.structure
        assert isinstance(workbook_ir, WorkbookIR)
        for index, chunk in enumerate(chunks):
            chunk.metadata["chunk_index"] = index
            chunk.metadata["chunk_profile"] = self.policy.name.value
            chunk.metadata["chunk_profile_version"] = self.policy.version
            ordinal = int(chunk.metadata["sheet_ordinal"])
            sheet_ir = workbook_ir.sheets[ordinal - 1]
            snapshot = workbook_ir.snapshot
            sheet_snapshot = snapshot.sheets[ordinal - 1] if snapshot is not None else None
            hidden_rows = set(sheet_snapshot.hidden_rows) if sheet_snapshot is not None else set()
            referenced_rows = set(chunk.metadata.get("row_numbers", []))
            if not referenced_rows:
                referenced_rows = _row_numbers_from_source_ranges(chunk.metadata["source_ranges"])
            chunk.metadata["sheet_visibility"] = (
                sheet_snapshot.visibility if sheet_snapshot is not None else sheet_ir.visibility
            )
            chunk.metadata["hidden_row_numbers"] = sorted(referenced_rows & hidden_rows)

    def _validate_chunks(self, parsed: ParsedDocumentResult, chunks: list[Chunk]) -> None:
        workbook_ir = parsed.structure
        assert isinstance(workbook_ir, WorkbookIR)
        expected_row_ids = [
            row.row_id
            for sheet in workbook_ir.sheets
            for block in sheet.blocks
            if block.logical_table is not None
            for row in block.logical_table.rows
            if row.role in {"data", "total"}
        ]
        actual_row_ids = [
            row_id
            for chunk in chunks
            if chunk.metadata["chunk_type"] == "table_rows"
            for row_id in chunk.metadata["row_ids"]
        ]
        if len(actual_row_ids) != len(set(actual_row_ids)) or set(actual_row_ids) != set(
            expected_row_ids
        ):
            raise ValueError("Workbook chunk row conservation failed")
        if [chunk.metadata["chunk_index"] for chunk in chunks] != list(range(len(chunks))):
            raise ValueError("Workbook chunk indexes are not contiguous")

        for chunk in chunks:
            source_ranges = chunk.metadata["source_ranges"]
            for source_range in source_ranges:
                _source_range_is_valid(workbook_ir.snapshot, source_range)
            if chunk.metadata["chunk_type"] != "table_rows":
                continue
            payload = chunk.structured_payload
            if len(chunk.metadata["row_ids"]) != len(payload["rows"]):
                raise ValueError("Workbook table chunk row payload mismatch")
            if self.policy.analysis_records:
                records = payload["records"]
                if len(chunk.metadata["row_ids"]) != len(records):
                    raise ValueError("Workbook table chunk analysis record mismatch")
                record_ranges = list(
                    dict.fromkeys(
                        source_ref for record in records for source_ref in record["source_refs"]
                    )
                )
                if source_ranges != record_ranges:
                    raise ValueError("Workbook table chunk source ranges mismatch")

    def _chunk_block(
        self,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        block: WorkbookBlock,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        if block.logical_table is not None:
            return self._chunk_logical_table(
                parsed,
                sheet_name,
                sheet_ordinal,
                block.logical_table,
                chunk_index_offset,
            )
        if block.form is not None:
            return self._chunk_form(
                parsed, sheet_name, sheet_ordinal, block.form, chunk_index_offset
            )
        if block.matrix is not None:
            return self._chunk_matrix(
                parsed, sheet_name, sheet_ordinal, block.matrix, chunk_index_offset
            )
        if block.text is not None:
            return self._chunk_text(
                parsed, sheet_name, sheet_ordinal, block.text, chunk_index_offset
            )
        source_range = block.source_refs[0].range
        columns, rows, row_numbers = _raw_block_grid(
            parsed.structure,
            sheet_ordinal,
            source_range,
        )
        return self._pack_table(
            parsed=parsed,
            sheet_name=sheet_name,
            sheet_ordinal=sheet_ordinal,
            columns=columns,
            rows=rows,
            row_numbers=row_numbers,
            confidence=block.confidence,
            chunk_index_offset=chunk_index_offset,
        )

    def _chunk_logical_table(
        self,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        table: LogicalTable,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        columns = [" / ".join(column.path) or column.coordinate for column in table.columns]
        continuation = _continuation_for_table(parsed.structure, table)
        eligible = [row for row in table.rows if row.role in {"data", "total"}]
        grouped: list[tuple[list[str], list[LogicalRow]]] = []
        for row in eligible:
            if not grouped or grouped[-1][0] != row.section_path:
                grouped.append((list(row.section_path), []))
            grouped[-1][1].append(row)

        chunks: list[Chunk] = []
        for section_path, rows in grouped:
            pending = []
            for row in rows:
                candidate = [*pending, row]
                content = _render_logical_chunk(table.title, section_path, columns, candidate)
                if pending and self.length_function(content) > self.max_chunk_size:
                    chunks.append(
                        _logical_chunk(
                            parsed,
                            table,
                            continuation,
                            sheet_name,
                            sheet_ordinal,
                            section_path,
                            columns,
                            pending,
                            chunk_index_offset + len(chunks),
                            self.max_chunk_size,
                            self.length_function,
                            self.policy,
                        )
                    )
                    pending = []
                pending.append(row)
            if pending:
                chunks.append(
                    _logical_chunk(
                        parsed,
                        table,
                        continuation,
                        sheet_name,
                        sheet_ordinal,
                        section_path,
                        columns,
                        pending,
                        chunk_index_offset + len(chunks),
                        self.max_chunk_size,
                        self.length_function,
                        self.policy,
                    )
                )
        return chunks

    def _chunk_form(
        self,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        form: FormBlock,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        chunks: list[Chunk] = []
        pending: list[FormField] = []

        def emit(*, oversized: bool = False) -> None:
            if not pending:
                return
            chunks.append(
                _form_chunk(
                    parsed,
                    form,
                    sheet_name,
                    sheet_ordinal,
                    pending,
                    [],
                    chunk_index_offset + len(chunks),
                    oversized,
                )
            )
            pending.clear()

        for field in form.fields:
            candidate = [*pending, field]
            content = _render_form_chunk(form.title, candidate, [])
            if pending and self.length_function(content) > self.max_chunk_size:
                emit()
                candidate = [field]
                content = _render_form_chunk(form.title, candidate, [])
            pending.append(field)
            if self.length_function(content) > self.max_chunk_size:
                emit(oversized=True)
        emit()
        if form.free_text:
            content = _render_form_chunk(form.title, [], form.free_text)
            chunks.append(
                _form_chunk(
                    parsed,
                    form,
                    sheet_name,
                    sheet_ordinal,
                    [],
                    form.free_text,
                    chunk_index_offset + len(chunks),
                    self.length_function(content) > self.max_chunk_size,
                )
            )
        return chunks

    def _chunk_matrix(
        self,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        matrix: MatrixBlock,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        chunks: list[Chunk] = []
        pending: list[tuple[MatrixHeader, list[str], list]] = []

        def emit(*, oversized: bool = False) -> None:
            if not pending:
                return
            chunks.append(
                _matrix_chunk(
                    parsed,
                    matrix,
                    sheet_name,
                    sheet_ordinal,
                    pending,
                    chunk_index_offset + len(chunks),
                    oversized,
                )
            )
            pending.clear()

        rows = zip(
            matrix.row_headers,
            matrix.values,
            matrix.value_source_refs,
            strict=True,
        )
        for header, values, refs in rows:
            candidate = [*pending, (header, values, refs)]
            content = _render_matrix_chunk(matrix, candidate)
            if pending and self.length_function(content) > self.max_chunk_size:
                emit()
                candidate = [(header, values, refs)]
                content = _render_matrix_chunk(matrix, candidate)
            pending.append((header, values, refs))
            if self.length_function(content) > self.max_chunk_size:
                emit(oversized=True)
        emit()
        return chunks

    def _chunk_text(
        self,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        text: TextBlock,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        chunks: list[Chunk] = []
        pending: list[TextLine] = []

        def emit(*, oversized: bool = False) -> None:
            if not pending:
                return
            chunks.append(
                _text_chunk(
                    parsed,
                    text,
                    sheet_name,
                    sheet_ordinal,
                    pending,
                    chunk_index_offset + len(chunks),
                    oversized,
                )
            )
            pending.clear()

        for line in text.lines:
            candidate = [*pending, line]
            content = "\n".join(item.text for item in candidate)
            if pending and self.length_function(content) > self.max_chunk_size:
                emit()
                candidate = [line]
                content = line.text
            pending.append(line)
            if self.length_function(content) > self.max_chunk_size:
                emit(oversized=True)
        emit()
        return chunks

    def _pack_table(
        self,
        *,
        parsed: ParsedDocumentResult,
        sheet_name: str,
        sheet_ordinal: int,
        columns: list[str],
        rows: list[list[str]],
        row_numbers: list[int],
        confidence: float,
        chunk_index_offset: int,
    ) -> list[Chunk]:
        packed: list[Chunk] = []
        pending_rows: list[list[str]] = []
        pending_numbers: list[int] = []

        def emit(*, oversized: bool = False) -> None:
            if not pending_rows:
                return
            source_range = _source_range(sheet_name, columns, pending_numbers)
            content = _render_chunk(sheet_name, source_range, columns, pending_rows)
            payload = {
                "columns": list(columns),
                "rows": [list(row) for row in pending_rows],
            }
            if self.policy.analysis_records:
                payload["column_schema"] = [
                    {
                        "column_index": index,
                        "coordinate": column,
                        "header_path": [],
                    }
                    for index, column in enumerate(columns)
                ]
                payload["records"] = [
                    {
                        "row_number": row_number,
                        "role": "raw",
                        "section_path": [],
                        "values": list(row),
                        "source_refs": [_source_range(sheet_name, columns, [row_number])],
                    }
                    for row, row_number in zip(pending_rows, pending_numbers, strict=True)
                ]
            metadata = document_metadata(parsed)
            metadata.update(
                {
                    "chunk_type": "raw_grid_rows",
                    "chunk_index": chunk_index_offset + len(packed),
                    "sheet_name": sheet_name,
                    "sheet_ordinal": sheet_ordinal,
                    "source_ranges": [source_range],
                    "row_numbers": list(pending_numbers),
                    "confidence": confidence,
                    "warnings": list(parsed.diagnostics.warnings)
                    if parsed.diagnostics is not None
                    else [],
                }
            )
            if oversized:
                metadata["oversized"] = True
            packed.append(
                Chunk(
                    content=content,
                    metadata=metadata,
                    structured_payload=payload,
                )
            )
            pending_rows.clear()
            pending_numbers.clear()

        for row, row_number in zip(rows, row_numbers, strict=True):
            candidate_rows = [*pending_rows, row]
            candidate_numbers = [*pending_numbers, row_number]
            candidate_range = _source_range(sheet_name, columns, candidate_numbers)
            candidate = _render_chunk(sheet_name, candidate_range, columns, candidate_rows)
            if pending_rows and self.length_function(candidate) > self.max_chunk_size:
                emit()
                candidate_rows = [row]
                candidate_numbers = [row_number]
                candidate_range = _source_range(sheet_name, columns, candidate_numbers)
                candidate = _render_chunk(sheet_name, candidate_range, columns, candidate_rows)

            pending_rows.append(row)
            pending_numbers.append(row_number)
            if self.length_function(candidate) > self.max_chunk_size:
                emit(oversized=True)

        emit()
        return packed


def _row_numbers(source_range: str | None, count: int) -> list[int]:
    if not source_range:
        return list(range(1, count + 1))
    _, min_row, _, _ = range_boundaries(source_range)
    return list(range(min_row, min_row + count))


def _row_numbers_from_source_ranges(source_ranges: list[str]) -> set[int]:
    row_numbers = set()
    for source_range in source_ranges:
        _, cell_range = source_range.rsplit("!", 1)
        _, min_row, _, max_row = range_boundaries(cell_range)
        row_numbers.update(range(min_row, max_row + 1))
    return row_numbers


def _source_range_is_valid(snapshot, source_ref: str) -> None:
    if snapshot is None:
        raise ValueError("WorkbookIR snapshot is required for source-range validation")
    sheet_name, cell_range = source_ref.rsplit("!", 1)
    sheet = next((item for item in snapshot.sheets if item.name == sheet_name), None)
    if sheet is None:
        raise ValueError(f"Workbook source range references unknown sheet: {sheet_name}")
    if sheet.used_range is None:
        raise ValueError(f"Workbook sheet used_range is required: {sheet_name}")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        used_min_col, used_min_row, used_max_col, used_max_row = range_boundaries(sheet.used_range)
    except ValueError as exc:
        raise ValueError(f"Workbook source range is invalid: {source_ref}") from exc
    if not (
        used_min_col <= min_col <= max_col <= used_max_col
        and used_min_row <= min_row <= max_row <= used_max_row
    ):
        raise ValueError(f"Workbook source range is outside sheet used_range: {source_ref}")


def _source_range(sheet_name: str, columns: list[str], row_numbers: list[int]) -> str:
    first_row = min(row_numbers)
    last_row = max(row_numbers)
    return f"{sheet_name}!{columns[0]}{first_row}:{columns[-1]}{last_row}"


def _render_chunk(
    sheet_name: str,
    source_range: str,
    columns: list[str],
    rows: list[list[str]],
) -> str:
    heading = f"## Sheet: {sheet_name}"
    source_comment = f"<!-- source_range: {source_range} -->"
    header = _markdown_row(columns)
    separator = "| " + " | ".join("---" for _ in columns) + " |"
    body = [_markdown_row(row) for row in rows]
    return "\n\n".join((heading, source_comment, "\n".join((header, separator, *body))))


def _markdown_row(row: list[str]) -> str:
    escaped = [
        str(value)
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("|", r"\|")
        .replace("\n", "<br>")
        for value in row
    ]
    return "| " + " | ".join(escaped) + " |"


def _render_form_chunk(
    title: str,
    fields: list[FormField],
    lines: list[TextLine],
) -> str:
    parts = [f"### Form: {title}"] if title else []
    if fields:
        parts.append(
            "\n".join(
                [
                    _markdown_row(["Field", "Value"]),
                    "| --- | --- |",
                    *[_markdown_row([field.label, field.value]) for field in fields],
                ]
            )
        )
    parts.extend(line.text for line in lines)
    return "\n\n".join(parts)


def _form_chunk(
    parsed: ParsedDocumentResult,
    form: FormBlock,
    sheet_name: str,
    sheet_ordinal: int,
    fields: list[FormField],
    lines: list[TextLine],
    chunk_index: int,
    oversized: bool,
) -> Chunk:
    metadata = document_metadata(parsed)
    source_ranges = [
        ref.key for field in fields for ref in [*field.label_source_refs, *field.value_source_refs]
    ]
    source_ranges.extend(ref.key for line in lines for ref in line.source_refs)
    metadata.update(
        {
            "chunk_type": "form_fields",
            "chunk_index": chunk_index,
            "sheet_name": sheet_name,
            "sheet_ordinal": sheet_ordinal,
            "form_id": form.form_id,
            "field_ids": [field.field_id for field in fields],
            "source_ranges": source_ranges,
            "confidence": min([form.confidence, *[field.confidence for field in fields]]),
            "warnings": list(parsed.diagnostics.warnings) if parsed.diagnostics is not None else [],
        }
    )
    if oversized:
        metadata["oversized"] = True
    return Chunk(
        content=_render_form_chunk(form.title, fields, lines),
        metadata=metadata,
        structured_payload={
            "fields": [[field.label, field.value] for field in fields],
            "free_text": [line.text for line in lines],
        },
    )


def _render_matrix_chunk(matrix: MatrixBlock, rows: list[tuple]) -> str:
    parts = [f"### Matrix: {matrix.title}"] if matrix.title else []
    columns = ["", *[header.value for header in matrix.column_headers]]
    table_lines = [
        _markdown_row(columns),
        "| " + " | ".join("---" for _ in columns) + " |",
        *[_markdown_row([header.value, *values]) for header, values, _ in rows],
    ]
    parts.append("\n".join(table_lines))
    return "\n\n".join(parts)


def _matrix_chunk(
    parsed: ParsedDocumentResult,
    matrix: MatrixBlock,
    sheet_name: str,
    sheet_ordinal: int,
    rows: list[tuple],
    chunk_index: int,
    oversized: bool,
) -> Chunk:
    metadata = document_metadata(parsed)
    source_ranges = []
    for header, _, refs in rows:
        source_ranges.extend(ref.key for ref in header.source_refs)
        source_ranges.extend(ref.key for ref in refs if ref is not None)
    metadata.update(
        {
            "chunk_type": "matrix_rows",
            "chunk_index": chunk_index,
            "sheet_name": sheet_name,
            "sheet_ordinal": sheet_ordinal,
            "matrix_id": matrix.matrix_id,
            "row_headers": [header.value for header, _, _ in rows],
            "source_ranges": source_ranges,
            "confidence": matrix.confidence,
            "warnings": list(parsed.diagnostics.warnings) if parsed.diagnostics is not None else [],
        }
    )
    if oversized:
        metadata["oversized"] = True
    return Chunk(
        content=_render_matrix_chunk(matrix, rows),
        metadata=metadata,
        structured_payload={
            "column_headers": [header.value for header in matrix.column_headers],
            "row_headers": [header.value for header, _, _ in rows],
            "values": [list(values) for _, values, _ in rows],
        },
    )


def _text_chunk(
    parsed: ParsedDocumentResult,
    text: TextBlock,
    sheet_name: str,
    sheet_ordinal: int,
    lines: list[TextLine],
    chunk_index: int,
    oversized: bool,
) -> Chunk:
    metadata = document_metadata(parsed)
    metadata.update(
        {
            "chunk_type": "text_block",
            "chunk_index": chunk_index,
            "sheet_name": sheet_name,
            "sheet_ordinal": sheet_ordinal,
            "text_id": text.text_id,
            "source_ranges": [ref.key for line in lines for ref in line.source_refs],
            "confidence": text.confidence,
            "warnings": list(parsed.diagnostics.warnings) if parsed.diagnostics is not None else [],
        }
    )
    if oversized:
        metadata["oversized"] = True
    return Chunk(
        content="\n".join(line.text for line in lines),
        metadata=metadata,
        structured_payload={"lines": [line.text for line in lines]},
    )


def _raw_block_grid(
    workbook_ir: WorkbookIR,
    sheet_ordinal: int,
    source_range: str,
) -> tuple[list[str], list[list[str]], list[int]]:
    if workbook_ir.snapshot is None:
        raise ValueError("WorkbookIR snapshot is required for raw block chunks")
    sheet = workbook_ir.snapshot.sheets[sheet_ordinal - 1]
    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    columns = [get_column_letter(column) for column in range(min_col, max_col + 1)]
    row_numbers = list(range(min_row, max_row + 1))
    rows = []
    for row_number in row_numbers:
        row = []
        for column in range(min_col, max_col + 1):
            coordinate = f"{get_column_letter(column)}{row_number}"
            cell = sheet.cells.get(coordinate)
            row.append("" if cell is None or cell.merge_anchor is not None else cell.display_value)
        rows.append(row)
    return columns, rows, row_numbers


def _render_logical_chunk(
    title: str,
    section_path: list[str],
    columns: list[str],
    rows: list[LogicalRow],
) -> str:
    headings = [f"### Table: {title}"] if title else []
    if section_path:
        headings.append(f"#### Section: {' / '.join(section_path)}")
    table_lines = [
        _markdown_row(columns),
        "| " + " | ".join("---" for _ in columns) + " |",
        *[_markdown_row(row.values) for row in rows],
    ]
    return "\n\n".join([*headings, "\n".join(table_lines)])


def _logical_chunk(
    parsed: ParsedDocumentResult,
    table: LogicalTable,
    continuation: TableContinuation | None,
    sheet_name: str,
    sheet_ordinal: int,
    section_path: list[str],
    columns: list[str],
    rows: list[LogicalRow],
    chunk_index: int,
    max_chunk_size: int,
    length_function: Callable[[str], int],
    policy: WorkbookChunkPolicy,
) -> Chunk:
    content = _render_logical_chunk(table.title, section_path, columns, rows)
    metadata = document_metadata(parsed)
    metadata.update(
        {
            "chunk_type": "table_rows",
            "chunk_index": chunk_index,
            "sheet_name": sheet_name,
            "sheet_ordinal": sheet_ordinal,
            "table_id": table.table_id,
            "section_path": list(section_path),
            "header_paths": [list(column.path) for column in table.columns],
            "row_ids": [row.row_id for row in rows],
            "row_numbers": [row.metadata["row_number"] for row in rows],
            "source_ranges": [row.source_ref.key for row in rows],
            "fragment_ranges": _fragment_ranges_for_rows(table, rows),
            "confidence": min([table.confidence, *[row.confidence for row in rows]]),
            "warnings": list(parsed.diagnostics.warnings) if parsed.diagnostics is not None else [],
        }
    )
    if continuation is not None:
        metadata.update(
            {
                "continuation_id": continuation.continuation_id,
                "continuation_role": table.continuation_role,
                "continuation_member_table_ids": list(continuation.member_table_ids),
                "continuation_source_ranges": [ref.key for ref in continuation.source_refs],
            }
        )
    if length_function(content) > max_chunk_size:
        metadata["oversized"] = True
    payload = {
        "columns": columns,
        "rows": [list(row.values) for row in rows],
        "roles": [row.role for row in rows],
    }
    if policy.analysis_records:
        payload["column_schema"] = [
            {
                "column_index": index,
                "coordinate": column.coordinate,
                "header_path": list(column.path),
            }
            for index, column in enumerate(table.columns)
        ]
        payload["records"] = [
            {
                "row_id": row.row_id,
                "row_number": int(row.metadata["row_number"]),
                "role": row.role,
                "section_path": list(row.section_path),
                "values": list(row.values),
                "source_refs": [row.source_ref.key],
            }
            for row in rows
        ]
    return Chunk(
        content=content,
        metadata=metadata,
        structured_payload=payload,
    )


def _continuation_for_table(
    workbook_ir: WorkbookIR,
    table: LogicalTable,
) -> TableContinuation | None:
    if table.continuation_id is None:
        return None
    return next(
        (
            group
            for group in workbook_ir.table_continuations
            if group.continuation_id == table.continuation_id
            and table.table_id in group.member_table_ids
        ),
        None,
    )


def _fragment_ranges_for_rows(table: LogicalTable, rows: list[LogicalRow]) -> list[str]:
    row_numbers = {int(row.metadata["row_number"]) for row in rows}
    ranges = []
    for fragment in table.fragments:
        _, min_row, _, max_row = range_boundaries(fragment.source_ref.range)
        if any(min_row <= row_number <= max_row for row_number in row_numbers):
            ranges.append(fragment.source_ref.key)
    return ranges
