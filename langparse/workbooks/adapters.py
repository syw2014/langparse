from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries

from langparse.workbooks.types import CellSnapshot, SheetSnapshot, WorkbookSnapshot


class WorkbookAdapter(Protocol):
    def snapshot(self, path: str | Path) -> WorkbookSnapshot: ...


class OOXMLWorkbookAdapter:
    """Extract OOXML workbook facts without interpreting table semantics."""

    def snapshot(self, path: str | Path) -> WorkbookSnapshot:
        workbook_path = Path(path)
        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        # File-like input intentionally bypasses openpyxl's extension gate.
        # LangParse routes by content, so a valid OOXML workbook renamed to
        # ``.csv`` must still be readable as a workbook.
        formula_stream = workbook_path.open("rb")
        value_stream = workbook_path.open("rb")
        warnings: list[str] = []
        try:
            formula_book = load_workbook(
                formula_stream,
                data_only=False,
                keep_vba=keep_vba,
                read_only=False,
            )
            value_book = load_workbook(
                value_stream,
                data_only=True,
                keep_vba=keep_vba,
                read_only=False,
            )
            sheets = [
                self._snapshot_sheet(
                    formula_sheet,
                    value_book[formula_sheet.title],
                    index,
                    warnings,
                )
                for index, formula_sheet in enumerate(formula_book.worksheets)
            ]
        finally:
            if "formula_book" in locals():
                formula_book.close()
            if "value_book" in locals():
                value_book.close()
            formula_stream.close()
            value_stream.close()

        return WorkbookSnapshot(
            source=str(workbook_path),
            filename=workbook_path.name,
            sheets=sheets,
            metadata={"format": workbook_path.suffix.lower(), "warnings": warnings},
        )

    def _snapshot_sheet(
        self,
        formula_sheet: Any,
        value_sheet: Any,
        index: int,
        warnings: list[str],
    ) -> SheetSnapshot:
        hidden_rows = sorted(
            row_index
            for row_index, dimension in formula_sheet.row_dimensions.items()
            if dimension.hidden
        )
        hidden_columns = sorted(
            column
            for column, dimension in formula_sheet.column_dimensions.items()
            if dimension.hidden
        )
        row_heights = {
            row_index: float(dimension.height)
            for row_index, dimension in formula_sheet.row_dimensions.items()
            if dimension.height is not None
        }
        column_widths = {
            column: float(dimension.width)
            for column, dimension in formula_sheet.column_dimensions.items()
            if dimension.width is not None
        }

        cells: dict[str, CellSnapshot] = {}
        for row in formula_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or not _should_capture(cell):
                    continue
                cached_cell = value_sheet[cell.coordinate]
                formula = cell.value if cell.data_type == "f" else None
                cached_value = cached_cell.value if formula is not None else cell.value
                cells[cell.coordinate] = CellSnapshot(
                    coordinate=cell.coordinate,
                    raw_value=cell.value,
                    display_value=_display_value(
                        cached_value if cached_value is not None else cell.value
                    ),
                    formula=formula,
                    cached_value=cached_value,
                    data_type=str(cell.data_type or ""),
                    number_format=cell.number_format or "General",
                    style_id=_style_fingerprint(cell),
                    hyperlink=_hyperlink_value(cell.hyperlink),
                    comment=cell.comment.text if cell.comment is not None else None,
                    hidden=cell.row in hidden_rows or cell.column_letter in hidden_columns,
                )

        merged_ranges = sorted(str(cell_range) for cell_range in formula_sheet.merged_cells.ranges)
        for merged_range in formula_sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            anchor = f"{get_column_letter(min_col)}{min_row}"
            anchor_cell = formula_sheet[anchor]
            anchor_snapshot = cells.setdefault(
                anchor,
                CellSnapshot(
                    coordinate=anchor,
                    raw_value=anchor_cell.value,
                    display_value=_display_value(anchor_cell.value),
                    data_type=str(anchor_cell.data_type or ""),
                    number_format=anchor_cell.number_format or "General",
                    style_id=_style_fingerprint(anchor_cell),
                ),
            )
            anchor_snapshot.rowspan = max_row - min_row + 1
            anchor_snapshot.colspan = max_col - min_col + 1
            for row_index in range(min_row, max_row + 1):
                for column_index in range(min_col, max_col + 1):
                    coordinate = f"{get_column_letter(column_index)}{row_index}"
                    if coordinate == anchor:
                        continue
                    cells.setdefault(
                        coordinate,
                        CellSnapshot(coordinate=coordinate, merge_anchor=anchor),
                    ).merge_anchor = anchor

        objects = []
        for chart in formula_sheet._charts:
            objects.append(
                {
                    "kind": "chart",
                    "anchor": _object_anchor(chart, formula_sheet.title, warnings),
                    "title": _chart_title(chart),
                }
            )
        for image in formula_sheet._images:
            objects.append(
                {
                    "kind": "image",
                    "anchor": _object_anchor(image, formula_sheet.title, warnings),
                    "width": getattr(image, "width", None),
                    "height": getattr(image, "height", None),
                }
            )

        return SheetSnapshot(
            name=formula_sheet.title,
            index=index,
            visibility=formula_sheet.sheet_state,
            used_range=_used_range(cells),
            print_area=_print_areas(formula_sheet.print_area),
            row_heights=row_heights,
            column_widths=column_widths,
            hidden_rows=hidden_rows,
            hidden_columns=hidden_columns,
            merged_ranges=merged_ranges,
            cells=dict(sorted(cells.items(), key=lambda item: _coordinate_sort_key(item[0]))),
            objects=objects,
        )


def _should_capture(cell: Any) -> bool:
    return any(
        (
            cell.value is not None,
            cell.has_style,
            cell.comment is not None,
            cell.hyperlink is not None,
        )
    )


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _hyperlink_value(hyperlink: Any) -> str | None:
    if hyperlink is None:
        return None
    return hyperlink.target or hyperlink.location


def _color_payload(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb if color.type == "rgb" else None,
        "indexed": color.indexed if color.type == "indexed" else None,
        "theme": color.theme if color.type == "theme" else None,
        "tint": color.tint,
    }


def _side_payload(side: Any) -> dict[str, Any]:
    return {"style": side.style, "color": _color_payload(side.color)}


def _style_fingerprint(cell: Any) -> str:
    if not cell.has_style:
        return ""
    payload = {
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.b,
            "italic": cell.font.i,
            "underline": cell.font.u,
            "strike": cell.font.strike,
            "color": _color_payload(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "foreground": _color_payload(cell.fill.fgColor),
            "background": _color_payload(cell.fill.bgColor),
        },
        "border": {
            "left": _side_payload(cell.border.left),
            "right": _side_payload(cell.border.right),
            "top": _side_payload(cell.border.top),
            "bottom": _side_payload(cell.border.bottom),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "text_rotation": cell.alignment.text_rotation,
        },
        "number_format": cell.number_format,
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()[:16]


def _coordinate_sort_key(coordinate: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", coordinate)
    if match is None:
        return (0, 0)
    column = 0
    for character in match.group(1):
        column = column * 26 + ord(character) - ord("A") + 1
    return (int(match.group(2)), column)


def _used_range(cells: dict[str, CellSnapshot]) -> str | None:
    if not cells:
        return None
    coordinates = [_coordinate_sort_key(coordinate) for coordinate in cells]
    rows = [row for row, _ in coordinates]
    columns = [column for _, column in coordinates]
    return (
        f"{get_column_letter(min(columns))}{min(rows)}:{get_column_letter(max(columns))}{max(rows)}"
    )


def _print_areas(print_area: Any) -> list[str]:
    if not print_area:
        return []
    parts = re.split(r",(?=(?:[^']*'[^']*')*[^']*$)", str(print_area))
    return [re.sub(r"^'([^']+)'!", r"\1!", part.strip()) for part in parts]


def _object_anchor(obj: Any, sheet_name: str, warnings: list[str]) -> str | None:
    anchor = getattr(obj, "anchor", None)
    if isinstance(anchor, str):
        return anchor
    marker = getattr(anchor, "_from", None)
    if marker is not None:
        return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
    warnings.append(f"Unable to resolve object anchor on sheet {sheet_name}")
    return None


def _chart_title(chart: Any) -> str | None:
    title = getattr(chart, "title", None)
    if title is None:
        return None
    return str(title)
