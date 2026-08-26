from __future__ import annotations

from collections import Counter
from copy import deepcopy
from dataclasses import asdict, dataclass, fields, replace

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from langparse.types import ParseDiagnostics
from langparse.workbooks.blocks import (
    interpret_form_block,
    interpret_matrix_block,
    interpret_text_block,
)
from langparse.workbooks.classification import (
    BlockClassification,
    RegionAssessment,
    assess_candidate_region,
    classify_candidate_region,
)
from langparse.workbooks.continuation import link_table_continuations
from langparse.workbooks.modeling import (
    InvalidRegionAmbiguityCaseError,
    ModelCallAudit,
    RegionAmbiguityCase,
    RequiredWorkbookDisambiguationError,
    WorkbookDisambiguation,
    WorkbookModelMode,
    WorkbookRegionDisambiguator,
    build_region_case,
)
from langparse.workbooks.modeling.disambiguation import _audit_payload
from langparse.workbooks.modeling.types import REGION_RULE_VERSION
from langparse.workbooks.regions import detect_candidate_regions
from langparse.workbooks.tables import interpret_logical_table
from langparse.workbooks.types import (
    CandidateRegion,
    CellSnapshot,
    LogicalTable,
    SheetIR,
    SheetSnapshot,
    SourceRef,
    WorkbookBlock,
    WorkbookIR,
    WorkbookSnapshot,
    stable_id,
)


@dataclass(frozen=True)
class _RegionDraft:
    sheet_index: int
    sheet: SheetSnapshot
    candidate: CandidateRegion
    assessment: RegionAssessment
    case: RegionAmbiguityCase | None
    case_id: str | None = None
    unavailable_audit: ModelCallAudit | None = None


@dataclass(frozen=True)
class _MaterializedRegion:
    draft: _RegionDraft
    block: WorkbookBlock
    deterministic_block: WorkbookBlock
    audit: ModelCallAudit | None = None
    model_selected: bool = False
    materialization_failed: bool = False


def assemble_workbook(
    snapshot: WorkbookSnapshot,
    *,
    disambiguation: WorkbookDisambiguation | None = None,
) -> tuple[WorkbookIR, ParseDiagnostics]:
    """Classify and interpret candidate regions with local raw-grid fallback."""

    configured = WorkbookDisambiguation.off() if disambiguation is None else disambiguation
    if configured.mode is WorkbookModelMode.OFF:
        return _assemble_deterministic(snapshot)

    drafts = _region_drafts(snapshot, configured)
    try:
        resolutions_by_case_id = _resolve_region_cases(drafts, configured)
    except RequiredWorkbookDisambiguationError as error:
        _raise_required_with_deterministic_fallback(snapshot, drafts, error)
    materialized = [
        _materialize_region(snapshot.source, draft, resolutions_by_case_id) for draft in drafts
    ]
    workbook_ir = _workbook_from_materialized(snapshot, materialized, rollback_selected=False)
    diagnostics, tentative_validation_codes = _finalize_workbook(snapshot, workbook_ir)

    selected_regions = [region for region in materialized if region.model_selected]
    reverted_case_ids: set[str] = set()
    if tentative_validation_codes and selected_regions:
        reverted_case_ids = {
            region.draft.case.case_id
            for region in selected_regions
            if region.draft.case is not None
        }
        workbook_ir = _workbook_from_materialized(snapshot, materialized, rollback_selected=True)
        diagnostics, rollback_validation_codes = _finalize_workbook(snapshot, workbook_ir)
        if set(rollback_validation_codes) - {"continuation_error"}:
            raise RuntimeError("deterministic workbook rollback failed validation")

    unresolved_case_ids: list[str] = []
    finalized_audits: list[ModelCallAudit] = []
    for region in materialized:
        if region.audit is None or region.draft.case_id is None:
            continue
        case_id = region.draft.case_id
        audit = region.audit
        if region.draft.case is None:
            if configured.mode is WorkbookModelMode.REQUIRED:
                unresolved_case_ids.append(case_id)
        elif region.materialization_failed:
            if configured.mode is WorkbookModelMode.REQUIRED:
                unresolved_case_ids.append(case_id)
        elif case_id in reverted_case_ids:
            audit = replace(
                audit,
                outcome="validation_error",
                validation_codes=_stable_codes(
                    *audit.validation_codes,
                    *tentative_validation_codes,
                ),
                reason_codes=("deterministic_fallback",),
                error_type=None,
            )
            if configured.mode is WorkbookModelMode.REQUIRED:
                unresolved_case_ids.append(case_id)
        elif region.model_selected:
            audit = replace(
                audit,
                outcome="accepted",
                reason_codes=("model_selected_choice",),
                error_type=None,
            )
        finalized_audits.append(audit)

    diagnostics.model_calls = [_audit_payload(audit) for audit in finalized_audits]
    if unresolved_case_ids:
        diagnostics.status = "failed"
        raise RequiredWorkbookDisambiguationError(
            tuple(unresolved_case_ids),
            diagnostics,
        )
    return workbook_ir, diagnostics


def _assemble_deterministic(
    snapshot: WorkbookSnapshot,
) -> tuple[WorkbookIR, ParseDiagnostics]:
    """Preserve the pre-model semantic assembly path byte-for-byte in behavior."""

    workbook_ir, diagnostics = assemble_baseline(snapshot)
    block_counts: Counter[str] = Counter()
    ambiguous_regions = []
    for sheet, sheet_ir in zip(snapshot.sheets, workbook_ir.sheets, strict=True):
        semantic_blocks: list[WorkbookBlock] = []
        for candidate in detect_candidate_regions(sheet):
            try:
                classification = classify_candidate_region(sheet, candidate)
                block = _block_for_candidate(
                    snapshot.source,
                    sheet,
                    candidate,
                    classification,
                )
            except Exception as exc:
                block = _unclassified_block(
                    snapshot.source,
                    candidate,
                    confidence=0.0,
                    reason_codes=["semantic_block_fallback"],
                    extra_diagnostic={"error_type": type(exc).__name__},
                )
            if block.kind == "unclassified":
                reason_codes = [
                    diagnostic["reason_code"]
                    for diagnostic in block.diagnostics
                    if "reason_code" in diagnostic
                ]
                ambiguous_regions.append(
                    {
                        "sheet_name": sheet.name,
                        "range": candidate.source_ref.range,
                        "candidate_kind": "unclassified",
                        "confidence": block.confidence,
                        "reason_codes": reason_codes,
                    }
                )
            semantic_blocks.append(block)
            block_counts[block.kind] += 1
        sheet_ir.blocks = semantic_blocks

    diagnostics.block_count_by_kind = dict(sorted(block_counts.items()))
    diagnostics.ambiguous_regions = ambiguous_regions
    try:
        groups, candidates = link_table_continuations(snapshot, workbook_ir)
    except Exception as exc:
        diagnostics.warnings.append(f"cross_sheet_continuation_fallback:{type(exc).__name__}")
    else:
        workbook_ir.table_continuations = groups
        diagnostics.continuation_candidates = candidates
        ambiguous_count = sum(item["status"] == "ambiguous" for item in candidates)
        if ambiguous_count:
            diagnostics.warnings.append(
                f"Workbook contains {ambiguous_count} ambiguous continuation candidates"
            )
    _update_coverage(snapshot, workbook_ir, diagnostics)
    validity_ratio, invalid_refs = validate_workbook_source_refs(snapshot, workbook_ir)
    diagnostics.source_ref_validity_ratio = validity_ratio
    if invalid_refs:
        diagnostics.status = "partial"
        diagnostics.warnings.append(
            f"Workbook IR contains {len(invalid_refs)} invalid source refs: {invalid_refs[:10]}"
        )
    return workbook_ir, diagnostics


def _region_drafts(
    snapshot: WorkbookSnapshot,
    configured: WorkbookDisambiguation,
) -> list[_RegionDraft]:
    drafts = []
    for sheet_index, sheet in enumerate(snapshot.sheets):
        for candidate in detect_candidate_regions(sheet):
            assessment = assess_candidate_region(sheet, candidate)
            case = None
            case_id = None
            unavailable_audit = None
            if configured.mode is not WorkbookModelMode.OFF and assessment.ambiguous:
                case_id = _local_region_case_id(candidate, assessment)
                unavailable_outcome = _unavailable_case_outcome(sheet, candidate)
                if unavailable_outcome is not None:
                    unavailable_audit = _local_unavailable_audit(
                        case_id,
                        candidate,
                        configured,
                        outcome=unavailable_outcome,
                    )
                else:
                    try:
                        case = build_region_case(sheet, candidate, assessment)
                    except InvalidRegionAmbiguityCaseError as error:
                        unavailable_audit = _local_unavailable_audit(
                            case_id,
                            candidate,
                            configured,
                            outcome="case_unavailable",
                            error_type=type(error).__name__,
                        )
                    else:
                        case_id = case.case_id
            drafts.append(
                _RegionDraft(
                    sheet_index=sheet_index,
                    sheet=sheet,
                    candidate=candidate,
                    assessment=assessment,
                    case=case,
                    case_id=case_id,
                    unavailable_audit=unavailable_audit,
                )
            )
    return drafts


def _local_region_case_id(
    candidate: CandidateRegion,
    assessment: RegionAssessment,
) -> str:
    return stable_id(
        "region_case_unavailable",
        REGION_RULE_VERSION,
        candidate.source_ref.key,
        *(choice.choice_id for choice in assessment.choices),
    )


def _unavailable_case_outcome(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
) -> str | None:
    if sheet.visibility != "visible":
        return "hidden_content"
    min_column, min_row, max_column, max_row = range_boundaries(candidate.source_ref.range)
    if any(min_row <= row <= max_row for row in sheet.hidden_rows):
        return "hidden_content"
    hidden_columns = {get_column_letter(column) for column in range(min_column, max_column + 1)}
    if hidden_columns.intersection(sheet.hidden_columns):
        return "hidden_content"
    for coordinate, cell in sheet.cells.items():
        row, column = coordinate_to_tuple(coordinate)
        if min_row <= row <= max_row and min_column <= column <= max_column and cell.hidden:
            return "hidden_content"
    return None


def _local_unavailable_audit(
    case_id: str,
    candidate: CandidateRegion,
    configured: WorkbookDisambiguation,
    *,
    outcome: str,
    error_type: str | None = None,
) -> ModelCallAudit:
    return ModelCallAudit(
        case_id=case_id,
        source_range=candidate.source_ref.range,
        mode=configured.mode.value,
        provider=None,
        model=None,
        model_revision=None,
        request_checksum=None,
        response_checksum=None,
        cache_status="not_checked",
        attempts=0,
        elapsed_ms=0,
        request_bytes=0,
        response_bytes=0,
        outcome=outcome,
        selected_choice_id=None,
        reported_confidence=None,
        validation_codes=(outcome,),
        reason_codes=("deterministic_fallback",),
        error_type=error_type,
    )


def _resolve_region_cases(
    drafts: list[_RegionDraft],
    configured: WorkbookDisambiguation,
):
    cases = [draft.case for draft in drafts if draft.case is not None]
    if not cases:
        return {}
    resolutions = WorkbookRegionDisambiguator().resolve(cases, configured)
    return {resolution.case_id: resolution for resolution in resolutions.resolutions}


def _raise_required_with_deterministic_fallback(
    snapshot: WorkbookSnapshot,
    drafts: list[_RegionDraft],
    error: RequiredWorkbookDisambiguationError,
) -> None:
    materialized = [_deterministic_materialized_region(snapshot.source, draft) for draft in drafts]
    workbook_ir = _workbook_from_materialized(snapshot, materialized, rollback_selected=False)
    diagnostics, _ = _finalize_workbook(snapshot, workbook_ir)
    diagnostics.model_calls = _ordered_error_audits(drafts, error.diagnostics.model_calls)
    diagnostics.status = "failed"
    raise RequiredWorkbookDisambiguationError(
        _ordered_unresolved_case_ids(drafts, error.case_ids),
        diagnostics,
    ) from None


def _ordered_error_audits(
    drafts: list[_RegionDraft],
    error_audits: list[dict[str, object]],
) -> list[dict[str, object]]:
    audits_by_case_id = {
        audit["case_id"]: _audit_field_payload(audit)
        for audit in error_audits
        if isinstance(audit.get("case_id"), str)
    }
    audits = []
    for draft in drafts:
        if draft.unavailable_audit is not None:
            audits.append(_audit_payload(draft.unavailable_audit))
        elif draft.case_id is not None and draft.case_id in audits_by_case_id:
            audits.append(audits_by_case_id[draft.case_id])
    return audits


def _audit_field_payload(audit: dict[str, object]) -> dict[str, object]:
    return {field.name: audit[field.name] for field in fields(ModelCallAudit)}


def _ordered_unresolved_case_ids(
    drafts: list[_RegionDraft],
    disambiguator_case_ids: tuple[str, ...],
) -> tuple[str, ...]:
    unresolved_case_ids = set(disambiguator_case_ids)
    unresolved_case_ids.update(
        draft.case_id
        for draft in drafts
        if draft.case_id is not None and draft.unavailable_audit is not None
    )
    ordered = [
        draft.case_id
        for draft in drafts
        if draft.case_id is not None and draft.case_id in unresolved_case_ids
    ]
    ordered.extend(case_id for case_id in disambiguator_case_ids if case_id not in ordered)
    return tuple(ordered)


def _deterministic_materialized_region(
    snapshot_source: str,
    draft: _RegionDraft,
) -> _MaterializedRegion:
    deterministic_block = _materialize_deterministic(snapshot_source, draft)
    return _MaterializedRegion(
        draft=draft,
        block=deterministic_block,
        deterministic_block=deterministic_block,
        audit=draft.unavailable_audit,
    )


def _materialize_region(snapshot_source, draft: _RegionDraft, resolutions_by_case_id):
    deterministic_region = _deterministic_materialized_region(snapshot_source, draft)
    deterministic_block = deterministic_region.deterministic_block
    if draft.case is None:
        return deterministic_region

    resolution = resolutions_by_case_id[draft.case.case_id]
    if resolution.status == "local_fallback":
        return _MaterializedRegion(
            draft=draft,
            block=deterministic_block,
            deterministic_block=deterministic_block,
            audit=_normalize_fallback_audit(resolution.audit),
        )

    choice = next(
        choice for choice in draft.case.choices if choice.choice_id == resolution.choice_id
    )
    classification = BlockClassification(
        kind=choice.kind,
        confidence=choice.local_score,
        reason_codes=[*choice.reason_codes, "model_selected_choice"],
        features=draft.assessment.deterministic.features,
    )
    assert resolution.audit is not None
    try:
        block = _block_for_candidate(
            snapshot_source,
            draft.sheet,
            draft.candidate,
            classification,
        )
    except Exception as exc:
        block = _unclassified_block(
            snapshot_source,
            draft.candidate,
            confidence=0.0,
            reason_codes=["semantic_block_fallback"],
            extra_diagnostic={"error_type": type(exc).__name__},
        )
        audit = replace(
            resolution.audit,
            outcome="materialization_error",
            validation_codes=_stable_codes(
                *resolution.audit.validation_codes,
                "materialization_error",
            ),
            reason_codes=("semantic_block_fallback",),
            error_type=type(exc).__name__,
        )
        return _MaterializedRegion(
            draft=draft,
            block=block,
            deterministic_block=deterministic_block,
            audit=audit,
            materialization_failed=True,
        )
    return _MaterializedRegion(
        draft=draft,
        block=block,
        deterministic_block=deterministic_block,
        audit=resolution.audit,
        model_selected=True,
    )


def _materialize_deterministic(snapshot_source, draft: _RegionDraft) -> WorkbookBlock:
    try:
        return _block_for_candidate(
            snapshot_source,
            draft.sheet,
            draft.candidate,
            draft.assessment.deterministic,
        )
    except Exception as exc:
        return _unclassified_block(
            snapshot_source,
            draft.candidate,
            confidence=0.0,
            reason_codes=["semantic_block_fallback"],
            extra_diagnostic={"error_type": type(exc).__name__},
        )


def _normalize_fallback_audit(audit: ModelCallAudit | None) -> ModelCallAudit | None:
    if audit is None:
        return None
    outcome = (
        "provider_error"
        if audit.outcome in {"adapter_error", "deadline_exceeded", "timeout"}
        else audit.outcome
    )
    return replace(
        audit,
        outcome=outcome,
        reason_codes=("deterministic_fallback",),
    )


def _workbook_from_materialized(
    snapshot: WorkbookSnapshot,
    materialized: list[_MaterializedRegion],
    *,
    rollback_selected: bool,
) -> WorkbookIR:
    workbook_ir, _ = assemble_baseline(snapshot)
    blocks_by_sheet: dict[int, list[WorkbookBlock]] = {
        index: [] for index in range(len(snapshot.sheets))
    }
    for region in materialized:
        block = (
            region.deterministic_block
            if rollback_selected and region.model_selected
            else region.block
        )
        blocks_by_sheet[region.draft.sheet_index].append(deepcopy(block))
    for sheet_index, sheet_ir in enumerate(workbook_ir.sheets):
        sheet_ir.blocks = blocks_by_sheet[sheet_index]
    return workbook_ir


def _finalize_workbook(
    snapshot: WorkbookSnapshot,
    workbook_ir: WorkbookIR,
) -> tuple[ParseDiagnostics, tuple[str, ...]]:
    _, diagnostics = assemble_baseline(snapshot)
    block_counts: Counter[str] = Counter()
    ambiguous_regions = []
    for sheet_ir in workbook_ir.sheets:
        for block in sheet_ir.blocks:
            if block.kind == "unclassified":
                reason_codes = [
                    diagnostic["reason_code"]
                    for diagnostic in block.diagnostics
                    if "reason_code" in diagnostic
                ]
                ambiguous_regions.append(
                    {
                        "sheet_name": sheet_ir.name,
                        "range": block.source_refs[0].range,
                        "candidate_kind": "unclassified",
                        "confidence": block.confidence,
                        "reason_codes": reason_codes,
                    }
                )
            block_counts[block.kind] += 1

    diagnostics.block_count_by_kind = dict(sorted(block_counts.items()))
    diagnostics.ambiguous_regions = ambiguous_regions
    continuation_failed = False
    try:
        groups, candidates = link_table_continuations(snapshot, workbook_ir)
    except Exception as exc:
        continuation_failed = True
        diagnostics.warnings.append(f"cross_sheet_continuation_fallback:{type(exc).__name__}")
    else:
        workbook_ir.table_continuations = groups
        diagnostics.continuation_candidates = candidates
        ambiguous_count = sum(item["status"] == "ambiguous" for item in candidates)
        if ambiguous_count:
            diagnostics.warnings.append(
                f"Workbook contains {ambiguous_count} ambiguous continuation candidates"
            )
    _update_coverage(snapshot, workbook_ir, diagnostics)
    row_conservation_passed = _row_conservation_passed(workbook_ir)
    if not row_conservation_passed:
        diagnostics.status = "partial"
        diagnostics.warnings.append("Workbook IR failed logical row conservation")
    validity_ratio, invalid_refs = validate_workbook_source_refs(snapshot, workbook_ir)
    diagnostics.source_ref_validity_ratio = validity_ratio
    if invalid_refs:
        diagnostics.status = "partial"
        diagnostics.warnings.append(
            f"Workbook IR contains {len(invalid_refs)} invalid source refs: {invalid_refs[:10]}"
        )
    validation_codes = []
    if diagnostics.coverage_ratio != 1.0:
        validation_codes.append("invalid_coverage")
    if not diagnostics.reconstruction_passed:
        validation_codes.append("reconstruction_failed")
    if not row_conservation_passed:
        validation_codes.append("row_conservation_failed")
    if diagnostics.source_ref_validity_ratio != 1.0 or invalid_refs:
        validation_codes.append("invalid_source_refs")
    if continuation_failed:
        validation_codes.append("continuation_error")
    return diagnostics, tuple(validation_codes)


def _row_conservation_passed(workbook_ir: WorkbookIR) -> bool:
    for sheet_ir in workbook_ir.sheets:
        for block in sheet_ir.blocks:
            if block.logical_table is None or not block.source_refs:
                continue
            min_col, min_row, max_col, max_row = range_boundaries(block.source_refs[0].range)
            expected = [
                (sheet_ir.name, min_col, row_number, max_col, row_number)
                for row_number in range(min_row, max_row + 1)
            ]
            actual = []
            for row in block.logical_table.rows:
                row_min_col, row_min, row_max_col, row_max = range_boundaries(row.source_ref.range)
                actual.append(
                    (
                        row.source_ref.sheet_name,
                        row_min_col,
                        row_min,
                        row_max_col,
                        row_max,
                    )
                )
            if actual != expected:
                return False
    return True


def _stable_codes(*codes: str) -> tuple[str, ...]:
    return tuple(dict.fromkeys(codes))


def _block_for_candidate(snapshot_source, sheet, candidate, classification):
    common = {
        "block_id": stable_id(
            "block", snapshot_source, candidate.source_ref.key, classification.kind
        ),
        "kind": classification.kind,
        "source_refs": [candidate.source_ref],
        "cell_refs": candidate.cell_refs,
        "confidence": classification.confidence,
        "metadata": {
            "view": "raw_grid"
            if classification.kind == "unclassified"
            else f"semantic_{classification.kind}",
            "cell_count": len(candidate.cell_refs),
            "features": asdict(classification.features),
            "reason_codes": list(classification.reason_codes),
        },
        "diagnostics": [
            {"reason_code": reason_code} for reason_code in classification.reason_codes
        ],
    }
    if classification.kind == "logical_table":
        table = interpret_logical_table(sheet, candidate)
        common["confidence"] = min(classification.confidence, table.confidence)
        return WorkbookBlock(**common, logical_table=table)
    if classification.kind == "form":
        return WorkbookBlock(
            **common,
            form=interpret_form_block(sheet, candidate, classification),
        )
    if classification.kind == "matrix":
        return WorkbookBlock(
            **common,
            matrix=interpret_matrix_block(sheet, candidate, classification),
        )
    if classification.kind == "text":
        return WorkbookBlock(
            **common,
            text=interpret_text_block(sheet, candidate, classification),
        )
    return WorkbookBlock(**common)


def _unclassified_block(
    snapshot_source,
    candidate,
    *,
    confidence,
    reason_codes,
    extra_diagnostic=None,
):
    diagnostics = [{"reason_code": reason_code} for reason_code in reason_codes]
    if extra_diagnostic:
        diagnostics[0].update(extra_diagnostic)
    return WorkbookBlock(
        block_id=stable_id("block", snapshot_source, candidate.source_ref.key, "unclassified"),
        kind="unclassified",
        source_refs=[candidate.source_ref],
        cell_refs=candidate.cell_refs,
        confidence=confidence,
        metadata={
            "view": "raw_grid",
            "cell_count": len(candidate.cell_refs),
            "reason_codes": list(reason_codes),
        },
        diagnostics=diagnostics,
    )


def validate_workbook_source_refs(
    snapshot: WorkbookSnapshot,
    workbook_ir: WorkbookIR,
) -> tuple[float, list[str]]:
    """Return the ratio of derived refs bounded by an existing source Sheet."""

    sheet_bounds = {
        sheet.name: range_boundaries(sheet.used_range or _range_for_coordinates(list(sheet.cells)))
        for sheet in snapshot.sheets
        if sheet.used_range or sheet.cells
    }
    refs = []
    for sheet_ir in workbook_ir.sheets:
        for block in sheet_ir.blocks:
            refs.extend(block.source_refs)
            if block.logical_table is not None:
                refs.extend(_logical_table_source_refs(block.logical_table))
            if block.form is not None:
                refs.extend(block.form.source_refs)
                refs.extend(ref for field in block.form.fields for ref in field.label_source_refs)
                refs.extend(ref for field in block.form.fields for ref in field.value_source_refs)
                refs.extend(ref for line in block.form.free_text for ref in line.source_refs)
            if block.matrix is not None:
                refs.extend(block.matrix.source_refs)
                refs.extend(
                    ref for header in block.matrix.row_headers for ref in header.source_refs
                )
                refs.extend(
                    ref for header in block.matrix.column_headers for ref in header.source_refs
                )
                refs.extend(
                    ref for row in block.matrix.value_source_refs for ref in row if ref is not None
                )
            if block.text is not None:
                refs.extend(block.text.source_refs)
                refs.extend(ref for line in block.text.lines for ref in line.source_refs)

    for continuation in workbook_ir.table_continuations:
        refs.extend(_logical_table_source_refs(continuation.logical_table))
        refs.extend(continuation.source_refs)

    invalid = sorted({ref.key for ref in refs if not _valid_source_ref(ref, sheet_bounds)})
    invalid_count = sum(not _valid_source_ref(ref, sheet_bounds) for ref in refs)
    ratio = (len(refs) - invalid_count) / len(refs) if refs else 1.0
    return ratio, invalid


def _logical_table_source_refs(table: LogicalTable) -> list[SourceRef]:
    return [
        *table.source_refs,
        *(ref for column in table.columns for ref in column.source_refs),
        *(row.source_ref for row in table.rows),
        *(fragment.source_ref for fragment in table.fragments),
        *(section.source_ref for section in table.sections),
    ]


def _valid_source_ref(ref: SourceRef, sheet_bounds) -> bool:
    bounds = sheet_bounds.get(ref.sheet_name)
    if bounds is None:
        return False
    min_col, min_row, max_col, max_row = range_boundaries(ref.range)
    sheet_min_col, sheet_min_row, sheet_max_col, sheet_max_row = bounds
    return (
        sheet_min_col <= min_col <= max_col <= sheet_max_col
        and sheet_min_row <= min_row <= max_row <= sheet_max_row
    )


def _update_coverage(snapshot, workbook_ir, diagnostics):
    source_refs = {
        f"{sheet.name}!{coordinate}"
        for sheet in snapshot.sheets
        for coordinate, cell in sheet.cells.items()
        if _is_assignable_cell(cell)
    }
    assigned_refs = {
        f"{sheet_ir.name}!{coordinate}"
        for sheet_ir in workbook_ir.sheets
        for block in sheet_ir.blocks
        for coordinate in block.cell_refs
    }
    diagnostics.coverage_ratio = (
        len(source_refs & assigned_refs) / len(source_refs) if source_refs else 1.0
    )
    diagnostics.reconstruction_passed = source_refs == assigned_refs
    if not diagnostics.reconstruction_passed:
        diagnostics.status = "partial"


def assemble_baseline(snapshot: WorkbookSnapshot) -> tuple[WorkbookIR, ParseDiagnostics]:
    """Create a lossless raw-grid IR before any semantic table interpretation."""

    sheet_irs: list[SheetIR] = []
    source_cell_keys: set[str] = set()
    assigned_cell_keys: set[str] = set()
    block_counts: Counter[str] = Counter()

    for sheet in snapshot.sheets:
        cell_refs = sorted(
            (coordinate for coordinate, cell in sheet.cells.items() if _is_assignable_cell(cell)),
            key=coordinate_to_tuple,
        )
        blocks: list[WorkbookBlock] = []
        if cell_refs:
            source_range = sheet.used_range or _range_for_coordinates(cell_refs)
            source_ref = SourceRef(sheet_name=sheet.name, range=source_range)
            block = WorkbookBlock(
                block_id=stable_id("block", snapshot.source, source_ref.key, "unclassified"),
                kind="unclassified",
                source_refs=[source_ref],
                cell_refs=cell_refs,
                metadata={"view": "raw_grid", "cell_count": len(cell_refs)},
            )
            blocks.append(block)
            block_counts[block.kind] += 1

        sheet_irs.append(
            SheetIR(
                sheet_id=stable_id("sheet", snapshot.source, str(sheet.index), sheet.name),
                name=sheet.name,
                index=sheet.index,
                blocks=blocks,
                visibility=sheet.visibility,
                metadata={
                    "used_range": sheet.used_range,
                    "print_area": sheet.print_area,
                    "merged_ranges": sheet.merged_ranges,
                    "object_count": len(sheet.objects),
                },
            )
        )

        qualified_refs = {f"{sheet.name}!{coordinate}" for coordinate in cell_refs}
        source_cell_keys.update(qualified_refs)
        assigned_cell_keys.update(qualified_refs if blocks else set())

    reconstruction_passed = assigned_cell_keys == source_cell_keys
    coverage_ratio = (
        len(assigned_cell_keys & source_cell_keys) / len(source_cell_keys)
        if source_cell_keys
        else 1.0
    )
    warnings = list(snapshot.metadata.get("warnings", []))
    if not reconstruction_passed:
        missing = sorted(source_cell_keys - assigned_cell_keys)
        warnings.append(f"Workbook IR omitted {len(missing)} source cells: {missing[:10]}")

    diagnostics = ParseDiagnostics(
        status="success" if reconstruction_passed and coverage_ratio == 1.0 else "partial",
        coverage_ratio=coverage_ratio,
        reconstruction_passed=reconstruction_passed,
        block_count_by_kind=dict(sorted(block_counts.items())),
        unsupported_features=list(snapshot.metadata.get("unsupported_features", [])),
        warnings=warnings,
    )
    workbook_ir = WorkbookIR(
        kind="workbook",
        workbook_id=stable_id("workbook", snapshot.source, snapshot.filename),
        source=snapshot.source,
        sheets=sheet_irs,
        filename=snapshot.filename,
        snapshot=snapshot,
        metadata={"snapshot": snapshot.metadata},
    )
    return workbook_ir, diagnostics


def _is_assignable_cell(cell: CellSnapshot) -> bool:
    return any(
        (
            cell.raw_value is not None,
            cell.formula is not None,
            cell.comment is not None,
            cell.hyperlink is not None,
            cell.merge_anchor is not None,
        )
    )


def _range_for_coordinates(coordinates: list[str]) -> str:
    positions = [coordinate_to_tuple(coordinate) for coordinate in coordinates]
    rows = [row for row, _ in positions]
    columns = [column for _, column in positions]
    return (
        f"{get_column_letter(min(columns))}{min(rows)}:{get_column_letter(max(columns))}{max(rows)}"
    )
