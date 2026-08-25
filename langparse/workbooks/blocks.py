from __future__ import annotations

from openpyxl.utils import get_column_letter, range_boundaries

from langparse.workbooks.classification import BlockClassification
from langparse.workbooks.types import (
    CandidateRegion,
    FormBlock,
    FormField,
    MatrixBlock,
    MatrixHeader,
    SheetSnapshot,
    SourceRef,
    TextBlock,
    TextLine,
    stable_id,
)


def interpret_form_block(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
    classification: BlockClassification,
) -> FormBlock:
    """Interpret adjacent label/value pairs without inventing missing fields."""

    min_col, min_row, max_col, max_row = range_boundaries(candidate.source_ref.range)
    title = ""
    fields: list[FormField] = []
    free_text: list[TextLine] = []
    for row_number in range(min_row, max_row + 1):
        entries = _row_entries(sheet, row_number, min_col, max_col)
        if row_number == min_row and len(entries) == 1:
            title = entries[0][1]
            continue
        pairs = _adjacent_pairs(entries)
        if pairs is not None:
            for (label_coordinate, label), (value_coordinate, value) in pairs:
                label_ref = SourceRef(sheet_name=sheet.name, range=label_coordinate)
                value_ref = SourceRef(sheet_name=sheet.name, range=value_coordinate)
                fields.append(
                    FormField(
                        field_id=stable_id(
                            "field",
                            candidate.source_ref.key,
                            label_ref.key,
                            value_ref.key,
                        ),
                        label=label,
                        value=value,
                        label_source_refs=[label_ref],
                        value_source_refs=[value_ref],
                        confidence=classification.confidence,
                    )
                )
            continue
        if entries:
            free_text.append(
                TextLine(
                    text=" ".join(value for _, value in entries),
                    source_refs=[
                        SourceRef(sheet_name=sheet.name, range=coordinate)
                        for coordinate, _ in entries
                    ],
                )
            )
    return FormBlock(
        form_id=stable_id("form", candidate.source_ref.key),
        title=title,
        fields=fields,
        free_text=free_text,
        source_refs=[candidate.source_ref],
        confidence=classification.confidence,
        diagnostics=[{"reason_codes": list(classification.reason_codes)}],
    )


def interpret_matrix_block(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
    classification: BlockClassification,
) -> MatrixBlock:
    """Preserve a two-axis matrix and its physical value grid."""

    min_col, min_row, max_col, max_row = range_boundaries(candidate.source_ref.range)
    if max_col - min_col < 2 or max_row - min_row < 2:
        raise ValueError("matrix requires at least two row and column dimensions")

    title = _display_value(sheet, min_row, min_col)
    column_headers = [
        MatrixHeader(
            value=_display_value(sheet, min_row, column),
            source_refs=[_source_ref(sheet, min_row, column)],
        )
        for column in range(min_col + 1, max_col + 1)
    ]
    row_headers = [
        MatrixHeader(
            value=_display_value(sheet, row, min_col),
            source_refs=[_source_ref(sheet, row, min_col)],
        )
        for row in range(min_row + 1, max_row + 1)
    ]
    if any(not header.value for header in [*column_headers, *row_headers]):
        raise ValueError("matrix axes must be complete")

    values = []
    value_source_refs = []
    for row in range(min_row + 1, max_row + 1):
        value_row = []
        ref_row = []
        for column in range(min_col + 1, max_col + 1):
            value_row.append(_display_value(sheet, row, column))
            coordinate = f"{get_column_letter(column)}{row}"
            ref_row.append(_source_ref(sheet, row, column) if coordinate in sheet.cells else None)
        values.append(value_row)
        value_source_refs.append(ref_row)

    return MatrixBlock(
        matrix_id=stable_id("matrix", candidate.source_ref.key),
        title=title,
        row_headers=row_headers,
        column_headers=column_headers,
        values=values,
        source_refs=[candidate.source_ref],
        value_source_refs=value_source_refs,
        confidence=classification.confidence,
        diagnostics=[{"reason_codes": list(classification.reason_codes)}],
    )


def interpret_text_block(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
    classification: BlockClassification,
) -> TextBlock:
    """Render source-ordered display text without merged-cell duplication."""

    min_col, min_row, max_col, max_row = range_boundaries(candidate.source_ref.range)
    lines = []
    for row_number in range(min_row, max_row + 1):
        entries = _row_entries(sheet, row_number, min_col, max_col)
        if entries:
            lines.append(
                TextLine(
                    text=" ".join(value for _, value in entries),
                    source_refs=[
                        SourceRef(sheet_name=sheet.name, range=coordinate)
                        for coordinate, _ in entries
                    ],
                )
            )
    return TextBlock(
        text_id=stable_id("text", candidate.source_ref.key),
        lines=lines,
        source_refs=[candidate.source_ref],
        confidence=classification.confidence,
        diagnostics=[{"reason_codes": list(classification.reason_codes)}],
    )


def _row_entries(
    sheet: SheetSnapshot,
    row_number: int,
    min_col: int,
    max_col: int,
) -> list[tuple[str, str]]:
    entries = []
    for column in range(min_col, max_col + 1):
        coordinate = f"{get_column_letter(column)}{row_number}"
        cell = sheet.cells.get(coordinate)
        if cell is None or cell.merge_anchor is not None or not cell.display_value.strip():
            continue
        entries.append((coordinate, cell.display_value))
    return entries


def _adjacent_pairs(
    entries: list[tuple[str, str]],
) -> list[tuple[tuple[str, str], tuple[str, str]]] | None:
    if len(entries) < 2 or len(entries) % 2:
        return None
    pairs = []
    for offset in range(0, len(entries), 2):
        label, value = entries[offset : offset + 2]
        label_column = _column_number(label[0])
        value_column = _column_number(value[0])
        if value_column != label_column + 1:
            return None
        pairs.append((label, value))
    return pairs


def _column_number(coordinate: str) -> int:
    from openpyxl.utils.cell import coordinate_to_tuple

    return coordinate_to_tuple(coordinate)[1]


def _display_value(sheet: SheetSnapshot, row: int, column: int) -> str:
    coordinate = f"{get_column_letter(column)}{row}"
    cell = sheet.cells.get(coordinate)
    if cell is None or cell.merge_anchor is not None:
        return ""
    return cell.display_value


def _source_ref(sheet: SheetSnapshot, row: int, column: int) -> SourceRef:
    return SourceRef(sheet_name=sheet.name, range=f"{get_column_letter(column)}{row}")
