from __future__ import annotations

import re
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple

from langparse.workbooks.types import (
    CandidateRegion,
    CellSnapshot,
    RegionAnchor,
    SheetSnapshot,
    SourceRef,
)

_FORMULA_CELL_REF = re.compile(r"(?<![A-Z0-9_!])\$?([A-Z]{1,3})\$?([1-9][0-9]*)(?![A-Z0-9_(])")
_REASON_ORDER = {
    "native_table_anchor": 0,
    "defined_name_anchor": 1,
    "print_area_anchor": 2,
    "merged_title_anchor": 3,
    "formula_continuity": 4,
    "style_boundary": 5,
    "density_boundary": 6,
    "blank_band": 7,
    "occupied_extent": 8,
}
_EXACT_ANCHOR_REASONS = {
    "native_table_anchor",
    "defined_name_anchor",
    "print_area_anchor",
}


@dataclass(frozen=True)
class _Rect:
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def area(self) -> int:
        return (self.max_column - self.min_column + 1) * (self.max_row - self.min_row + 1)

    @property
    def range(self) -> str:
        return (
            f"{get_column_letter(self.min_column)}{self.min_row}:"
            f"{get_column_letter(self.max_column)}{self.max_row}"
        )


@dataclass(frozen=True)
class _UsableAnchor:
    anchor: RegionAnchor
    rect: _Rect


@dataclass(frozen=True)
class _Cut:
    orientation: str
    boundary: int
    reason: str


def detect_candidate_regions(sheet: SheetSnapshot) -> list[CandidateRegion]:
    """Partition assignable cells using stable structural evidence.

    The public interface intentionally remains a single deterministic function.
    Native anchors, print areas, visual discontinuities, merged ranges and formula
    references are implementation details hidden behind that seam.
    """

    occupied = {coordinate: cell for coordinate, cell in sheet.cells.items() if _is_occupied(cell)}
    if not occupied:
        return []

    positions = {coordinate: coordinate_to_tuple(coordinate) for coordinate in occupied}
    coarse_regions = _initial_coarse_regions(sheet, positions)
    blank_partitioned = len(coarse_regions) > 1
    regions: list[CandidateRegion] = []
    for coarse in coarse_regions:
        regions.extend(
            _partition_coarse_region(
                sheet,
                occupied,
                positions,
                coarse,
                blank_partitioned=blank_partitioned,
            )
        )

    regions.sort(
        key=lambda region: (
            coordinate_to_tuple(region.source_ref.range.split(":", 1)[0]),
            region.source_ref.range,
        )
    )
    assigned = [coordinate for region in regions for coordinate in region.cell_refs]
    if Counter(assigned) != Counter(occupied.keys()):
        raise RuntimeError("Candidate region partition violated cell ownership")
    return regions


def _initial_coarse_regions(
    sheet: SheetSnapshot,
    positions: dict[str, tuple[int, int]],
) -> list[_Rect]:
    coarse_regions = []
    for min_row, max_row in _consecutive_groups(row for row, _ in positions.values()):
        columns = {column for row, column in positions.values() if min_row <= row <= max_row}
        coarse_regions.extend(
            _Rect(min_column, min_row, max_column, max_row)
            for min_column, max_column in _consecutive_groups(columns)
        )

    for anchor in sheet.region_anchors:
        if anchor.kind != "excel_table" or anchor.source_ref.sheet_name != sheet.name:
            continue
        try:
            anchor_rect = _rect_from_range(anchor.source_ref.range)
        except ValueError:
            continue
        if not any(_contains(anchor_rect, *position) for position in positions.values()):
            continue
        overlapping = [rect for rect in coarse_regions if _rectangles_overlap(rect, anchor_rect)]
        if not overlapping:
            continue
        coarse_regions = [
            rect for rect in coarse_regions if not _rectangles_overlap(rect, anchor_rect)
        ]
        coarse_regions.append(_bounding_rect([anchor_rect, *overlapping]))
        coarse_regions = _merge_overlapping_rects(coarse_regions)

    return sorted(
        coarse_regions,
        key=lambda rect: (rect.min_row, rect.min_column, rect.max_row, rect.max_column),
    )


def _bounding_rect(rectangles: list[_Rect]) -> _Rect:
    return _Rect(
        min(rect.min_column for rect in rectangles),
        min(rect.min_row for rect in rectangles),
        max(rect.max_column for rect in rectangles),
        max(rect.max_row for rect in rectangles),
    )


def _merge_overlapping_rects(rectangles: list[_Rect]) -> list[_Rect]:
    pending = list(rectangles)
    merged: list[_Rect] = []
    while pending:
        current = pending.pop()
        overlaps = [rect for rect in pending if _rectangles_overlap(current, rect)]
        if overlaps:
            pending = [rect for rect in pending if rect not in overlaps]
            pending.append(_bounding_rect([current, *overlaps]))
        else:
            merged.append(current)
    return merged


def _partition_coarse_region(
    sheet: SheetSnapshot,
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    coarse: _Rect,
    *,
    blank_partitioned: bool,
) -> list[CandidateRegion]:
    selected_anchors, conflicts = _select_anchors(sheet, positions, coarse)
    print_rects = _print_area_rects(sheet, coarse)
    protected_rects = [selected.rect for selected in selected_anchors]
    use_print_rects = len(print_rects) >= 2 and _pairwise_non_overlapping(print_rects)
    if use_print_rects:
        protected_rects.extend(print_rects)
    return _partition_rect(
        sheet,
        occupied,
        positions,
        coarse,
        selected_anchors,
        conflicts,
        print_rects,
        use_print_rects=use_print_rects,
        protected_rects=protected_rects,
        partition_reasons=frozenset(),
        blank_partitioned=blank_partitioned,
    )


def _partition_rect(
    sheet: SheetSnapshot,
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    selected_anchors: list[_UsableAnchor],
    conflicts: list[tuple[_UsableAnchor, _UsableAnchor]],
    print_rects: list[_Rect],
    *,
    use_print_rects: bool,
    protected_rects: list[_Rect],
    partition_reasons: frozenset[str],
    blank_partitioned: bool,
) -> list[CandidateRegion]:
    cell_refs = sorted(
        (coordinate for coordinate, position in positions.items() if _contains(rect, *position)),
        key=coordinate_to_tuple,
    )
    if not cell_refs:
        return []

    cuts = _candidate_cuts(
        sheet,
        occupied,
        positions,
        rect,
        selected_anchors,
        print_rects,
        use_print_rects=use_print_rects,
        protected_rects=protected_rects,
    )
    if cuts:
        cut = cuts[0]
        first, second = _split_rect(rect, cut)
        inherited_reasons = partition_reasons
        if cut.reason not in _EXACT_ANCHOR_REASONS:
            inherited_reasons = frozenset((*partition_reasons, cut.reason))
        return [
            *_partition_rect(
                sheet,
                occupied,
                positions,
                first,
                selected_anchors,
                conflicts,
                print_rects,
                use_print_rects=use_print_rects,
                protected_rects=protected_rects,
                partition_reasons=inherited_reasons,
                blank_partitioned=blank_partitioned,
            ),
            *_partition_rect(
                sheet,
                occupied,
                positions,
                second,
                selected_anchors,
                conflicts,
                print_rects,
                use_print_rects=use_print_rects,
                protected_rects=protected_rects,
                partition_reasons=inherited_reasons,
                blank_partitioned=blank_partitioned,
            ),
        ]

    source_rect = _anchored_source_rect(rect, cell_refs, positions, selected_anchors)
    reasons = _region_reasons(
        sheet,
        occupied,
        positions,
        source_rect,
        selected_anchors,
        print_rects,
        partition_reasons=partition_reasons,
        blank_partitioned=blank_partitioned,
    )
    diagnostics = [
        {
            "reason_code": "overlapping_native_anchors",
            "kept_kind": kept.anchor.kind,
            "kept_range": kept.rect.range,
            "kept_name": kept.anchor.name,
            "kept_scope": kept.anchor.scope,
            "rejected_kind": rejected.anchor.kind,
            "rejected_range": rejected.rect.range,
            "rejected_name": rejected.anchor.name,
            "rejected_scope": rejected.anchor.scope,
        }
        for kept, rejected in conflicts
        if _rectangles_overlap(source_rect, kept.rect)
        or _rectangles_overlap(source_rect, rejected.rect)
    ]
    return [
        CandidateRegion(
            source_ref=SourceRef(sheet_name=sheet.name, range=source_rect.range),
            cell_refs=cell_refs,
            confidence=_region_confidence(reasons, diagnostics),
            features={
                "row_count": source_rect.max_row - source_rect.min_row + 1,
                "column_count": source_rect.max_column - source_rect.min_column + 1,
                "occupied_count": len(cell_refs),
                "density": len(cell_refs) / source_rect.area,
            },
            diagnostics=diagnostics,
            reason_codes=reasons,
        )
    ]


def _candidate_cuts(
    sheet: SheetSnapshot,
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    selected_anchors: list[_UsableAnchor],
    print_rects: list[_Rect],
    *,
    use_print_rects: bool,
    protected_rects: list[_Rect],
) -> list[_Cut]:
    cuts: dict[tuple[str, int], _Cut] = {}
    evidence = [
        (selected.rect, _anchor_reason(selected.anchor.kind))
        for selected in selected_anchors
        if _rectangles_overlap(selected.rect, rect)
    ]
    if use_print_rects:
        evidence.extend(
            (print_rect, "print_area_anchor")
            for print_rect in print_rects
            if _rectangles_overlap(print_rect, rect)
        )
    for evidence_rect, reason in evidence:
        for cut in _rect_edge_cuts(evidence_rect, rect, reason):
            if _cut_is_safe(sheet, rect, cut, protected_rects):
                _offer_cut(cuts, cut)

    for boundary in range(rect.min_column, rect.max_column):
        key = ("vertical", boundary)
        if key in cuts:
            continue
        cut = _Cut("vertical", boundary, "style_boundary")
        if not _cut_is_safe(sheet, rect, cut, protected_rects):
            continue
        if _formula_crosses_column(occupied, positions, rect, boundary):
            continue
        if _is_style_boundary(sheet, rect, boundary):
            cuts[key] = cut
        elif _is_density_boundary(occupied, positions, rect, boundary):
            cuts[key] = _Cut("vertical", boundary, "density_boundary")

    for boundary in range(rect.min_row, rect.max_row):
        key = ("horizontal", boundary)
        if key in cuts:
            continue
        cut = _Cut("horizontal", boundary, "style_boundary")
        if not _cut_is_safe(sheet, rect, cut, protected_rects):
            continue
        if _formula_crosses_row(occupied, positions, rect, boundary):
            continue
        if _is_row_style_boundary(sheet, rect, boundary):
            cuts[key] = cut
        elif _is_row_density_boundary(occupied, positions, rect, boundary):
            cuts[key] = _Cut("horizontal", boundary, "density_boundary")

    return sorted(
        cuts.values(),
        key=lambda cut: (
            _REASON_ORDER.get(cut.reason, 99),
            0 if cut.orientation == "vertical" else 1,
            cut.boundary,
        ),
    )


def _rect_edge_cuts(evidence: _Rect, rect: _Rect, reason: str) -> list[_Cut]:
    cuts = []
    if rect.min_column < evidence.min_column <= rect.max_column:
        cuts.append(_Cut("vertical", evidence.min_column - 1, reason))
    if rect.min_column <= evidence.max_column < rect.max_column:
        cuts.append(_Cut("vertical", evidence.max_column, reason))
    if rect.min_row < evidence.min_row <= rect.max_row:
        cuts.append(_Cut("horizontal", evidence.min_row - 1, reason))
    if rect.min_row <= evidence.max_row < rect.max_row:
        cuts.append(_Cut("horizontal", evidence.max_row, reason))
    return cuts


def _offer_cut(cuts: dict[tuple[str, int], _Cut], candidate: _Cut) -> None:
    key = (candidate.orientation, candidate.boundary)
    current = cuts.get(key)
    if current is None or _REASON_ORDER.get(candidate.reason, 99) < _REASON_ORDER.get(
        current.reason,
        99,
    ):
        cuts[key] = candidate


def _cut_is_safe(
    sheet: SheetSnapshot,
    rect: _Rect,
    cut: _Cut,
    protected_rects: list[_Rect],
) -> bool:
    if cut.orientation == "vertical":
        if _merged_range_crosses_column(sheet, rect, cut.boundary):
            return False
        return not any(
            _rectangles_overlap(protected, rect)
            and protected.min_column <= cut.boundary < protected.max_column
            for protected in protected_rects
        )
    if _merged_range_crosses_row(sheet, rect, cut.boundary):
        return False
    return not any(
        _rectangles_overlap(protected, rect)
        and protected.min_row <= cut.boundary < protected.max_row
        for protected in protected_rects
    )


def _split_rect(rect: _Rect, cut: _Cut) -> tuple[_Rect, _Rect]:
    if cut.orientation == "vertical":
        return (
            _Rect(rect.min_column, rect.min_row, cut.boundary, rect.max_row),
            _Rect(cut.boundary + 1, rect.min_row, rect.max_column, rect.max_row),
        )
    return (
        _Rect(rect.min_column, rect.min_row, rect.max_column, cut.boundary),
        _Rect(rect.min_column, cut.boundary + 1, rect.max_column, rect.max_row),
    )


def _select_anchors(
    sheet: SheetSnapshot,
    positions: dict[str, tuple[int, int]],
    coarse: _Rect,
) -> tuple[list[_UsableAnchor], list[tuple[_UsableAnchor, _UsableAnchor]]]:
    usable = []
    for anchor in sheet.region_anchors:
        if anchor.source_ref.sheet_name != sheet.name:
            continue
        try:
            rect = _rect_from_range(anchor.source_ref.range)
        except ValueError:
            continue
        if not _rectangles_overlap(rect, coarse) or not _rect_inside(rect, coarse):
            continue
        count = sum(_contains(rect, row, column) for row, column in positions.values())
        if not count:
            continue
        if anchor.kind == "defined_name" and (
            rect.max_column == rect.min_column
            or rect.max_row == rect.min_row
            or count / rect.area < 0.5
        ):
            continue
        if anchor.kind not in {"excel_table", "defined_name"}:
            continue
        usable.append(_UsableAnchor(anchor, rect))

    usable.sort(
        key=lambda item: (
            {"excel_table": 0, "defined_name": 1}.get(item.anchor.kind, 99),
            -item.rect.area,
            item.rect.min_row,
            item.rect.min_column,
            item.anchor.name or "",
        )
    )
    selected: list[_UsableAnchor] = []
    conflicts: list[tuple[_UsableAnchor, _UsableAnchor]] = []
    for candidate in usable:
        conflicting = next(
            (item for item in selected if _rectangles_overlap(item.rect, candidate.rect)),
            None,
        )
        if conflicting is not None:
            if conflicting.rect == candidate.rect:
                continue
            conflicts.append((conflicting, candidate))
            continue
        selected.append(candidate)
    return selected, conflicts


def _print_area_rects(sheet: SheetSnapshot, coarse: _Rect) -> list[_Rect]:
    rectangles = []
    for value in sheet.print_area:
        local_range = value.rsplit("!", 1)[-1].replace("$", "")
        try:
            rect = _rect_from_range(local_range)
        except ValueError:
            continue
        if _rect_inside(rect, coarse):
            rectangles.append(rect)
    return sorted(
        set(rectangles),
        key=lambda item: (item.min_row, item.min_column, item.max_row, item.max_column),
    )


def _anchored_source_rect(
    segment_rect: _Rect,
    cell_refs: list[str],
    positions: dict[str, tuple[int, int]],
    anchors: list[_UsableAnchor],
) -> _Rect:
    for selected in anchors:
        anchor_cells = {
            coordinate
            for coordinate, position in positions.items()
            if _contains(selected.rect, *position)
        }
        if anchor_cells == set(cell_refs) and _rect_inside(selected.rect, segment_rect):
            return selected.rect
    return segment_rect


def _region_reasons(
    sheet: SheetSnapshot,
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    anchors: list[_UsableAnchor],
    print_rects: list[_Rect],
    *,
    partition_reasons: frozenset[str],
    blank_partitioned: bool,
) -> list[str]:
    reasons = set(partition_reasons)
    for selected in anchors:
        if selected.rect == rect:
            reasons.add(_anchor_reason(selected.anchor.kind))
    if rect in print_rects:
        reasons.add("print_area_anchor")
    if any(
        cell.formula is not None and _local_formula_refs(cell.formula)
        for coordinate, cell in occupied.items()
        if _contains(rect, *positions[coordinate])
    ):
        reasons.add("formula_continuity")
    if any(_rectangles_overlap(rect, merged) for merged in _merged_rects(sheet)):
        reasons.add("merged_title_anchor")
    if blank_partitioned:
        reasons.add("blank_band")
    if not reasons or reasons == {"formula_continuity"}:
        reasons.add("occupied_extent")
    return sorted(reasons, key=lambda item: (_REASON_ORDER.get(item, 99), item))


def _region_confidence(reasons: list[str], diagnostics: list[dict[str, Any]]) -> float:
    if "native_table_anchor" in reasons:
        confidence = 0.98
    elif "print_area_anchor" in reasons or "defined_name_anchor" in reasons:
        confidence = 0.95
    elif "style_boundary" in reasons:
        confidence = 0.9
    elif "density_boundary" in reasons:
        confidence = 0.82
    else:
        confidence = 1.0
    if diagnostics:
        confidence = min(confidence, 0.75)
    return confidence


def _is_style_boundary(sheet: SheetSnapshot, rect: _Rect, boundary: int) -> bool:
    if boundary - rect.min_column + 1 < 2 or rect.max_column - boundary < 2:
        return False
    comparisons = []
    for row in range(rect.min_row, rect.max_row + 1):
        left = sheet.cells.get(f"{get_column_letter(boundary)}{row}")
        right = sheet.cells.get(f"{get_column_letter(boundary + 1)}{row}")
        if left is not None and right is not None and _is_occupied(left) and _is_occupied(right):
            comparisons.append(
                (left.visual_style_id or left.style_id) != (right.visual_style_id or right.style_id)
            )
    return len(comparisons) >= 2 and sum(comparisons) / len(comparisons) >= 0.8


def _is_row_style_boundary(sheet: SheetSnapshot, rect: _Rect, boundary: int) -> bool:
    if boundary - rect.min_row + 1 < 2 or rect.max_row - boundary < 2:
        return False
    top_rect = _Rect(rect.min_column, rect.min_row, rect.max_column, boundary)
    bottom_rect = _Rect(rect.min_column, boundary + 1, rect.max_column, rect.max_row)
    if not (
        _looks_like_complete_table(sheet, top_rect)
        and _looks_like_complete_table(sheet, bottom_rect)
    ):
        return False
    if not (
        _rows_have_stable_visual_style(sheet, rect, boundary - 1, boundary)
        and _rows_have_stable_visual_style(sheet, rect, boundary + 1, boundary + 2)
    ):
        return False
    comparisons = []
    for column in range(rect.min_column, rect.max_column + 1):
        top = sheet.cells.get(f"{get_column_letter(column)}{boundary}")
        bottom = sheet.cells.get(f"{get_column_letter(column)}{boundary + 1}")
        if top is not None and bottom is not None and _is_occupied(top) and _is_occupied(bottom):
            comparisons.append(
                (top.visual_style_id or top.style_id) != (bottom.visual_style_id or bottom.style_id)
            )
    return len(comparisons) >= 2 and sum(comparisons) / len(comparisons) >= 0.8


def _looks_like_complete_table(sheet: SheetSnapshot, rect: _Rect) -> bool:
    header_cells = [
        cell
        for column in range(rect.min_column, rect.max_column + 1)
        if (cell := sheet.cells.get(f"{get_column_letter(column)}{rect.min_row}")) is not None
        and _is_occupied(cell)
        and cell.merge_anchor is None
    ]
    if len(header_cells) < 2 or any(
        cell.formula is not None
        or not isinstance(cell.raw_value, str)
        or not cell.display_value.strip()
        for cell in header_cells
    ):
        return False
    # Text-only records are valid table bodies too. Require a populated record
    # under the prospective header, rather than requiring numeric values.
    return any(
        all(
            (cell := sheet.cells.get(f"{get_column_letter(column)}{row}")) is not None
            and _is_occupied(cell)
            and cell.merge_anchor is None
            for column in range(rect.min_column, rect.max_column + 1)
        )
        for row in range(rect.min_row + 1, rect.max_row + 1)
    )


def _rows_have_stable_visual_style(
    sheet: SheetSnapshot,
    rect: _Rect,
    first_row: int,
    second_row: int,
) -> bool:
    comparisons = []
    for column in range(rect.min_column, rect.max_column + 1):
        first = sheet.cells.get(f"{get_column_letter(column)}{first_row}")
        second = sheet.cells.get(f"{get_column_letter(column)}{second_row}")
        if (
            first is not None
            and second is not None
            and _is_occupied(first)
            and _is_occupied(second)
        ):
            comparisons.append(
                (first.visual_style_id or first.style_id)
                == (second.visual_style_id or second.style_id)
            )
    return len(comparisons) >= 2 and sum(comparisons) / len(comparisons) >= 0.8


def _is_density_boundary(
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    boundary: int,
) -> bool:
    left = _Rect(rect.min_column, rect.min_row, boundary, rect.max_row)
    right = _Rect(boundary + 1, rect.min_row, rect.max_column, rect.max_row)
    left_width = left.max_column - left.min_column + 1
    right_width = right.max_column - right.min_column + 1
    if min(left_width, right_width) != 1 or max(left_width, right_width) < 2:
        return False
    left_cells = [
        occupied[coordinate]
        for coordinate, position in positions.items()
        if _contains(left, *position)
    ]
    right_cells = [
        occupied[coordinate]
        for coordinate, position in positions.items()
        if _contains(right, *position)
    ]
    left_density = len(left_cells) / left.area
    right_density = len(right_cells) / right.area
    dense, sparse = (
        (left_density, right_cells)
        if left_density >= right_density
        else (right_density, left_cells)
    )
    sparse_density = min(left_density, right_density)
    return (
        dense >= 0.75
        and sparse_density <= 0.5
        and any(len(cell.display_value.strip()) >= 20 for cell in sparse)
    )


def _is_row_density_boundary(
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    boundary: int,
) -> bool:
    top = _Rect(rect.min_column, rect.min_row, rect.max_column, boundary)
    bottom = _Rect(rect.min_column, boundary + 1, rect.max_column, rect.max_row)
    top_height = top.max_row - top.min_row + 1
    bottom_height = bottom.max_row - bottom.min_row + 1
    if min(top_height, bottom_height) != 1 or max(top_height, bottom_height) < 2:
        return False
    top_cells = [
        occupied[coordinate]
        for coordinate, position in positions.items()
        if _contains(top, *position)
    ]
    bottom_cells = [
        occupied[coordinate]
        for coordinate, position in positions.items()
        if _contains(bottom, *position)
    ]
    top_density = len(top_cells) / top.area
    bottom_density = len(bottom_cells) / bottom.area
    dense, sparse = (
        (top_density, bottom_cells)
        if top_density >= bottom_density
        else (bottom_density, top_cells)
    )
    sparse_density = min(top_density, bottom_density)
    return (
        dense >= 0.75
        and sparse_density <= 0.5
        and any(len(cell.display_value.strip()) >= 20 for cell in sparse)
    )


def _formula_crosses_column(
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    boundary: int,
) -> bool:
    for coordinate, cell in occupied.items():
        row, formula_column = positions[coordinate]
        if not _contains(rect, row, formula_column) or cell.formula is None:
            continue
        for reference_column, reference_row in _local_formula_refs(cell.formula):
            if not _contains(rect, reference_row, reference_column):
                continue
            if (formula_column <= boundary < reference_column) or (
                reference_column <= boundary < formula_column
            ):
                return True
    return False


def _formula_crosses_row(
    occupied: dict[str, CellSnapshot],
    positions: dict[str, tuple[int, int]],
    rect: _Rect,
    boundary: int,
) -> bool:
    for coordinate, cell in occupied.items():
        formula_row, column = positions[coordinate]
        if not _contains(rect, formula_row, column) or cell.formula is None:
            continue
        for reference_column, reference_row in _local_formula_refs(cell.formula):
            if not _contains(rect, reference_row, reference_column):
                continue
            if (formula_row <= boundary < reference_row) or (
                reference_row <= boundary < formula_row
            ):
                return True
    return False


def _local_formula_refs(formula: str) -> list[tuple[int, int]]:
    try:
        range_tokens = (
            token.value
            for token in Tokenizer(formula).items
            if token.type == "OPERAND" and token.subtype == "RANGE" and "!" not in token.value
        )
    except Exception:
        return []
    return [
        (column_index_from_string(column_name), int(row_number))
        for value in range_tokens
        for column_name, row_number in _FORMULA_CELL_REF.findall(value.upper())
    ]


def _merged_range_crosses_column(sheet: SheetSnapshot, rect: _Rect, boundary: int) -> bool:
    return any(
        merged.min_column <= boundary < merged.max_column
        and not (merged.max_row < rect.min_row or merged.min_row > rect.max_row)
        for merged in _merged_rects(sheet)
    )


def _merged_range_crosses_row(sheet: SheetSnapshot, rect: _Rect, boundary: int) -> bool:
    return any(
        merged.min_row <= boundary < merged.max_row
        and not (merged.max_column < rect.min_column or merged.min_column > rect.max_column)
        for merged in _merged_rects(sheet)
    )


def _merged_rects(sheet: SheetSnapshot) -> list[_Rect]:
    rectangles = []
    for value in sheet.merged_ranges:
        try:
            rectangles.append(_rect_from_range(value.replace("$", "")))
        except ValueError:
            continue
    return rectangles


def _anchor_reason(kind: str) -> str:
    return "native_table_anchor" if kind == "excel_table" else "defined_name_anchor"


def _rect_from_range(value: str) -> _Rect:
    boundaries = range_boundaries(value)
    if any(boundary is None for boundary in boundaries):
        raise ValueError("region ranges must have finite row and column bounds")
    min_column, min_row, max_column, max_row = boundaries
    return _Rect(min_column, min_row, max_column, max_row)


def _contains(rect: _Rect, row: int, column: int) -> bool:
    return rect.min_row <= row <= rect.max_row and rect.min_column <= column <= rect.max_column


def _rect_inside(inner: _Rect, outer: _Rect) -> bool:
    return (
        outer.min_column <= inner.min_column <= inner.max_column <= outer.max_column
        and outer.min_row <= inner.min_row <= inner.max_row <= outer.max_row
    )


def _rectangles_overlap(left: _Rect, right: _Rect) -> bool:
    return not (
        left.max_column < right.min_column
        or right.max_column < left.min_column
        or left.max_row < right.min_row
        or right.max_row < left.min_row
    )


def _pairwise_non_overlapping(rectangles: list[_Rect]) -> bool:
    return all(
        not _rectangles_overlap(left, right)
        for index, left in enumerate(rectangles)
        for right in rectangles[index + 1 :]
    )


def _is_occupied(cell: CellSnapshot) -> bool:
    return any(
        (
            cell.raw_value is not None,
            cell.formula is not None,
            cell.comment is not None,
            cell.hyperlink is not None,
            cell.merge_anchor is not None,
        )
    )


def _consecutive_groups(values: Iterable[int]) -> list[tuple[int, int]]:
    ordered = sorted(set(values))
    if not ordered:
        return []
    groups: list[tuple[int, int]] = []
    start = previous = ordered[0]
    for value in ordered[1:]:
        if value != previous + 1:
            groups.append((start, previous))
            start = value
        previous = value
    groups.append((start, previous))
    return groups
