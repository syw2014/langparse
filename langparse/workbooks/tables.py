from __future__ import annotations

import re

from openpyxl.utils import get_column_letter, range_boundaries

from langparse.workbooks.types import (
    CandidateRegion,
    HeaderColumn,
    LogicalTable,
    SheetSnapshot,
    SourceRef,
    TableFragment,
    stable_id,
)

PAGE_RE = re.compile(r"第\s*(\d+)\s*页\s*共\s*(\d+)\s*页")


def interpret_logical_table(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
) -> LogicalTable:
    """Interpret deterministic print fragments and their shared header."""

    min_col, min_row, max_col, max_row = range_boundaries(candidate.source_ref.range)
    page_markers = _page_markers(sheet, min_row, max_row, min_col, max_col)
    fragments = _build_fragments(
        sheet,
        candidate,
        page_markers,
        min_col,
        min_row,
        max_col,
        max_row,
    )
    header_rows = fragments[0].header_row_numbers if fragments else []
    columns = _build_header_columns(
        sheet,
        candidate.source_ref,
        header_rows,
        min_col,
        max_col,
    )
    title_row = fragments[0].title_row_numbers[0] if fragments[0].title_row_numbers else min_row
    title = _first_text(sheet, title_row, min_col, max_col)
    context = [
        text
        for row_number in fragments[0].context_row_numbers
        for text in [_row_text(sheet, row_number, min_col, max_col)]
        if text
    ]
    return LogicalTable(
        table_id=stable_id("table", candidate.source_ref.key),
        title=title,
        context=context,
        columns=columns,
        fragments=fragments,
        source_refs=[candidate.source_ref],
    )


def _page_markers(
    sheet: SheetSnapshot,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[tuple[int, int, int]]:
    markers = []
    for row_number in range(min_row, max_row + 1):
        match = PAGE_RE.search(_row_text(sheet, row_number, min_col, max_col))
        if match:
            markers.append((row_number, int(match.group(1)), int(match.group(2))))
    return markers


def _build_fragments(
    sheet: SheetSnapshot,
    candidate: CandidateRegion,
    page_markers: list[tuple[int, int, int]],
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
) -> list[TableFragment]:
    if not _valid_page_sequence(page_markers):
        header_rows = _header_rows_after(sheet, min_row - 1, min_row, max_row, min_col, max_col)
        return [
            TableFragment(
                fragment_id=stable_id("fragment", candidate.source_ref.key),
                source_ref=candidate.source_ref,
                title_row_numbers=[min_row] if header_rows and min_row < header_rows[0] else [],
                header_row_numbers=header_rows,
                diagnostics=[{"reason_code": "no_consistent_print_sequence"}],
            )
        ]

    starts = [max(min_row, row_number - 1) for row_number, _, _ in page_markers]
    fragments: list[TableFragment] = []
    header_fingerprints: list[tuple[tuple[str, ...], ...]] = []
    for index, ((context_row, page_number, total_pages), start_row) in enumerate(
        zip(page_markers, starts, strict=True)
    ):
        end_row = starts[index + 1] - 1 if index + 1 < len(starts) else max_row
        header_rows = _header_rows_after(
            sheet,
            context_row,
            start_row,
            end_row,
            min_col,
            max_col,
        )
        header_fingerprints.append(
            tuple(
                tuple(
                    _cell_text(sheet, row_number, column) for column in range(min_col, max_col + 1)
                )
                for row_number in header_rows
            )
        )
        source_range = (
            f"{get_column_letter(min_col)}{start_row}:{get_column_letter(max_col)}{end_row}"
        )
        fragments.append(
            TableFragment(
                fragment_id=stable_id("fragment", candidate.source_ref.key, str(page_number)),
                source_ref=SourceRef(sheet_name=sheet.name, range=source_range),
                page_number=page_number,
                total_pages=total_pages,
                title_row_numbers=[start_row] if start_row < context_row else [],
                context_row_numbers=[context_row],
                header_row_numbers=header_rows,
            )
        )

    if len(set(header_fingerprints)) != 1:
        diagnostic = {"reason_code": "header_fingerprint_mismatch"}
        for fragment in fragments:
            fragment.confidence = 0.5
            fragment.diagnostics.append(diagnostic)
    return fragments


def _valid_page_sequence(markers: list[tuple[int, int, int]]) -> bool:
    if len(markers) < 2:
        return False
    pages = [page for _, page, _ in markers]
    totals = {total for _, _, total in markers}
    return len(totals) == 1 and pages == list(range(pages[0], pages[0] + len(pages)))


def _header_rows_after(
    sheet: SheetSnapshot,
    context_row: int,
    start_row: int,
    end_row: int,
    min_col: int,
    max_col: int,
) -> list[int]:
    rows = []
    for row_number in range(max(start_row, context_row + 1), end_row + 1):
        if _looks_like_data_or_section(sheet, row_number, min_col, max_col):
            break
        if _row_text(sheet, row_number, min_col, max_col):
            rows.append(row_number)
    return rows


def _looks_like_data_or_section(
    sheet: SheetSnapshot,
    row_number: int,
    min_col: int,
    max_col: int,
) -> bool:
    first = _cell_text(sheet, row_number, min_col).strip()
    if re.fullmatch(r"\d+", first):
        return True
    row_text = _row_text(sheet, row_number, min_col, max_col).replace(" ", "")
    return "合计" in row_text or "总计" in row_text


def _build_header_columns(
    sheet: SheetSnapshot,
    candidate_ref: SourceRef,
    header_rows: list[int],
    min_col: int,
    max_col: int,
) -> list[HeaderColumn]:
    columns = []
    for column in range(min_col, max_col + 1):
        coordinate = get_column_letter(column)
        path = []
        source_refs = []
        for row_number in header_rows:
            text = _cell_text(sheet, row_number, column).strip()
            if text and (not path or path[-1] != text):
                path.append(text)
            source_refs.append(SourceRef(sheet_name=sheet.name, range=f"{coordinate}{row_number}"))
        columns.append(
            HeaderColumn(
                column_id=stable_id("column", candidate_ref.key, coordinate),
                coordinate=coordinate,
                path=path,
                source_refs=source_refs,
            )
        )
    return columns


def _cell_text(sheet: SheetSnapshot, row_number: int, column: int) -> str:
    coordinate = f"{get_column_letter(column)}{row_number}"
    cell = sheet.cells.get(coordinate)
    if cell is None:
        return ""
    if cell.merge_anchor:
        anchor = sheet.cells.get(cell.merge_anchor)
        return anchor.display_value if anchor is not None else ""
    return cell.display_value


def _row_text(sheet: SheetSnapshot, row_number: int, min_col: int, max_col: int) -> str:
    return " | ".join(
        text
        for column in range(min_col, max_col + 1)
        for text in [_cell_text(sheet, row_number, column).strip()]
        if text
    )


def _first_text(sheet: SheetSnapshot, row_number: int, min_col: int, max_col: int) -> str:
    for column in range(min_col, max_col + 1):
        text = _cell_text(sheet, row_number, column).strip()
        if text:
            return text
    return ""
