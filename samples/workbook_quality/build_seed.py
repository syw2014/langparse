"""Build the public, synthetic workbook-quality tuning seed.

The fixtures encode deliberately small business structures. Truth is written
explicitly below; it is not inferred from LangParse output.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent
FIXTURES = ROOT / "fixtures"
FIXED_TIME = datetime(2026, 1, 1, 0, 0, 0)
FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)


def _block(
    source_range: str,
    kind: str,
    *,
    headers: list[dict] | None = None,
    rows: list[dict] | None = None,
    form_fields: list[dict] | None = None,
    matrix_rows: list[str] | None = None,
    matrix_columns: list[str] | None = None,
) -> dict:
    return {
        "source_range": source_range,
        "kind": kind,
        "headers": headers or [],
        "rows": rows or [],
        "form_fields": form_fields or [],
        "matrix_axes": {
            "rows": matrix_rows or [],
            "columns": matrix_columns or [],
        },
    }


def _row(source_range: str, role: str) -> dict:
    return {"source_range": source_range, "role": role}


def _header(coordinate: str, *path: str) -> dict:
    return {"coordinate": coordinate, "path": list(path)}


def _expectation(
    sheets: list[dict],
    *,
    continuations: list[list[str]] | None = None,
    required_source_refs: list[str] | None = None,
    objects: list[dict] | None = None,
) -> dict:
    return {
        "sheets": sheets,
        "continuations": continuations or [],
        "required_source_refs": required_source_refs or [],
        "objects": objects or [],
    }


def _sheet(name: str, *blocks: dict) -> dict:
    return {"name": name, "blocks": list(blocks)}


def _save(sample_id: str, filename: str, workbook: Workbook, expectation: dict) -> dict:
    path = FIXTURES / filename
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.save(path)
    _normalize_xlsx_archive(path)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "sample_id": sample_id,
        "path": filename,
        "sha256": f"sha256:{digest}",
        "expectation": expectation,
    }


def _normalize_xlsx_archive(path: Path) -> None:
    temporary = path.with_suffix(".normalized.xlsx")
    with ZipFile(path, "r") as source, ZipFile(temporary, "w", ZIP_DEFLATED) as target:
        for source_info in sorted(source.infolist(), key=lambda item: item.filename):
            target_info = ZipInfo(source_info.filename, date_time=FIXED_ZIP_TIME)
            target_info.compress_type = ZIP_DEFLATED
            target_info.external_attr = source_info.external_attr
            target_info.create_system = source_info.create_system
            payload = source.read(source_info.filename)
            if source_info.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                    payload,
                )
            target.writestr(target_info, payload)
    temporary.replace(path)


def _simple_table() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    sheet.append(["Beta", 2])
    expected = _expectation(
        [
            _sheet(
                "Data",
                _block(
                    "A1:B3",
                    "logical_table",
                    headers=[_header("A", "Name"), _header("B", "Value")],
                    rows=[
                        _row("A1:B1", "header"),
                        _row("A2:B2", "data"),
                        _row("A3:B3", "data"),
                    ],
                ),
            )
        ],
        required_source_refs=["Data!A1:B3"],
    )
    return workbook, expected


def _two_tables() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for coordinate, value in {
        "A1": "Name",
        "B1": "Value",
        "A2": "Alpha",
        "B2": 1,
        "A5": "Code",
        "B5": "Amount",
        "A6": "B-1",
        "B6": 20,
    }.items():
        sheet[coordinate] = value
    expected = _expectation(
        [
            _sheet(
                "Data",
                _block(
                    "A1:B2",
                    "logical_table",
                    headers=[_header("A", "Name"), _header("B", "Value")],
                    rows=[_row("A1:B1", "header"), _row("A2:B2", "data")],
                ),
                _block(
                    "A5:B6",
                    "logical_table",
                    headers=[_header("A", "Code"), _header("B", "Amount")],
                    rows=[_row("A5:B5", "header"), _row("A6:B6", "data")],
                ),
            )
        ],
        required_source_refs=["Data!A1:B2", "Data!A5:B6"],
    )
    return workbook, expected


def _form() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Form"
    sheet.append(["项目登记"])
    sheet.append(["项目名称", "道路工程"])
    sheet.append(["建设单位", "示例公司"])
    expected = _expectation(
        [
            _sheet(
                "Form",
                _block(
                    "A1:B3",
                    "form",
                    form_fields=[
                        {"label": "项目名称", "value": "道路工程"},
                        {"label": "建设单位", "value": "示例公司"},
                    ],
                ),
            )
        ],
        required_source_refs=["Form!A1:B3"],
    )
    return workbook, expected


def _matrix() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet.append([None, "Q1", "Q2"])
    sheet.append(["North", 1, 2])
    sheet.append(["South", 3, 4])
    expected = _expectation(
        [
            _sheet(
                "Matrix",
                _block(
                    "A1:C3",
                    "matrix",
                    matrix_rows=["North", "South"],
                    matrix_columns=["Q1", "Q2"],
                ),
            )
        ],
        required_source_refs=["Matrix!A1:C3"],
    )
    return workbook, expected


def _text() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Overview"
    sheet["A2"] = "First note"
    sheet["A3"] = "Second note"
    expected = _expectation(
        [_sheet("Notes", _block("A1:A3", "text"))],
        required_source_refs=["Notes!A1:A3"],
    )
    return workbook, expected


def _sparse() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sparse"
    sheet["A1"] = "Top left"
    sheet["B2"] = "Bottom right"
    expected = _expectation(
        [_sheet("Sparse", _block("A1:B2", "unclassified"))],
        required_source_refs=["Sparse!A1:B2"],
    )
    return workbook, expected


def _continuation() -> tuple[Workbook, dict]:
    workbook = Workbook()
    expected_sheets = []
    refs = []
    for index, item in enumerate(("Alpha", "Beta"), start=1):
        sheet = workbook.active if index == 1 else workbook.create_sheet()
        sheet.title = f"清单{index}"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 10
        sheet.append(["工程清单"])
        sheet.append([f"第 {index} 页 共 2 页 ({index}/2)"])
        sheet.append(["Name", "Code", "Value"])
        sheet.append([item, f"Item {index}", index])
        source = f"清单{index}!A1:C4"
        refs.append(source)
        expected_sheets.append(
            _sheet(
                sheet.title,
                _block(
                    "A1:C4",
                    "logical_table",
                    headers=[
                        _header("A", "Name"),
                        _header("B", "Code"),
                        _header("C", "Value"),
                    ],
                    rows=[
                        _row("A1:C1", "title"),
                        _row("A2:C2", "context"),
                        _row("A3:C3", "header"),
                        _row("A4:C4", "data"),
                    ],
                ),
            )
        )
    return workbook, _expectation(
        expected_sheets,
        continuations=[refs],
        required_source_refs=refs,
    )


def _repeated_fragments() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget"
    header_1 = [
        "序号",
        "项目编码",
        "项目名称",
        "项目特征描述",
        "计量单位",
        "工程量",
        "综合单价(元)",
        "合价(元)",
        "其中",
        None,
        None,
        "备注",
    ]
    header_2 = [None, None, None, None, None, None, None, None, "人工费", "机械费", "管理费", None]
    for offset, page in ((0, 1), (6, 2)):
        rows = [
            ["表1-2 清单"],
            [
                "单位工程名称",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"第 {page} 页 共 2 页",
            ],
            header_1,
            header_2,
            [1 + offset, "code", "item"],
            [2 + offset, "code", "item"],
        ]
        for row_index, values in enumerate(rows, start=1 + offset):
            for column_index, value in enumerate(values, start=1):
                if value is not None:
                    sheet.cell(row=row_index, column=column_index, value=value)
        for column in range(1, 9):
            sheet.merge_cells(
                start_row=3 + offset,
                start_column=column,
                end_row=4 + offset,
                end_column=column,
            )
        sheet.merge_cells(
            start_row=3 + offset,
            start_column=9,
            end_row=3 + offset,
            end_column=11,
        )
        sheet.merge_cells(
            start_row=3 + offset,
            start_column=12,
            end_row=4 + offset,
            end_column=12,
        )
    headers = [
        _header("A", "序号"),
        _header("B", "项目编码"),
        _header("C", "项目名称"),
        _header("D", "项目特征描述"),
        _header("E", "计量单位"),
        _header("F", "工程量"),
        _header("G", "综合单价(元)"),
        _header("H", "合价(元)"),
        _header("I", "其中", "人工费"),
        _header("J", "其中", "机械费"),
        _header("K", "其中", "管理费"),
        _header("L", "备注"),
    ]
    roles = [
        "title",
        "context",
        "header",
        "header",
        "data",
        "data",
        "repeated_title",
        "repeated_context",
        "repeated_header",
        "repeated_header",
        "data",
        "data",
    ]
    expected = _expectation(
        [
            _sheet(
                "Budget",
                _block(
                    "A1:L12",
                    "logical_table",
                    headers=headers,
                    rows=[_row(f"A{row}:L{row}", role) for row, role in enumerate(roles, 1)],
                ),
            )
        ],
        required_source_refs=["Budget!A1:L12"],
    )
    return workbook, expected


def _chart() -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ChartData"
    sheet.append(["Category", "Sales"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])
    chart = BarChart()
    chart.title = "Sales by category"
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")
    expected = _expectation(
        [
            _sheet(
                "ChartData",
                _block(
                    "A1:B3",
                    "logical_table",
                    headers=[_header("A", "Category"), _header("B", "Sales")],
                    rows=[
                        _row("A1:B1", "header"),
                        _row("A2:B2", "data"),
                        _row("A3:B3", "data"),
                    ],
                ),
            )
        ],
        required_source_refs=["ChartData!A1:B3"],
        objects=[{"sheet_name": "ChartData", "kind": "chart", "anchor": "D2"}],
    )
    return workbook, expected


def _formula_and_hidden_sheet() -> tuple[Workbook, dict]:
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["Item", "Rate", "Amount"])
    data.append(["Alpha", 2, "=B2*2"])
    data.append(["Beta", 3, "=B3*2"])
    workbook.defined_names.add(DefinedName("RateInputs", attr_text="'Data'!$B$2:$B$3"))
    lookup = workbook.create_sheet("Lookup")
    lookup.append(["Name", "Value"])
    lookup.append(["Base", 2])
    lookup.sheet_state = "hidden"
    expected = _expectation(
        [
            _sheet(
                "Data",
                _block(
                    "A1:C3",
                    "logical_table",
                    headers=[
                        _header("A", "Item"),
                        _header("B", "Rate"),
                        _header("C", "Amount"),
                    ],
                    rows=[
                        _row("A1:C1", "header"),
                        _row("A2:C2", "data"),
                        _row("A3:C3", "data"),
                    ],
                ),
            ),
            _sheet(
                "Lookup",
                _block(
                    "A1:B2",
                    "logical_table",
                    headers=[_header("A", "Name"), _header("B", "Value")],
                    rows=[_row("A1:B1", "header"), _row("A2:B2", "data")],
                ),
            ),
        ],
        required_source_refs=["Data!A1:C3", "Lookup!A1:B2"],
    )
    return workbook, expected


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    builders = [
        ("simple-table", "simple_table.xlsx", _simple_table),
        ("two-tables", "two_tables.xlsx", _two_tables),
        ("form", "form.xlsx", _form),
        ("matrix", "matrix.xlsx", _matrix),
        ("text", "text.xlsx", _text),
        ("sparse-unclassified", "sparse_unclassified.xlsx", _sparse),
        ("cross-sheet-continuation", "continuation.xlsx", _continuation),
        ("repeated-print-fragments", "repeated_fragments.xlsx", _repeated_fragments),
        ("chart-facts", "chart.xlsx", _chart),
        ("formula-hidden-sheet", "formula_hidden.xlsx", _formula_and_hidden_sheet),
    ]
    samples = []
    for sample_id, filename, builder in builders:
        workbook, expectation = builder()
        samples.append(_save(sample_id, filename, workbook, expectation))

    manifest = {
        "schema_version": 1,
        "dataset_id": "workbook-quality-seed",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "quality_gate": {
            "minimum": {
                "block_precision": 1.0,
                "block_recall": 1.0,
                "header_path_accuracy": 1.0,
                "row_role_f1": 1.0,
                "form_field_exact_match": 1.0,
                "matrix_axis_accuracy": 1.0,
                "continuation_precision": 1.0,
                "continuation_recall": 1.0,
                "source_ref_completeness": 1.0,
                "source_ref_validity_ratio": 1.0,
                "cell_coverage_ratio": 1.0,
                "object_fact_precision": 1.0,
                "object_fact_recall": 1.0,
                "object_semantic_recall": 0.0,
            },
            "maximum": {"fallback_rate": 0.1},
        },
        "samples": samples,
    }
    (ROOT / "public-manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
