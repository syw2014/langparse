from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from langparse.services.workbook_ambiguity_benchmark import (
    WorkbookAmbiguityBenchmarkService,
    _EvaluationCaptureAdapter,
)
from langparse.workbooks.adapters import OOXMLWorkbookAdapter
from langparse.workbooks.assembly import assemble_workbook
from langparse.workbooks.evaluation.schema import compute_choices_digest
from langparse.workbooks.modeling.policy import WorkbookDisambiguation
from langparse.workbooks.modeling.ports import WorkbookStructureModelAdapter
from langparse.workbooks.modeling.types import (
    ModelIdentity,
    ProviderReply,
    WorkbookModelRequest,
)


def _create_synthetic_ambiguous_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "LeftTop"
    ws["B2"] = "RightBottom"
    wb.save(path)


class _ChecksumCorruptedMockAdapter(WorkbookStructureModelAdapter):
    """Adapter that returns a choice with an invalid request_checksum."""

    def __init__(self, target_choice_id: str) -> None:
        self._identity = ModelIdentity(provider="openai", model="gpt-4o-mini", revision="1")
        self.target_choice_id = target_choice_id

    @property
    def identity(self) -> ModelIdentity:
        return self._identity

    def complete(self, request: WorkbookModelRequest, *, timeout_seconds: float) -> ProviderReply:
        body = json.dumps(
            {
                "schema_version": 1,
                "request_checksum": "sha256:corrupted_checksum_does_not_match",
                "decisions": [
                    {
                        "case_id": cid,
                        "status": "selected",
                        "choice_id": self.target_choice_id,
                        "confidence": 0.99,
                        "reason_codes": ["corrupted_checksum_attack"],
                    }
                    for cid in request.case_ids
                ],
            }
        ).encode("utf-8")
        return ProviderReply(
            body=body,
            provider_request_id="corrupt-req",
            usage={"prompt_tokens": 50, "completion_tokens": 20, "total_tokens": 70},
        )


def test_corrupted_response_is_rejected_in_evaluation(tmp_path: Path):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    sample_file = source_root / "ambiguous.xlsx"
    _create_synthetic_ambiguous_workbook(sample_file)

    import hashlib

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    adapter = _EvaluationCaptureAdapter()
    snapshot = OOXMLWorkbookAdapter().snapshot(sample_file)
    assemble_workbook(snapshot, disambiguation=WorkbookDisambiguation.auto(adapter))
    case0 = adapter.captured_cases[0]
    text_choice = next(c for c in case0.choices if c.kind == "text")

    manifest_data = {
        "schema_version": 1,
        "dataset_id": "strictness-test",
        "dataset_version": "1",
        "split": "holdout",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-01",
                "path": "ambiguous.xlsx",
                "sha256": file_hash(sample_file),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-01",
                        "sheet_name": case0.sheet_name,
                        "source_range": case0.source_range,
                        "expected": "text",
                        "fact_digest": case0.fact_digest,
                        "choices_digest": compute_choices_digest(case0.choices),
                    }
                ],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    out_dir = tmp_path / "eval_out"

    corrupt_adapter = _ChecksumCorruptedMockAdapter(target_choice_id=text_choice.choice_id)
    report = WorkbookAmbiguityBenchmarkService().run(
        manifest_path,
        output_dir=out_dir,
        adapter=corrupt_adapter,
    )

    # Must be recorded as failed, NOT fixed_error
    assert report.metrics.fixed_error_count == 0
    assert report.metrics.model_failed_count == 1
    assert report.metrics.model_correct_acceptance_count == 0
    assert report.summary["production_ready"] is False
    assert report.results[0].observation_status == "failed"
    assert report.results[0].transition == "failed"
