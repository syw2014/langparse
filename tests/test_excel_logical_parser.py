from pathlib import Path

import pytest
from openpyxl import Workbook

from langparse.chunkers.workbook import WorkbookStructuralChunker
from langparse.parsers.excel_parser import ExcelParser
from langparse.workbooks.assembly import assemble_workbook
from langparse.workbooks.types import CellSnapshot, SheetSnapshot, WorkbookSnapshot

PRIVATE_BUDGET_WORKBOOK = Path("/Users/jerryshi/Desktop/download/预算清单-gXF6T6B.xlsx")


def _snapshot(values: dict[str, object]) -> WorkbookSnapshot:
    return WorkbookSnapshot(
        source="book.xlsx",
        filename="book.xlsx",
        sheets=[
            SheetSnapshot(
                name="Data",
                index=0,
                cells={
                    coordinate: CellSnapshot(
                        coordinate=coordinate,
                        raw_value=value,
                        display_value=str(value),
                    )
                    for coordinate, value in values.items()
                },
            )
        ],
    )


def test_assembly_promotes_two_blank_band_regions_to_two_logical_tables():
    snapshot = _snapshot(
        {
            "A1": "Name",
            "B1": "Value",
            "A2": "Alpha",
            "B2": 1,
            "A5": "Name",
            "B5": "Value",
            "A6": "Beta",
            "B6": 2,
        }
    )

    ir, diagnostics = assemble_workbook(snapshot)

    assert [block.kind for block in ir.sheets[0].blocks] == ["logical_table", "logical_table"]
    assert diagnostics.coverage_ratio == 1.0
    assert diagnostics.reconstruction_passed is True
    assert diagnostics.block_count_by_kind == {"logical_table": 2}


def test_assembly_keeps_one_cell_candidate_unclassified():
    ir, diagnostics = assemble_workbook(_snapshot({"C3": "note"}))

    assert [block.kind for block in ir.sheets[0].blocks] == ["unclassified"]
    assert diagnostics.coverage_ratio == 1.0


def test_excel_parser_uses_semantic_workbook_assembly(tmp_path):
    path = tmp_path / "table.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)

    assert parsed.structure is not None
    block = parsed.structure.sheets[0].blocks[0]
    assert block.kind == "logical_table"
    assert block.logical_table is not None
    assert [row.role for row in block.logical_table.rows] == ["header", "data"]
    assert parsed.markdown_content.count("| Name | Value |") == 1


def test_excel_parser_links_cross_sheet_continuation_and_chunks_source_members(tmp_path):
    path = tmp_path / "continued-table.xlsx"
    workbook = Workbook()
    for index, item in enumerate(("Alpha", "Beta"), start=1):
        sheet = workbook.active if index == 1 else workbook.create_sheet()
        sheet.title = f"清单{index}"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 10
        sheet.append(["工程清单"])
        sheet.append([f"第 {index} 页 共 2 页 ({index}/2)"])
        sheet.append(["Name", "Code", "Value"])
        sheet.append([item, f"Item {index}", index])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)

    assert len(parsed.structure.table_continuations) == 1
    group = parsed.structure.table_continuations[0]
    assert [row.values[0] for row in group.logical_table.rows if row.role == "data"] == [
        "Alpha",
        "Beta",
    ]
    assert [item["status"] for item in parsed.diagnostics.continuation_candidates] == ["accepted"]
    chunks = WorkbookStructuralChunker().chunk(parsed)
    assert len(chunks) == 2
    assert {chunk.metadata["continuation_id"] for chunk in chunks} == {group.continuation_id}


@pytest.mark.skipif(
    not PRIVATE_BUDGET_WORKBOOK.exists(),
    reason="private budget workbook is not available",
)
def test_private_budget_workbook_sheet_8_acceptance():
    parsed = ExcelParser().parse_result(PRIVATE_BUDGET_WORKBOOK)

    assert parsed.structure is not None
    assert len(parsed.structure.sheets) == 15
    assert parsed.diagnostics.block_count_by_kind == {"logical_table": 14, "text": 1}
    assert len(parsed.structure.table_continuations) == 0
    assert not any(
        item["status"] == "accepted" for item in parsed.diagnostics.continuation_candidates
    )
    sheet = parsed.structure.sheets[7]
    assert sheet.name == "8.表1-2分部分项工程量清单与计价表(资格后审专用)"
    assert len(sheet.blocks) == 1
    block = sheet.blocks[0]
    assert block.kind == "logical_table"
    assert block.logical_table is not None
    table = block.logical_table

    assert len(table.fragments) == 6
    assert [fragment.page_number for fragment in table.fragments] == [1, 2, 3, 4, 5, 6]
    assert len(table.columns) == 12
    assert [column.path for column in table.columns[8:11]] == [
        ["其中", "人工费"],
        ["其中", "机械费"],
        ["其中", "管理费"],
    ]
    assert [section.title for section in table.sections] == ["土方", "管道部分"]

    data_rows = [row for row in table.rows if row.role == "data"]
    total_rows = [row for row in table.rows if row.role == "total"]
    assert [row.values[0] for row in data_rows] == [str(number) for number in range(1, 48)]
    assert len(total_rows) == 1
    assert parsed.diagnostics is not None
    assert parsed.diagnostics.coverage_ratio == 1.0
    assert parsed.diagnostics.reconstruction_passed is True
    assert parsed.diagnostics.source_ref_validity_ratio == 1.0
    assert parsed.structure.sheets[0].blocks[0].kind == "text"
    block_kinds = {
        block.kind for workbook_sheet in parsed.structure.sheets for block in workbook_sheet.blocks
    }
    assert block_kinds == {"logical_table", "text"}
    chunks = WorkbookStructuralChunker().chunk(parsed)
    expected_chunk_types = {
        {
            "logical_table": "table_rows",
            "form": "form_fields",
            "matrix": "matrix_rows",
            "text": "text_block",
            "unclassified": "raw_grid_rows",
        }[kind]
        for kind in block_kinds
    }
    assert {chunk.metadata["chunk_type"] for chunk in chunks} == expected_chunk_types
    assert len(chunks) == 39
    logical_row_ids = [
        row.row_id
        for workbook_sheet in parsed.structure.sheets
        for workbook_block in workbook_sheet.blocks
        if workbook_block.logical_table is not None
        for row in workbook_block.logical_table.rows
        if row.role in {"data", "total"}
    ]
    chunk_row_ids = [
        row_id
        for chunk in chunks
        if chunk.metadata["chunk_type"] == "table_rows"
        for row_id in chunk.metadata["row_ids"]
    ]
    assert len(logical_row_ids) == 228
    assert len(chunk_row_ids) == 228
    assert len(logical_row_ids) == len(set(logical_row_ids))
    assert len(chunk_row_ids) == len(set(chunk_row_ids))
    assert len(chunk_row_ids) == len(logical_row_ids)
    assert set(chunk_row_ids) == set(logical_row_ids)
    assert "Unnamed:" not in parsed.markdown_content
