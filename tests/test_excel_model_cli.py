from __future__ import annotations

import json
import os
from pathlib import Path
from unittest import mock

import pytest
from openpyxl import Workbook

from langparse.cli import build_parser, main
from langparse.workbooks.evaluation.schema import compute_choices_digest


def _create_synthetic_ambiguous_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "LeftTop"
    ws["B2"] = "RightBottom"
    wb.save(path)


def test_cli_does_not_accept_api_keys_in_process_arguments():
    with pytest.raises(SystemExit):
        build_parser().parse_args(
            ["parse", "book.xlsx", "--model", "gpt-4o", "--api-key", "sk-secret"]
        )


def test_root_help_describes_workbook_evaluation_command():
    help_text = build_parser().format_help()

    assert "benchmark-workbook-ambiguity" in help_text
    assert "evaluate workbook ambiguity" in help_text


def test_cli_parse_with_model_args(tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    sample_file = tmp_path / "test_sample.xlsx"
    _create_synthetic_ambiguous_workbook(sample_file)

    mock_client = mock.MagicMock()
    mock_choice = mock.MagicMock()
    mock_choice.message.content = json.dumps(
        {
            "schema_version": 1,
            "request_checksum": "dummy",
            "decisions": [
                {
                    "case_id": "dummy",
                    "status": "abstained",
                    "confidence": 0.0,
                    "reason_codes": ["mock_decision"],
                }
            ],
        }
    )
    mock_response = mock.MagicMock()
    mock_response.choices = [mock_choice]
    mock_response.id = "cmpl-mock"
    mock_response.usage = None
    mock_client.chat.completions.create.return_value = mock_response

    with (
        mock.patch.dict(os.environ, {"OPENAI_API_KEY": "sk-test-key"}),
        mock.patch("openai.OpenAI", return_value=mock_client),
    ):
        exit_code = main(
            [
                "parse",
                str(sample_file),
                "--model",
                "gpt-4o-mini",
            ]
        )
        assert exit_code == 0
        captured = capsys.readouterr()
        assert "LeftTop" in captured.out


def test_cli_eval_subcommand_with_model(tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    sample_file = source_root / "ambiguous.xlsx"
    _create_synthetic_ambiguous_workbook(sample_file)

    import hashlib

    from langparse.services.workbook_ambiguity_benchmark import _EvaluationCaptureAdapter
    from langparse.workbooks.adapters import OOXMLWorkbookAdapter
    from langparse.workbooks.assembly import assemble_workbook
    from langparse.workbooks.modeling.policy import WorkbookDisambiguation

    adapter = _EvaluationCaptureAdapter()
    snapshot = OOXMLWorkbookAdapter().snapshot(sample_file)
    assemble_workbook(snapshot, disambiguation=WorkbookDisambiguation.auto(adapter))
    case0 = adapter.captured_cases[0]

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    manifest_data = {
        "schema_version": 1,
        "dataset_id": "cli-eval-test",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-01",
                "path": "ambiguous.xlsx",
                "sha256": file_hash(sample_file),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-01",
                        "sheet_name": case0.sheet_name,
                        "source_range": case0.source_range,
                        "expected": "text",
                        "fact_digest": case0.fact_digest,
                        "choices_digest": compute_choices_digest(case0.choices),
                    }
                ],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    out_dir = tmp_path / "eval_reports"

    mock_client = mock.MagicMock()
    mock_choice = mock.MagicMock()
    text_choice = next(c for c in case0.choices if c.kind == "text")
    mock_choice.message.content = json.dumps(
        {
            "schema_version": 1,
            "request_checksum": "dummy",
            "decisions": [
                {
                    "case_id": case0.case_id,
                    "status": "selected",
                    "choice_id": text_choice.choice_id,
                    "confidence": 0.98,
                    "reason_codes": ["mock_fixed_error"],
                }
            ],
        }
    )
    mock_response = mock.MagicMock()
    mock_response.choices = [mock_choice]
    mock_response.id = "cmpl-eval-mock"
    mock_response.usage = None
    mock_client.chat.completions.create.return_value = mock_response

    with (
        mock.patch.dict(os.environ, {"OPENAI_API_KEY": "sk-test-key"}),
        mock.patch("openai.OpenAI", return_value=mock_client),
    ):
        exit_code = main(
            [
                "eval",
                str(manifest_path),
                "--output-dir",
                str(out_dir),
                "--model",
                "gpt-4o",
            ]
        )
        assert exit_code == 0
        captured = capsys.readouterr()
        assert "workbook ambiguity benchmark completed" in captured.out


def test_cli_missing_api_key_error_handling(tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    sample_file = tmp_path / "test.xlsx"
    _create_synthetic_ambiguous_workbook(sample_file)

    with mock.patch.dict(os.environ, {}, clear=True):
        exit_code = main(["parse", str(sample_file), "--model", "gpt-4o"])
        assert exit_code == 2
        captured = capsys.readouterr()
        assert "OPENAI_API_KEY is required" in captured.err
        assert "--api-key" not in captured.err
