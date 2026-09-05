from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from langparse.workbooks.adapters import OOXMLWorkbookAdapter


def test_ooxml_adapter_preserves_workbook_facts(tmp_path):
    path = tmp_path / "facts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Amount"
    sheet["A1"].font = Font(bold=True)
    sheet["A1"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["A1"].comment = Comment("Source value", "LangParse")
    sheet["A1"].hyperlink = "https://example.com/source"
    sheet["A2"] = 2
    sheet["B2"] = "=A2*2"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Merged"
    sheet.row_dimensions[4].hidden = True
    sheet.row_dimensions[2].height = 24
    sheet.column_dimensions["C"].hidden = True
    sheet.column_dimensions["A"].width = 18
    sheet.print_area = "A1:B3"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    chart = BarChart()
    chart.add_data(
        Reference(sheet, min_col=1, min_row=1, max_row=2),
        titles_from_data=True,
    )
    sheet.add_chart(chart, "D2")
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)
    data = snapshot.sheets[0]

    assert data.used_range == "A1:B3"
    assert data.cells["B2"].raw_value == "=A2*2"
    assert data.cells["B2"].formula == "=A2*2"
    assert data.cells["A1"].style_id
    assert data.cells["A1"].comment == "Source value"
    assert data.cells["A1"].hyperlink == "https://example.com/source"
    assert data.cells["A3"].rowspan == 1
    assert data.cells["A3"].colspan == 2
    assert data.cells["B3"].merge_anchor == "A3"
    assert data.merged_ranges == ["A3:B3"]
    assert data.row_heights[2] == 24
    assert data.column_widths["A"] == 18
    assert 4 in data.hidden_rows
    assert "C" in data.hidden_columns
    assert data.print_area == ["Data!$A$1:$B$3"]
    assert snapshot.sheets[1].visibility == "hidden"
    assert any(item["kind"] == "chart" and item["anchor"] == "D2" for item in data.objects)


def test_ooxml_adapter_captures_native_region_anchors(tmp_path):
    path = tmp_path / "anchors.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value", None, "Input", "Rate"])
    sheet.append(["Alpha", 1, None, "Base", 2])
    sheet.add_table(Table(displayName="SalesTable", ref="A1:B2"))
    workbook.defined_names.add(DefinedName("InputBlock", attr_text="'Data'!$D$1:$E$2"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert [
        (anchor.kind, anchor.source_ref.range, anchor.name)
        for anchor in snapshot.sheets[0].region_anchors
    ] == [
        ("excel_table", "A1:B2", "SalesTable"),
        ("defined_name", "D1:E2", "InputBlock"),
    ]


def test_visual_style_fingerprint_ignores_number_format_only_changes(tmp_path):
    path = tmp_path / "number-formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["A1"].number_format = "0"
    sheet["B1"] = 1
    sheet["B1"].number_format = "0.00"
    workbook.save(path)

    data = OOXMLWorkbookAdapter().snapshot(path).sheets[0]

    assert data.cells["A1"].style_id != data.cells["B1"].style_id
    assert data.cells["A1"].visual_style_id == data.cells["B1"].visual_style_id


def test_visual_style_fingerprint_treats_default_and_number_format_only_as_equal(
    tmp_path,
):
    path = tmp_path / "default-and-number-format.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["B1"] = 1
    sheet["B1"].number_format = "0.00"
    workbook.save(path)

    data = OOXMLWorkbookAdapter().snapshot(path).sheets[0]

    assert data.cells["A1"].style_id != data.cells["B1"].style_id
    assert data.cells["A1"].visual_style_id == data.cells["B1"].visual_style_id


def test_whole_column_defined_name_is_not_used_as_a_region_anchor(tmp_path):
    path = tmp_path / "whole-column-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.defined_names.add(DefinedName("AllNames", attr_text="'Data'!$A:$A"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert snapshot.sheets[0].region_anchors == []
    assert snapshot.metadata["warnings"] == ["defined_name_anchor_unsupported"]


def test_sheet_local_defined_name_is_captured_as_a_region_anchor(tmp_path):
    path = tmp_path / "sheet-local-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    sheet.defined_names.add(DefinedName("LocalBlock", attr_text="'Data'!$A$1:$B$2"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert [
        (anchor.kind, anchor.source_ref.key, anchor.name)
        for anchor in snapshot.sheets[0].region_anchors
    ] == [("defined_name", "Data!A1:B2", "LocalBlock")]


def test_excel_table_extent_expands_snapshot_used_range(tmp_path):
    path = tmp_path / "table-with-empty-tail.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    sheet.add_table(Table(displayName="ReservedRows", ref="A1:B4"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert snapshot.sheets[0].used_range == "A1:B4"


def test_sheet_local_defined_name_shadows_same_named_workbook_definition(tmp_path):
    path = tmp_path / "shadowed-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in (["A", "B", "C", "D"], [1, 2, 3, 4]):
        sheet.append(row)
    workbook.defined_names.add(DefinedName("Block", attr_text="'Data'!$A$1:$C$2"))
    sheet.defined_names.add(DefinedName("Block", attr_text="'Data'!$A$1:$B$2"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert [
        (anchor.name, anchor.source_ref.range, anchor.scope)
        for anchor in snapshot.sheets[0].region_anchors
    ] == [("Block", "A1:B2", "worksheet")]


def test_unusable_sheet_local_defined_name_still_shadows_workbook_definition(tmp_path):
    path = tmp_path / "invalid-local-shadow.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.defined_names.add(DefinedName("Block", attr_text="'Data'!$A$1:$B$2"))
    sheet.defined_names.add(DefinedName("Block", attr_text="'Data'!$A:$A"))
    workbook.save(path)

    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    assert snapshot.sheets[0].region_anchors == []
    assert snapshot.metadata["warnings"] == ["defined_name_anchor_unsupported"]
