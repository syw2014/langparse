from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from langparse.workbooks.adapters import OOXMLWorkbookAdapter
from langparse.workbooks.assembly import assemble_workbook
from langparse.workbooks.modeling.policy import WorkbookDisambiguation
from langparse.workbooks.modeling.ports import WorkbookStructureModelAdapter
from langparse.workbooks.modeling.types import (
    ModelIdentity,
    ProviderReply,
    WorkbookModelPolicy,
    WorkbookModelRequest,
)


def _create_synthetic_multi_region_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    # Create two disjoint ambiguous sparse regions
    ws["A1"] = "R1_Top"
    ws["B2"] = "R1_Bottom"
    ws["D10"] = "R2_Top"
    ws["E11"] = "R2_Bottom"
    wb.save(path)


class _QuotaExceedingMockAdapter(WorkbookStructureModelAdapter):
    """Adapter that returns large usage on first call."""

    def __init__(self, token_charge: int = 100_000) -> None:
        self._identity = ModelIdentity(provider="mock", model="gpt-4o-mini", revision="1")
        self.call_count = 0
        self.token_charge = token_charge

    @property
    def identity(self) -> ModelIdentity:
        return self._identity

    def complete(self, request: WorkbookModelRequest, *, timeout_seconds: float) -> ProviderReply:
        self.call_count += 1
        body = json.dumps({"malformed_json": True}).encode("utf-8")
        return ProviderReply(
            body=body,
            provider_request_id=f"req-{self.call_count}",
            usage={
                "prompt_tokens": self.token_charge,
                "completion_tokens": self.token_charge // 2,
                "total_tokens": self.token_charge + (self.token_charge // 2),
            },
        )


def test_malformed_response_still_consumes_quota_and_blocks_further_calls(tmp_path: Path):
    sample = tmp_path / "multi.xlsx"
    _create_synthetic_multi_region_workbook(sample)
    snapshot = OOXMLWorkbookAdapter().snapshot(sample)

    adapter = _QuotaExceedingMockAdapter(token_charge=50_000)
    # Set max_tokens_per_workbook = 10_000
    policy = WorkbookModelPolicy(max_tokens_per_workbook=10_000, max_attempts=2)
    disambiguation = WorkbookDisambiguation.auto(adapter, policy=policy)

    structure, diagnostics = assemble_workbook(snapshot, disambiguation=disambiguation)
    assert structure is not None

    # First case consumes 75,000 tokens during attempt 1.
    # The second attempt / second case MUST be blocked by quota exceeded!
    quota_calls = [c for c in diagnostics.model_calls if c.get("outcome") == "quota_exceeded"]
    assert len(quota_calls) > 0
