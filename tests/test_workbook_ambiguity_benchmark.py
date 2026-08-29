from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from langparse.services.workbook_ambiguity_benchmark import (
    WorkbookAmbiguityBenchmarkService,
    WorkbookEvaluationReport,
)
from langparse.workbooks.evaluation.schema import (
    GoldenSetDriftError,
    WorkbookEvaluationError,
    compute_choices_digest,
)
from langparse.workbooks.modeling.types import ModelIdentity, ProviderReply


def _create_synthetic_fixture_ambiguous(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    # Diagonal sparse text produces deterministic unclassified with weak text choice
    ws["A1"] = "LeftTop"
    ws["B2"] = "RightBottom"
    wb.save(path)


def _create_synthetic_fixture_clear(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    # Standard clean 2D table
    ws["A1"] = "ColA"
    ws["B1"] = "ColB"
    ws["A2"] = "Val1"
    ws["B2"] = "Val2"
    ws["A3"] = "Val3"
    ws["B3"] = "Val4"
    wb.save(path)


def test_benchmark_service_e2e_run(tmp_path: Path):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    f_ambiguous = source_root / "ambiguous.xlsx"
    f_clear = source_root / "clear.xlsx"
    _create_synthetic_fixture_ambiguous(f_ambiguous)
    _create_synthetic_fixture_clear(f_clear)

    # First run a dry capture to get exact runtime facts and choices digest
    import hashlib

    from langparse.services.workbook_ambiguity_benchmark import _EvaluationCaptureAdapter
    from langparse.workbooks.adapters import OOXMLWorkbookAdapter
    from langparse.workbooks.assembly import assemble_workbook
    from langparse.workbooks.modeling.policy import WorkbookDisambiguation

    adapter = _EvaluationCaptureAdapter()
    snapshot = OOXMLWorkbookAdapter().snapshot(f_ambiguous)
    assemble_workbook(snapshot, disambiguation=WorkbookDisambiguation.auto(adapter))

    assert len(adapter.captured_cases) >= 1
    case0 = adapter.captured_cases[0]
    runtime_fact_digest = case0.fact_digest
    runtime_choices_digest = compute_choices_digest(case0.choices)

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    manifest_data = {
        "schema_version": 1,
        "dataset_id": "e2e-seed",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-ambiguous",
                "path": "ambiguous.xlsx",
                "sha256": file_hash(f_ambiguous),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-01",
                        "sheet_name": case0.sheet_name,
                        "source_range": case0.source_range,
                        "expected": "text",
                        "fact_digest": runtime_fact_digest,
                        "choices_digest": runtime_choices_digest,
                    }
                ],
            },
            {
                "sample_id": "sample-clear",
                "path": "clear.xlsx",
                "sha256": file_hash(f_clear),
                "cohort": "clear_no_call",
                "cases": [],
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    service = WorkbookAmbiguityBenchmarkService()
    output_dir = tmp_path / "reports"
    report = service.run(manifest_path, output_dir=output_dir, markdown=True)

    assert isinstance(report, WorkbookEvaluationReport)
    assert report.run_digest.startswith("sha256:")
    assert report.output_path.is_dir()
    assert report.output_path.name == report.run_digest.removeprefix("sha256:")

    # Check generated files
    results_jsonl = report.output_path / "workbook-ambiguity-results.jsonl"
    summary_json = report.output_path / "workbook-ambiguity-summary.json"
    summary_md = report.output_path / "workbook-ambiguity-summary.md"

    assert results_jsonl.is_file()
    assert summary_json.is_file()
    assert summary_md.is_file()

    # Check results.jsonl content (privacy check: no file paths or sheet names)
    lines = results_jsonl.read_text(encoding="utf-8").strip().split("\n")
    assert len(lines) == 1
    row = json.loads(lines[0])
    assert row["run_digest"] == report.run_digest
    assert row["cohort"] == "ambiguous"
    assert row["expected_kind"] == "text"
    assert "Sheet" not in json.dumps(row)
    assert "xlsx" not in json.dumps(row)

    # Check summary.json
    summary = json.loads(summary_json.read_text(encoding="utf-8"))
    assert summary["schema_version"] == 1
    assert summary["dataset_id"] == "e2e-seed"
    assert summary["status"] == "valid"
    assert summary["effectiveness_evidence"] is False
    assert summary["model_contract"] == {
        "schema_version": 1,
        "prompt_version": "region-choice-v2",
        "privacy_version": "region-privacy-v1",
        "rule_version": "region-rules-v1",
        "validator_version": "region-validator-v1",
    }
    assert summary["ambiguous_case_count"] == 1
    assert summary["clear_sample_count"] == 1
    assert summary["clear_unexpected_call_count"] == 0

    # Test idempotent replay (running again with same output_dir should succeed)
    report2 = service.run(manifest_path, output_dir=output_dir, markdown=True)
    assert report2.run_digest == report.run_digest


def test_drift_detection_case_missing(tmp_path: Path):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    f_clear = source_root / "clear.xlsx"
    _create_synthetic_fixture_clear(f_clear)

    import hashlib

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    # Manifest claims an ambiguous case on a clear sheet (will not generate case)
    manifest_data = {
        "schema_version": 1,
        "dataset_id": "e2e-drift",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-01",
                "path": "clear.xlsx",
                "sha256": file_hash(f_clear),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-fake",
                        "sheet_name": "Data",
                        "source_range": "Z100:Z200",
                        "expected": "text",
                        "fact_digest": "sha256:" + "0" * 64,
                        "choices_digest": "sha256:" + "0" * 64,
                    }
                ],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    service = WorkbookAmbiguityBenchmarkService()
    with pytest.raises(GoldenSetDriftError) as exc_info:
        service.run(manifest_path, output_dir=tmp_path / "reports")
    assert exc_info.value.code == "case_missing"


def test_drift_detection_choices_changed(tmp_path: Path):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    f_ambiguous = source_root / "ambiguous.xlsx"
    _create_synthetic_fixture_ambiguous(f_ambiguous)

    import hashlib

    from langparse.services.workbook_ambiguity_benchmark import _EvaluationCaptureAdapter
    from langparse.workbooks.adapters import OOXMLWorkbookAdapter
    from langparse.workbooks.assembly import assemble_workbook
    from langparse.workbooks.modeling.policy import WorkbookDisambiguation

    adapter = _EvaluationCaptureAdapter()
    snapshot = OOXMLWorkbookAdapter().snapshot(f_ambiguous)
    assemble_workbook(snapshot, disambiguation=WorkbookDisambiguation.auto(adapter))
    assert len(adapter.captured_cases) >= 1
    case0 = adapter.captured_cases[0]

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    manifest_data = {
        "schema_version": 1,
        "dataset_id": "e2e-drift-choices",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-01",
                "path": "ambiguous.xlsx",
                "sha256": file_hash(f_ambiguous),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-01",
                        "sheet_name": case0.sheet_name,
                        "source_range": case0.source_range,
                        "expected": "text",
                        "fact_digest": case0.fact_digest,
                        "choices_digest": "sha256:" + "f" * 64,  # wrong choices digest
                    }
                ],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    service = WorkbookAmbiguityBenchmarkService()
    with pytest.raises(GoldenSetDriftError) as exc_info:
        service.run(manifest_path, output_dir=tmp_path / "reports")
    assert exc_info.value.code == "choices_changed"


def test_cli_workbook_ambiguity_benchmark(tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    f_ambiguous = source_root / "ambiguous.xlsx"
    _create_synthetic_fixture_ambiguous(f_ambiguous)

    import hashlib

    from langparse.cli import main
    from langparse.services.workbook_ambiguity_benchmark import _EvaluationCaptureAdapter
    from langparse.workbooks.adapters import OOXMLWorkbookAdapter
    from langparse.workbooks.assembly import assemble_workbook
    from langparse.workbooks.modeling.policy import WorkbookDisambiguation

    adapter = _EvaluationCaptureAdapter()
    snapshot = OOXMLWorkbookAdapter().snapshot(f_ambiguous)
    assemble_workbook(snapshot, disambiguation=WorkbookDisambiguation.auto(adapter))
    case0 = adapter.captured_cases[0]

    def file_hash(p: Path) -> str:
        return f"sha256:{hashlib.sha256(p.read_bytes()).hexdigest()}"

    manifest_data = {
        "schema_version": 1,
        "dataset_id": "cli-test",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "samples": [
            {
                "sample_id": "sample-01",
                "path": "ambiguous.xlsx",
                "sha256": file_hash(f_ambiguous),
                "cohort": "ambiguous",
                "cases": [
                    {
                        "label_id": "label-01",
                        "sheet_name": case0.sheet_name,
                        "source_range": case0.source_range,
                        "expected": "text",
                        "fact_digest": case0.fact_digest,
                        "choices_digest": compute_choices_digest(case0.choices),
                    }
                ],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest_data), encoding="utf-8")

    out_dir = tmp_path / "cli_reports"
    exit_code = main(
        ["benchmark-workbook-ambiguity", str(manifest_path), "--output-dir", str(out_dir)]
    )
    assert exit_code == 0
    captured = capsys.readouterr()
    assert "workbook ambiguity benchmark completed" in captured.out


class _InvalidChecksumAdapter:
    identity = ModelIdentity(provider="scripted-provider", model="fixture", revision="1")

    def complete(self, request, *, timeout_seconds: float) -> ProviderReply:
        request_payload = json.loads(request.body)
        case = request_payload["cases"][0]
        selected = next(choice for choice in case["choices"] if choice["kind"] == "text")
        return ProviderReply(
            body=json.dumps(
                {
                    "schema_version": 1,
                    "request_checksum": "sha256:" + "0" * 64,
                    "decisions": [
                        {
                            "case_id": case["case_id"],
                            "status": "selected",
                            "choice_id": selected["choice_id"],
                            "confidence": 0.99,
                            "reason_codes": ["invalid_checksum"],
                        }
                    ],
                }
            ).encode(),
            provider_request_id="invalid-checksum",
        )


class _AbstainingAdapter:
    def __init__(self, model: str) -> None:
        self._identity = ModelIdentity(provider="scripted-provider", model=model, revision="1")

    @property
    def identity(self) -> ModelIdentity:
        return self._identity

    def complete(self, request, *, timeout_seconds: float) -> ProviderReply:
        return ProviderReply(
            body=json.dumps(
                {
                    "schema_version": request.schema_version,
                    "request_checksum": request.request_checksum,
                    "decisions": [
                        {
                            "case_id": request.case_ids[0],
                            "status": "abstained",
                            "confidence": 0.0,
                            "reason_codes": ["scripted_abstention"],
                        }
                    ],
                }
            ).encode(),
            provider_request_id="scripted-abstention",
        )


def test_live_evaluation_uses_final_production_audits(tmp_path: Path):
    report = WorkbookAmbiguityBenchmarkService().run(
        Path("samples/workbook_ambiguity/public-manifest.json"),
        output_dir=tmp_path / "reports",
        adapter=_InvalidChecksumAdapter(),
    )

    assert report.summary["effectiveness_evidence"] is False
    assert report.summary["production_ready"] is False
    assert report.metrics.model_failed_count == 2
    assert {result.observation_status for result in report.results} == {"failed"}


def test_run_digest_and_summary_include_model_identity(tmp_path: Path):
    service = WorkbookAmbiguityBenchmarkService()
    manifest = Path("samples/workbook_ambiguity/public-manifest.json")

    first = service.run(
        manifest,
        output_dir=tmp_path / "first",
        adapter=_AbstainingAdapter("router/model-a"),
    )
    second = service.run(
        manifest,
        output_dir=tmp_path / "second",
        adapter=_AbstainingAdapter("router/model-b"),
    )

    assert first.run_digest != second.run_digest
    assert first.summary["model_identity"] == {
        "provider": "scripted-provider",
        "model": "router/model-a",
        "revision": "1",
    }


def test_idempotent_replay_rejects_tampered_or_incomplete_report(tmp_path: Path):
    service = WorkbookAmbiguityBenchmarkService()
    output_dir = tmp_path / "reports"
    first = service.run(
        Path("samples/workbook_ambiguity/public-manifest.json"),
        output_dir=output_dir,
        markdown=False,
    )
    results_path = first.output_path / "workbook-ambiguity-results.jsonl"
    results_path.write_text('{"tampered":true}\n', encoding="utf-8")

    with pytest.raises(WorkbookEvaluationError) as exc_info:
        service.run(
            Path("samples/workbook_ambiguity/public-manifest.json"),
            output_dir=output_dir,
            markdown=True,
        )

    assert exc_info.value.code == "report_conflict"
    assert results_path.read_text(encoding="utf-8") == '{"tampered":true}\n'
