from openpyxl.utils import get_column_letter

from langparse.workbooks.blocks import (
    interpret_form_block,
    interpret_matrix_block,
    interpret_text_block,
)
from langparse.workbooks.classification import classify_candidate_region
from langparse.workbooks.types import CandidateRegion, CellSnapshot, SheetSnapshot, SourceRef


def _fixture(rows: list[list[object | None]]) -> tuple[SheetSnapshot, CandidateRegion]:
    sheet = SheetSnapshot(name="Data", index=0)
    for row_number, values in enumerate(rows, start=1):
        for column_number, value in enumerate(values, start=1):
            if value is None:
                continue
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            sheet.cells[coordinate] = CellSnapshot(
                coordinate=coordinate,
                raw_value=value,
                display_value=str(value),
            )
    max_column = max(len(row) for row in rows)
    candidate = CandidateRegion(
        source_ref=SourceRef(
            sheet_name=sheet.name,
            range=f"A1:{get_column_letter(max_column)}{len(rows)}",
        ),
        cell_refs=list(sheet.cells),
    )
    return sheet, candidate


def test_interprets_form_fields_and_preserves_unmatched_text():
    sheet, candidate = _fixture(
        [
            ["项目登记", None],
            ["项目名称", "道路工程"],
            ["建设单位", "示例公司"],
            ["本表由建设单位填写", None],
        ]
    )
    classification = classify_candidate_region(sheet, candidate)

    form = interpret_form_block(sheet, candidate, classification)

    assert form.title == "项目登记"
    assert [(field.label, field.value) for field in form.fields] == [
        ("项目名称", "道路工程"),
        ("建设单位", "示例公司"),
    ]
    assert form.fields[0].label_source_refs[0].key == "Data!A2"
    assert form.fields[0].value_source_refs[0].key == "Data!B2"
    assert [line.text for line in form.free_text] == ["本表由建设单位填写"]
    assert form.free_text[0].source_refs[0].key == "Data!A4"
    assert form.source_refs == [candidate.source_ref]


def test_interprets_matrix_axes_values_and_value_sources():
    sheet, candidate = _fixture(
        [
            ["指标", "1月", "2月"],
            ["收入", 10, 12],
            ["成本", 3, 4],
        ]
    )
    classification = classify_candidate_region(sheet, candidate)

    matrix = interpret_matrix_block(sheet, candidate, classification)

    assert [header.value for header in matrix.column_headers] == ["1月", "2月"]
    assert [header.value for header in matrix.row_headers] == ["收入", "成本"]
    assert matrix.values == [["10", "12"], ["3", "4"]]
    assert [[ref.key for ref in row] for row in matrix.value_source_refs] == [
        ["Data!B2", "Data!C2"],
        ["Data!B3", "Data!C3"],
    ]
    assert matrix.column_headers[0].source_refs[0].key == "Data!B1"
    assert matrix.row_headers[0].source_refs[0].key == "Data!A2"


def test_interprets_text_lines_in_source_order_without_merge_duplicates():
    sheet = SheetSnapshot(name="Notes", index=0, used_range="A1:B3")
    sheet.cells = {
        "A1": CellSnapshot(
            coordinate="A1",
            raw_value="项目说明",
            display_value="项目说明",
            colspan=2,
        ),
        "B1": CellSnapshot(coordinate="B1", merge_anchor="A1"),
        "A2": CellSnapshot(coordinate="A2", raw_value="第一行", display_value="第一行"),
        "B2": CellSnapshot(coordinate="B2", raw_value="补充", display_value="补充"),
        "A3": CellSnapshot(coordinate="A3", raw_value="第二行", display_value="第二行"),
    }
    candidate = CandidateRegion(
        source_ref=SourceRef(sheet_name=sheet.name, range="A1:B3"),
        cell_refs=list(sheet.cells),
    )
    classification = classify_candidate_region(sheet, candidate)

    text = interpret_text_block(sheet, candidate, classification)

    assert [line.text for line in text.lines] == ["项目说明", "第一行 补充", "第二行"]
    assert [ref.key for ref in text.lines[0].source_refs] == ["Notes!A1"]
    assert [ref.key for ref in text.lines[1].source_refs] == ["Notes!A2", "Notes!B2"]
