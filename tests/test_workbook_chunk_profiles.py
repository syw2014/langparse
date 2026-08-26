import pytest
from openpyxl import Workbook

from langparse.chunkers.profiles import (
    ChunkProfileNotSupportedError,
    WorkbookChunkProfile,
    resolve_workbook_chunk_policy,
)
from langparse.chunkers.workbook import WorkbookStructuralChunker
from langparse.parsers.excel_parser import ExcelParser
from langparse.types import ParsedDocumentResult, ParsedPageResult
from langparse.workbooks.types import (
    FormBlock,
    MatrixBlock,
    MatrixHeader,
    SheetIR,
    SheetSnapshot,
    SourceRef,
    TextBlock,
    TextLine,
    WorkbookBlock,
    WorkbookIR,
    WorkbookSnapshot,
)


def test_workbook_profile_defaults_and_budgets_are_stable():
    default = resolve_workbook_chunk_policy(None)
    retrieval = resolve_workbook_chunk_policy("retrieval")
    analysis = resolve_workbook_chunk_policy(WorkbookChunkProfile.ANALYSIS)

    assert default is retrieval
    assert retrieval.name is WorkbookChunkProfile.RETRIEVAL
    assert retrieval.version == 1
    assert retrieval.default_max_chunk_size == 1000
    assert retrieval.analysis_records is False
    assert analysis.name is WorkbookChunkProfile.ANALYSIS
    assert analysis.version == 1
    assert analysis.default_max_chunk_size == 4000
    assert analysis.analysis_records is True


def test_unknown_workbook_profile_lists_the_supported_values():
    with pytest.raises(
        ValueError,
        match="Unknown workbook chunk profile 'balanced'. Available: analysis, retrieval",
    ):
        resolve_workbook_chunk_policy("balanced")


def test_workbook_chunker_uses_profile_budget_unless_explicitly_overridden():
    retrieval = WorkbookStructuralChunker()
    analysis = WorkbookStructuralChunker(profile="analysis")
    override = WorkbookStructuralChunker(profile="analysis", max_chunk_size=321)

    assert retrieval.policy.name is WorkbookChunkProfile.RETRIEVAL
    assert retrieval.max_chunk_size == 1000
    assert analysis.policy.name is WorkbookChunkProfile.ANALYSIS
    assert analysis.max_chunk_size == 4000
    assert override.max_chunk_size == 321


def test_workbook_chunker_rejects_non_positive_explicit_budget():
    with pytest.raises(ValueError, match="max_chunk_size must be positive"):
        WorkbookStructuralChunker(max_chunk_size=0)


def test_profile_not_supported_error_is_a_value_error():
    assert issubclass(ChunkProfileNotSupportedError, ValueError)


def test_legacy_positional_max_chunk_size_remains_supported():
    chunker = WorkbookStructuralChunker(120)

    assert chunker.policy.name is WorkbookChunkProfile.RETRIEVAL
    assert chunker.max_chunk_size == 120


def _large_table(tmp_path):
    path = tmp_path / "large-table.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Visible"
    sheet.append(["Name", "Description", "Value"])
    for index in range(1, 41):
        sheet.append([f"Item {index}", f"description {index} " * 8, index])
    sheet.row_dimensions[2].hidden = True
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden.append(["Name", "Value"])
    hidden.append(["Secret", 7])
    workbook.save(path)
    return ExcelParser().parse_result(path)


def _semantic_workbook_result(
    *,
    sheet_name: str,
    used_range: str,
    block: WorkbookBlock,
) -> ParsedDocumentResult:
    snapshot = WorkbookSnapshot(
        source="semantic.xlsx",
        filename="semantic.xlsx",
        sheets=[SheetSnapshot(name=sheet_name, index=0, used_range=used_range)],
    )
    structure = WorkbookIR(
        kind="workbook",
        workbook_id="workbook_1",
        source="semantic.xlsx",
        filename="semantic.xlsx",
        snapshot=snapshot,
        sheets=[SheetIR(sheet_id="sheet_1", name=sheet_name, index=0, blocks=[block])],
    )
    return ParsedDocumentResult(
        source="semantic.xlsx",
        filename="semantic.xlsx",
        engine="excel",
        pages=[
            ParsedPageResult(
                page_number=1,
                markdown_content="",
                metadata={"sheet_name": sheet_name},
            )
        ],
        paginated=False,
        structure=structure,
    )


def _simple_workbook_result(tmp_path, *, sheet_name: str = "Sheet"):
    path = tmp_path / "simple.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Name", "Value"])
    sheet.append(["Road", 10])
    workbook.save(path)
    return ExcelParser().parse_result(path)


def test_profiles_add_versioned_metadata_and_analysis_packs_more_rows(tmp_path):
    parsed = _large_table(tmp_path)

    retrieval = WorkbookStructuralChunker(profile="retrieval").chunk(parsed)
    analysis = WorkbookStructuralChunker(profile="analysis").chunk(parsed)
    retrieval_table = [chunk for chunk in retrieval if chunk.metadata["chunk_type"] == "table_rows"]
    analysis_table = [chunk for chunk in analysis if chunk.metadata["chunk_type"] == "table_rows"]

    assert len(analysis_table) < len(retrieval_table)
    assert {chunk.metadata["chunk_profile"] for chunk in retrieval} == {"retrieval"}
    assert {chunk.metadata["chunk_profile"] for chunk in analysis} == {"analysis"}
    assert {chunk.metadata["chunk_profile_version"] for chunk in retrieval + analysis} == {1}
    assert [chunk.metadata["chunk_index"] for chunk in retrieval] == list(range(len(retrieval)))
    assert [chunk.metadata["chunk_index"] for chunk in analysis] == list(range(len(analysis)))


def test_profile_metadata_exposes_hidden_sources_without_filtering_them(tmp_path):
    parsed = _large_table(tmp_path)
    chunks = WorkbookStructuralChunker().chunk(parsed)

    visible = [chunk for chunk in chunks if chunk.metadata["sheet_name"] == "Visible"]
    hidden = [chunk for chunk in chunks if chunk.metadata["sheet_name"] == "Hidden"]

    assert any(2 in chunk.metadata["hidden_row_numbers"] for chunk in visible)
    assert {chunk.metadata["sheet_visibility"] for chunk in visible} == {"visible"}
    assert {chunk.metadata["sheet_visibility"] for chunk in hidden} == {"hidden"}
    assert any("Secret" in chunk.content for chunk in hidden)


def test_analysis_table_payload_has_source_linked_schema_and_records(tmp_path):
    parsed = _large_table(tmp_path)
    chunk = next(
        item
        for item in WorkbookStructuralChunker(profile="analysis").chunk(parsed)
        if item.metadata["chunk_type"] == "table_rows" and item.metadata["sheet_name"] == "Visible"
    )

    payload = chunk.structured_payload
    assert payload["column_schema"] == [
        {"column_index": 0, "coordinate": "A", "header_path": ["Name"]},
        {"column_index": 1, "coordinate": "B", "header_path": ["Description"]},
        {"column_index": 2, "coordinate": "C", "header_path": ["Value"]},
    ]
    assert len(payload["records"]) == len(payload["rows"]) == len(chunk.metadata["row_ids"])
    first = payload["records"][0]
    assert first == {
        "row_id": chunk.metadata["row_ids"][0],
        "row_number": chunk.metadata["row_numbers"][0],
        "role": "data",
        "section_path": [],
        "values": payload["rows"][0],
        "source_refs": [chunk.metadata["source_ranges"][0]],
    }


def test_retrieval_payload_keeps_existing_keys_without_analysis_records(tmp_path):
    parsed = _large_table(tmp_path)
    chunk = next(
        item
        for item in WorkbookStructuralChunker(profile="retrieval").chunk(parsed)
        if item.metadata["chunk_type"] == "table_rows"
    )

    assert set(chunk.structured_payload) == {"columns", "rows", "roles"}


def test_analysis_raw_grid_payload_has_source_linked_records(tmp_path):
    path = tmp_path / "raw.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "left"
    sheet["B2"] = "right"
    workbook.save(path)
    parsed = ExcelParser().parse_result(path)

    chunk = WorkbookStructuralChunker(profile="analysis").chunk(parsed)[0]

    assert chunk.metadata["chunk_type"] == "raw_grid_rows"
    assert chunk.structured_payload["column_schema"] == [
        {"column_index": 0, "coordinate": "A", "header_path": []},
        {"column_index": 1, "coordinate": "B", "header_path": []},
    ]
    assert chunk.structured_payload["records"] == [
        {
            "row_number": 1,
            "role": "raw",
            "section_path": [],
            "values": ["left", ""],
            "source_refs": ["Sheet!A1:B1"],
        },
        {
            "row_number": 2,
            "role": "raw",
            "section_path": [],
            "values": ["", "right"],
            "source_refs": ["Sheet!A2:B2"],
        },
    ]


def test_chunking_rejects_missing_logical_table_rows(tmp_path):
    class DroppingTableChunker(WorkbookStructuralChunker):
        def _chunk_logical_table(self, *args, **kwargs):
            return []

    with pytest.raises(ValueError, match="Workbook chunk row conservation failed"):
        DroppingTableChunker().chunk(_large_table(tmp_path))


def test_source_range_validation_preserves_sheet_names_containing_an_exclamation_mark(tmp_path):
    # Break caught: splitting a qualified range at its first separator rejects valid sheet names.
    parsed = _simple_workbook_result(tmp_path, sheet_name="Budget!FY26")

    chunks = WorkbookStructuralChunker().chunk(parsed)

    assert chunks
    assert {chunk.metadata["sheet_name"] for chunk in chunks} == {"Budget!FY26"}
    assert all(
        source_range.startswith("Budget!FY26!")
        for chunk in chunks
        for source_range in chunk.metadata["source_ranges"]
    )


def test_source_range_validation_requires_a_workbook_snapshot(tmp_path):
    # Break caught: source references must not be accepted when their workbook boundary is unavailable.
    parsed = _simple_workbook_result(tmp_path)
    assert isinstance(parsed.structure, WorkbookIR)
    parsed.structure.snapshot = None

    with pytest.raises(
        ValueError,
        match="WorkbookIR snapshot is required for source-range validation",
    ):
        WorkbookStructuralChunker().chunk(parsed)


def test_source_range_validation_requires_the_referenced_sheet_used_range(tmp_path):
    # Break caught: validating a range without its sheet boundary accepts unverifiable references.
    parsed = _simple_workbook_result(tmp_path)
    assert isinstance(parsed.structure, WorkbookIR)
    parsed.structure.snapshot.sheets[0].used_range = None

    with pytest.raises(ValueError, match="Workbook sheet used_range is required: Sheet"):
        WorkbookStructuralChunker().chunk(parsed)


def test_source_range_validation_rejects_ranges_outside_the_sheet_used_range(tmp_path):
    # Break caught: a chunk source range must not escape the worksheet's known used range.
    parsed = _simple_workbook_result(tmp_path)
    assert isinstance(parsed.structure, WorkbookIR)
    parsed.structure.snapshot.sheets[0].used_range = "A1:A1"

    with pytest.raises(
        ValueError,
        match=r"Workbook source range is outside sheet used_range: Sheet!A2:B2",
    ):
        WorkbookStructuralChunker().chunk(parsed)


def test_analysis_form_free_text_record_has_literal_keys_and_source_refs():
    # Break caught: free text must not be coerced into a field record with an invented field id.
    source_ref = SourceRef(sheet_name="Cover", range="A2")
    parsed = _semantic_workbook_result(
        sheet_name="Cover",
        used_range="A1:A2",
        block=WorkbookBlock(
            block_id="form_1",
            kind="form",
            source_refs=[source_ref],
            form=FormBlock(
                form_id="form_1",
                free_text=[TextLine(text="Tender note", source_refs=[source_ref])],
            ),
        ),
    )

    chunk = WorkbookStructuralChunker(profile="analysis").chunk(parsed)[0]

    assert chunk.metadata["field_ids"] == []
    assert chunk.metadata["source_ranges"] == ["Cover!A2"]
    assert chunk.structured_payload["records"] == [
        {
            "record_type": "text",
            "text": "Tender note",
            "source_refs": ["Cover!A2"],
        }
    ]


def test_analysis_matrix_records_keep_missing_value_refs_out_of_metadata_ranges():
    # Break caught: absent matrix cells are not source-backed values and must remain None.
    row_ref = SourceRef(sheet_name="Matrix", range="A2")
    value_ref = SourceRef(sheet_name="Matrix", range="B2")
    parsed = _semantic_workbook_result(
        sheet_name="Matrix",
        used_range="A1:C2",
        block=WorkbookBlock(
            block_id="matrix_1",
            kind="matrix",
            source_refs=[SourceRef(sheet_name="Matrix", range="A1:C2")],
            matrix=MatrixBlock(
                matrix_id="matrix_1",
                row_headers=[MatrixHeader(value="Revenue", source_refs=[row_ref])],
                column_headers=[MatrixHeader(value="Jan"), MatrixHeader(value="Feb")],
                values=[["10", ""]],
                value_source_refs=[[value_ref, None]],
            ),
        ),
    )

    chunk = WorkbookStructuralChunker(profile="analysis").chunk(parsed)[0]

    assert chunk.metadata["source_ranges"] == ["Matrix!A2", "Matrix!B2"]
    assert chunk.structured_payload["records"] == [
        {
            "row_header": "Revenue",
            "row_header_source_refs": ["Matrix!A2"],
            "values": ["10", ""],
            "value_source_refs": ["Matrix!B2", None],
        }
    ]


def test_retrieval_matrix_payload_keeps_only_its_legacy_structured_keys():
    # Break caught: retrieval matrix payloads must not gain analysis-only records.
    source_ref = SourceRef(sheet_name="Matrix", range="A1:B2")
    parsed = _semantic_workbook_result(
        sheet_name="Matrix",
        used_range="A1:B2",
        block=WorkbookBlock(
            block_id="matrix_1",
            kind="matrix",
            source_refs=[source_ref],
            matrix=MatrixBlock(
                matrix_id="matrix_1",
                row_headers=[
                    MatrixHeader(value="Revenue", source_refs=[SourceRef("Matrix", "A2")])
                ],
                column_headers=[MatrixHeader(value="Jan")],
                values=[["10"]],
                value_source_refs=[[SourceRef("Matrix", "B2")]],
            ),
        ),
    )

    chunk = WorkbookStructuralChunker(profile="retrieval").chunk(parsed)[0]

    assert set(chunk.structured_payload) == {"column_headers", "row_headers", "values"}


def test_retrieval_text_payload_keeps_only_its_legacy_structured_keys():
    # Break caught: retrieval text payloads must not gain analysis-only records.
    source_ref = SourceRef(sheet_name="Notes", range="A1")
    parsed = _semantic_workbook_result(
        sheet_name="Notes",
        used_range="A1",
        block=WorkbookBlock(
            block_id="text_1",
            kind="text",
            source_refs=[source_ref],
            text=TextBlock(
                text_id="text_1",
                lines=[TextLine(text="Tender note", source_refs=[source_ref])],
            ),
        ),
    )

    chunk = WorkbookStructuralChunker(profile="retrieval").chunk(parsed)[0]

    assert set(chunk.structured_payload) == {"lines"}
