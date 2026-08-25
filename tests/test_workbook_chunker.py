from langparse.chunkers.workbook import WorkbookStructuralChunker
from langparse.parsers.excel_parser import ExcelParser


def test_workbook_chunker_emits_source_aware_chunks(sample_excel_file):
    parsed = ExcelParser().parse_result(sample_excel_file)

    chunks = WorkbookStructuralChunker(max_chunk_size=120).chunk(parsed)

    assert chunks
    assert chunks[0].metadata["chunk_type"] == "table_rows"
    assert chunks[0].metadata["sheet_name"] == "Sheet1"
    assert chunks[0].metadata["source_ranges"]
    assert chunks[0].metadata["fragment_ranges"]
    assert chunks[0].metadata["row_numbers"]
    assert chunks[0].structured_payload["columns"] == ["A", "B"]
    assert chunks[0].structured_payload["rows"]
    assert "| A | B |" in chunks[0].content


def test_workbook_chunker_never_splits_an_oversized_source_row(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "wide.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "x" * 200
    workbook.save(path)
    parsed = ExcelParser().parse_result(path)

    chunks = WorkbookStructuralChunker(max_chunk_size=40).chunk(parsed)

    assert len(chunks) == 1
    assert chunks[0].metadata["oversized"] is True
    assert chunks[0].structured_payload["rows"] == [["x" * 200]]
    assert chunks[0].metadata["row_numbers"] == [1]


def test_parse_result_chunk_true_populates_result(sample_excel_file):
    from langparse.services.parse_service import ParseService

    parsed = ParseService().parse_result(sample_excel_file, chunk=True)

    assert parsed.chunks
    assert parsed.chunks[0].metadata["chunk_type"] == "table_rows"


def test_logical_chunks_do_not_mix_sections(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "sections.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["序号", "名称", "说明", "金额"])
    sheet.append([0, "土方", "", 100])
    sheet.append([1, "挖土", "", 40])
    sheet.append([0, "管道", "", 200])
    sheet.append([2, "安装", "", 80])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    chunks = WorkbookStructuralChunker(max_chunk_size=1000).chunk(parsed)

    assert [chunk.metadata["section_path"] for chunk in chunks] == [["土方"], ["管道"]]
    assert all(chunk.metadata["table_id"] for chunk in chunks)
    assert all(chunk.metadata["header_paths"] for chunk in chunks)
    assert chunks[0].metadata["row_ids"]


def test_workbook_chunker_emits_every_block_in_a_mixed_sheet(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "项目登记"
    sheet["A2"] = "项目名称"
    sheet["B2"] = "道路工程"
    sheet["A3"] = "建设单位"
    sheet["B3"] = "示例公司"
    sheet["A6"] = "Name"
    sheet["B6"] = "Value"
    sheet["A7"] = "Alpha"
    sheet["B7"] = 1
    sheet["A8"] = "Beta"
    sheet["B8"] = 2
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    chunks = WorkbookStructuralChunker(max_chunk_size=1000).chunk(parsed)

    assert [chunk.metadata["chunk_type"] for chunk in chunks] == [
        "form_fields",
        "table_rows",
    ]
    assert chunks[0].metadata["form_id"]
    assert len(chunks[0].metadata["field_ids"]) == 2
    assert chunks[0].metadata["source_ranges"] == [
        "Sheet!A2",
        "Sheet!B2",
        "Sheet!A3",
        "Sheet!B3",
    ]


def test_workbook_chunker_emits_matrix_and_text_payloads(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "semantic-blocks.xlsx"
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "Matrix"
    for row in [["指标", "1月", "2月"], ["收入", 10, 12], ["成本", 3, 4]]:
        matrix.append(row)
    notes = workbook.create_sheet("Notes")
    notes.append(["第一行"])
    notes.append(["第二行"])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    chunks = WorkbookStructuralChunker(max_chunk_size=1000).chunk(parsed)

    assert [chunk.metadata["chunk_type"] for chunk in chunks] == [
        "matrix_rows",
        "text_block",
    ]
    assert chunks[0].metadata["matrix_id"]
    assert chunks[0].structured_payload["column_headers"] == ["1月", "2月"]
    assert chunks[1].metadata["text_id"]
    assert chunks[1].structured_payload["lines"] == ["第一行", "第二行"]


def test_workbook_chunker_keeps_oversized_form_field_intact(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "form.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["登记表"])
    sheet.append(["项目名称", "x" * 200])
    sheet.append(["建设单位", "示例公司"])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    chunks = WorkbookStructuralChunker(max_chunk_size=40).chunk(parsed)

    assert chunks[0].metadata["chunk_type"] == "form_fields"
    assert chunks[0].metadata["oversized"] is True
    assert chunks[0].structured_payload["fields"] == [["项目名称", "x" * 200]]


def test_workbook_chunker_limits_raw_fallback_to_its_candidate_range(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "raw-and-table.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "左上"
    sheet["B2"] = "右下"
    sheet["A5"] = "Name"
    sheet["B5"] = "Value"
    sheet["A6"] = "Alpha"
    sheet["B6"] = 1
    sheet["A7"] = "Beta"
    sheet["B7"] = 2
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    chunks = WorkbookStructuralChunker(max_chunk_size=1000).chunk(parsed)

    assert [chunk.metadata["chunk_type"] for chunk in chunks] == [
        "raw_grid_rows",
        "table_rows",
    ]
    assert chunks[0].metadata["source_ranges"] == ["Sheet!A1:B2"]
    assert chunks[0].structured_payload["rows"] == [["左上", ""], ["", "右下"]]
