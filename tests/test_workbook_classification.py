from openpyxl.utils import get_column_letter

from langparse.workbooks.classification import (
    classify_candidate_region,
    extract_region_features,
)
from langparse.workbooks.types import CandidateRegion, CellSnapshot, SheetSnapshot, SourceRef


def _fixture(rows: list[list[object | None]]) -> tuple[SheetSnapshot, CandidateRegion]:
    sheet = SheetSnapshot(name="Data", index=0)
    for row_number, values in enumerate(rows, start=1):
        for column_number, value in enumerate(values, start=1):
            if value is None:
                continue
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            sheet.cells[coordinate] = CellSnapshot(
                coordinate=coordinate,
                raw_value=value,
                display_value=str(value),
            )
    max_column = max(len(row) for row in rows)
    source_range = f"A1:{get_column_letter(max_column)}{len(rows)}"
    candidate = CandidateRegion(
        source_ref=SourceRef(sheet_name=sheet.name, range=source_range),
        cell_refs=list(sheet.cells),
    )
    return sheet, candidate


def test_extracts_matrix_shape_without_form_pairs():
    sheet, candidate = _fixture(
        [
            ["指标", "1月", "2月"],
            ["收入", 10, 12],
            ["成本", 3, 4],
        ]
    )

    features = extract_region_features(sheet, candidate)

    assert features.row_count == 3
    assert features.column_count == 3
    assert features.numeric_grid_rows == 2
    assert features.numeric_grid_columns == 2
    assert features.label_value_pairs == 0


def test_extracts_form_pairs_and_table_schema_as_distinct_signals():
    form_sheet, form_candidate = _fixture(
        [
            ["项目登记", None],
            ["项目名称", "道路工程"],
            ["建设单位", "示例公司"],
        ]
    )
    table_sheet, table_candidate = _fixture(
        [
            ["Name", "Value"],
            ["Alpha", 1],
            ["Beta", 2],
        ]
    )

    form_features = extract_region_features(form_sheet, form_candidate)
    table_features = extract_region_features(table_sheet, table_candidate)

    assert form_features.label_value_pairs == 2
    assert form_features.label_value_coverage == 2 / 3
    assert form_features.has_stable_table_schema is False
    assert table_features.has_stable_table_schema is True


def test_classifies_five_region_shapes_deterministically():
    fixtures = {
        "logical_table": _fixture([["Name", "Value"], ["Alpha", 1], ["Beta", 2]]),
        "form": _fixture([["项目登记", None], ["项目名称", "道路工程"], ["建设单位", "示例公司"]]),
        "matrix": _fixture([["指标", "1月", "2月"], ["收入", 10, 12], ["成本", 3, 4]]),
        "text": _fixture(
            [["项目说明"], ["本项目包含道路、排水及附属设施建设内容。"], ["请按图施工。"]]
        ),
    }
    ambiguous_sheet, ambiguous_candidate = _fixture([["左上", None], [None, "右下"]])

    results = {
        expected: classify_candidate_region(sheet, candidate)
        for expected, (sheet, candidate) in fixtures.items()
    }
    ambiguous = classify_candidate_region(ambiguous_sheet, ambiguous_candidate)

    assert {expected: result.kind for expected, result in results.items()} == {
        expected: expected for expected in fixtures
    }
    assert results["form"].reason_codes == ["stable_label_value_pairs"]
    assert results["matrix"].reason_codes == ["numeric_matrix_with_axes"]
    assert results["logical_table"].reason_codes == ["stable_header_data_schema"]
    assert all(result.confidence >= 0.8 for result in results.values())
    assert ambiguous.kind == "unclassified"
    assert ambiguous.confidence < 0.8
    assert ambiguous.reason_codes == ["insufficient_semantic_evidence"]


def test_classifies_presentation_cover_as_text_instead_of_table():
    sheet, candidate = _fixture(
        [
            ["工程名称", None, "示例道路工程"],
            ["招 标 控 制 价", None, None],
            ["招标控制价", "(小写)：", "1584176元"],
            ["招 标 人：", None, "中介机构："],
            ["法定代表人：", None, "复核人："],
        ]
    )
    sheet.cells["A2"].colspan = 3

    result = classify_candidate_region(sheet, candidate)

    assert result.kind == "text"
    assert result.reason_codes == ["presentation_text_region"]


def test_native_excel_table_anchor_overrides_sparse_form_shape():
    sheet, candidate = _fixture(
        [
            ["Name", "Value"],
            ["Alpha", 1],
            [None, None],
            [None, None],
        ]
    )
    candidate.reason_codes = ["native_table_anchor"]

    result = classify_candidate_region(sheet, candidate)

    assert result.kind == "logical_table"
    assert result.confidence == 0.98
    assert result.reason_codes == ["native_table_anchor"]
