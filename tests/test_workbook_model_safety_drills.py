from __future__ import annotations

import json
import os
from pathlib import Path
from unittest import mock

import pytest
from openpyxl import Workbook

from langparse.workbooks.adapters import OOXMLWorkbookAdapter
from langparse.workbooks.assembly import assemble_workbook
from langparse.workbooks.evaluation.evaluator import (
    WorkbookEvaluationMetrics,
    assess_production_readiness,
)
from langparse.workbooks.modeling.policy import WorkbookDisambiguation
from langparse.workbooks.modeling.ports import (
    RequiredWorkbookDisambiguationError,
    WorkbookStructureModelAdapter,
)
from langparse.workbooks.modeling.types import (
    ModelIdentity,
    ProviderReply,
    WorkbookModelPolicy,
    WorkbookModelRequest,
)


def _create_synthetic_ambiguous_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "LeftTop"
    ws["B2"] = "RightBottom"
    wb.save(path)


class _MockAdapter(WorkbookStructureModelAdapter):
    def __init__(
        self, reply_payload: dict | None = None, should_fail: bool = False, tokens: int = 100
    ) -> None:
        self._identity = ModelIdentity(provider="mock", model="gpt-4o-mini", revision="1")
        self.call_count = 0
        self.reply_payload = reply_payload
        self.should_fail = should_fail
        self.tokens = tokens

    @property
    def identity(self) -> ModelIdentity:
        return self._identity

    def complete(self, request: WorkbookModelRequest, *, timeout_seconds: float) -> ProviderReply:
        self.call_count += 1
        if self.should_fail:
            raise RuntimeError("Upstream API server error 500")

        if self.reply_payload is not None:
            body = json.dumps(self.reply_payload).encode("utf-8")
        else:
            body = json.dumps(
                {
                    "schema_version": 1,
                    "request_checksum": request.request_checksum,
                    "decisions": [
                        {
                            "case_id": cid,
                            "status": "abstained",
                            "confidence": 0.0,
                            "reason_codes": ["mock"],
                        }
                        for cid in request.case_ids
                    ],
                }
            ).encode("utf-8")

        return ProviderReply(
            body=body,
            provider_request_id="mock-req",
            usage={
                "prompt_tokens": self.tokens,
                "completion_tokens": self.tokens // 2,
                "total_tokens": self.tokens + self.tokens // 2,
            },
        )


class _NoUsageAdapter(WorkbookStructureModelAdapter):
    identity = ModelIdentity(provider="mock", model="gpt-4o-mini", revision="1")

    def __init__(self) -> None:
        self.call_count = 0

    def complete(self, request: WorkbookModelRequest, *, timeout_seconds: float) -> ProviderReply:
        self.call_count += 1
        return ProviderReply(
            body=json.dumps(
                {
                    "schema_version": 1,
                    "request_checksum": request.request_checksum,
                    "decisions": [
                        {
                            "case_id": request.case_ids[0],
                            "status": "abstained",
                            "confidence": 0.0,
                            "reason_codes": ["mock"],
                        }
                    ],
                }
            ).encode(),
            provider_request_id="mock-no-usage",
        )


def test_kill_switch_via_env_var(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter()
    disambiguation = WorkbookDisambiguation.auto(adapter)

    with mock.patch.dict(os.environ, {"LANGPARSE_DISABLE_MODEL": "1"}):
        structure, diagnostics = assemble_workbook(snapshot, disambiguation=disambiguation)
        assert adapter.call_count == 0
        assert len(diagnostics.model_calls) > 0
        assert diagnostics.model_calls[0]["outcome"] == "kill_switch_activated"


def test_kill_switch_via_policy(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter()
    policy = WorkbookModelPolicy(kill_switch=True)
    disambiguation = WorkbookDisambiguation.auto(adapter, policy=policy)

    structure, diagnostics = assemble_workbook(snapshot, disambiguation=disambiguation)
    assert adapter.call_count == 0
    assert len(diagnostics.model_calls) > 0
    assert diagnostics.model_calls[0]["outcome"] == "kill_switch_activated"


def test_token_quota_limit_exceeded(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter(tokens=500)
    # Set max_tokens_per_workbook lower than initial calls or 1
    policy = WorkbookModelPolicy(max_tokens_per_workbook=1)
    # If quota is 1 and tracker starts at 0, first call happens or exceeds depending on tracker check
    # In disambiguation: before call, tracker["tokens"] (0) < 1, so call 1 happens. If there were subsequent cases, quota exceeds.
    disambiguation = WorkbookDisambiguation.auto(adapter, policy=policy)
    structure, diagnostics = assemble_workbook(snapshot, disambiguation=disambiguation)
    assert structure is not None


def test_paid_invalid_response_consumes_quota_before_retry(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter(reply_payload={"invalid_schema": True}, tokens=500)
    policy = WorkbookModelPolicy(max_tokens_per_workbook=1)
    _, diagnostics = assemble_workbook(
        snapshot,
        disambiguation=WorkbookDisambiguation.auto(adapter, policy=policy),
    )

    assert adapter.call_count == 1
    assert diagnostics.model_calls[0]["outcome"] == "quota_exceeded"


def test_missing_usage_fails_closed_when_quota_is_enabled(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _NoUsageAdapter()
    _, diagnostics = assemble_workbook(
        snapshot,
        disambiguation=WorkbookDisambiguation.auto(
            adapter,
            policy=WorkbookModelPolicy(max_tokens_per_workbook=100),
        ),
    )

    assert adapter.call_count == 1
    assert diagnostics.model_calls[0]["outcome"] == "quota_unavailable"


def test_cost_quota_uses_explicit_rates_for_custom_provider(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter(reply_payload={"invalid_schema": True}, tokens=500)
    adapter._identity = ModelIdentity(provider="mock", model="custom-model", revision="1")
    _, diagnostics = assemble_workbook(
        snapshot,
        disambiguation=WorkbookDisambiguation.auto(
            adapter,
            policy=WorkbookModelPolicy(
                max_cost_usd_per_workbook=0.000001,
                input_cost_usd_per_million=1.0,
                output_cost_usd_per_million=1.0,
                cost_pricing_version="mock-contract-2026-q3",
            ),
        ),
    )

    assert adapter.call_count == 1
    assert diagnostics.model_calls[0]["outcome"] == "quota_exceeded"


def test_transactional_rollback_on_malformed_response(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    # Return invalid JSON body
    adapter = _MockAdapter(should_fail=False)
    adapter.reply_payload = {"invalid_schema": True}
    disambiguation = WorkbookDisambiguation.auto(adapter)

    structure, diagnostics = assemble_workbook(snapshot, disambiguation=disambiguation)
    assert structure is not None
    # Auto mode safely rolled back / fell back to deterministic baseline
    assert len(diagnostics.model_calls) > 0
    assert diagnostics.model_calls[0]["outcome"] in ("invalid_response", "request_error")


def test_required_mode_failure_drill(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    _create_synthetic_ambiguous_workbook(path)
    snapshot = OOXMLWorkbookAdapter().snapshot(path)

    adapter = _MockAdapter(should_fail=True)
    disambiguation = WorkbookDisambiguation.required(adapter)

    with pytest.raises(RequiredWorkbookDisambiguationError) as exc_info:
        assemble_workbook(snapshot, disambiguation=disambiguation)
    assert len(exc_info.value.case_ids) > 0


def test_assess_production_readiness_rules():
    # 1. Missing provider evidence
    metrics_no_evidence = WorkbookEvaluationMetrics(
        ambiguous_case_count=10,
        resolvable_case_count=10,
        unresolvable_case_count=0,
        baseline_correct_count=5,
        baseline_wrong_count=5,
        baseline_accuracy=0.5,
        effectiveness_evidence=False,
    )
    ready, reasons = assess_production_readiness(metrics_no_evidence)
    assert not ready
    assert "effectiveness_evidence_missing" in reasons

    # 2. Fully compliant provider metrics (holdout, sufficient cases, operational evidence)
    metrics_good = WorkbookEvaluationMetrics(
        ambiguous_case_count=35,
        resolvable_case_count=35,
        unresolvable_case_count=0,
        baseline_correct_count=15,
        baseline_wrong_count=20,
        baseline_accuracy=15 / 35,
        model_selected_count=30,
        model_correct_acceptance_count=30,
        model_wrong_acceptance_count=0,
        model_abstained_count=5,
        model_unresolved_count=0,
        model_failed_count=0,
        model_selection_accuracy=1.0,
        wrong_acceptance_rate=0.0,
        model_coverage=30 / 35,
        abstain_rate=5 / 35,
        unresolved_rate=0.0,
        failure_rate=0.0,
        fixed_error_count=15,
        introduced_error_count=0,
        unchanged_correct_count=15,
        unchanged_wrong_count=5,
        net_correct_delta=15,
        clear_sample_count=10,
        clear_unexpected_call_count=0,
        clear_call_rate=0.0,
        effectiveness_evidence=True,
    )
    ready, reasons = assess_production_readiness(
        metrics_good,
        split="holdout",
        operational_evidence=True,
    )
    assert ready
    assert len(reasons) == 0

    # 3. High wrong acceptance rate
    metrics_bad_acceptance = WorkbookEvaluationMetrics(
        ambiguous_case_count=10,
        resolvable_case_count=10,
        unresolvable_case_count=0,
        baseline_correct_count=5,
        baseline_wrong_count=5,
        baseline_accuracy=0.5,
        model_selected_count=8,
        model_correct_acceptance_count=5,
        model_wrong_acceptance_count=3,
        model_abstained_count=2,
        model_unresolved_count=0,
        model_failed_count=0,
        model_selection_accuracy=0.625,
        wrong_acceptance_rate=0.30,  # > 5%
        model_coverage=0.8,
        abstain_rate=0.2,
        unresolved_rate=0.0,
        failure_rate=0.0,
        fixed_error_count=2,
        introduced_error_count=3,
        unchanged_correct_count=3,
        unchanged_wrong_count=2,
        net_correct_delta=-1,
        clear_sample_count=5,
        clear_unexpected_call_count=0,
        clear_call_rate=0.0,
        effectiveness_evidence=True,
    )
    ready, reasons = assess_production_readiness(metrics_bad_acceptance)
    assert not ready
    assert "wrong_acceptance_rate_exceeds_threshold" in reasons
    assert "net_correct_delta_non_positive" in reasons


def test_release_readiness_requires_holdout_scale_and_operational_evidence():
    metrics = WorkbookEvaluationMetrics(
        ambiguous_case_count=2,
        resolvable_case_count=2,
        unresolvable_case_count=0,
        baseline_correct_count=1,
        baseline_wrong_count=1,
        baseline_accuracy=0.5,
        model_selected_count=1,
        model_correct_acceptance_count=1,
        model_wrong_acceptance_count=0,
        model_abstained_count=1,
        model_unresolved_count=0,
        model_failed_count=0,
        fixed_error_count=1,
        introduced_error_count=0,
        unchanged_correct_count=0,
        unchanged_wrong_count=0,
        wrong_acceptance_rate=0.0,
        clear_sample_count=1,
        clear_unexpected_call_count=0,
        clear_call_rate=0.0,
        net_correct_delta=1,
        effectiveness_evidence=True,
    )

    ready, reasons = assess_production_readiness(
        metrics,
        split="tuning",
        operational_evidence=False,
    )

    assert ready is False
    assert set(reasons) >= {
        "holdout_evidence_required",
        "minimum_ambiguous_cases_not_met",
        "operational_evidence_missing",
    }
