import hashlib
import json
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _write_table(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.save(path)


def _write_manifest(tmp_path: Path) -> tuple[Path, Path]:
    fixtures = tmp_path / "fixtures"
    fixtures.mkdir()
    workbook_path = fixtures / "table.xlsx"
    _write_table(workbook_path)
    manifest = {
        "schema_version": 1,
        "dataset_id": "quality-test",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": "fixtures",
        "quality_gate": {
            "minimum": {
                "block_precision": 1.0,
                "block_recall": 1.0,
                "header_path_accuracy": 1.0,
                "row_role_f1": 1.0,
                "source_ref_completeness": 1.0,
                "cell_coverage_ratio": 1.0,
            },
            "maximum": {"fallback_rate": 0.0},
        },
        "samples": [
            {
                "sample_id": "simple-table",
                "path": "table.xlsx",
                "sha256": "sha256:" + hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
                "expectation": {
                    "sheets": [
                        {
                            "name": "Data",
                            "blocks": [
                                {
                                    "source_range": "A1:B2",
                                    "kind": "logical_table",
                                    "headers": [
                                        {"coordinate": "A", "path": ["Name"]},
                                        {"coordinate": "B", "path": ["Value"]},
                                    ],
                                    "rows": [
                                        {"source_range": "A1:B1", "role": "header"},
                                        {"source_range": "A2:B2", "role": "data"},
                                    ],
                                    "form_fields": [],
                                    "matrix_axes": {"rows": [], "columns": []},
                                }
                            ],
                        }
                    ],
                    "continuations": [],
                    "required_source_refs": ["Data!A1:B2"],
                    "objects": [],
                },
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    return manifest_path, workbook_path


def test_runs_workbook_quality_benchmark_end_to_end(tmp_path: Path):
    from langparse.services import WorkbookQualityBenchmarkService

    manifest_path, workbook_path = _write_manifest(tmp_path)
    before = (workbook_path.stat().st_size, workbook_path.stat().st_mtime_ns)

    report = WorkbookQualityBenchmarkService().run(
        manifest_path,
        output_dir=tmp_path / "reports",
    )

    assert report.summary["status"] == "passed"
    assert report.summary["metric_schema_version"] == 1
    assert report.summary["metrics"]["block_recall"] == 1.0
    assert report.summary["metrics"]["fallback_rate"] == 0.0
    assert report.output_path.joinpath("workbook-quality-results.jsonl").is_file()
    assert report.output_path.joinpath("workbook-quality-summary.json").is_file()
    assert report.output_path.joinpath("workbook-quality-summary.md").is_file()
    after = (workbook_path.stat().st_size, workbook_path.stat().st_mtime_ns)
    assert after == before


def test_cli_runs_workbook_quality_benchmark(
    tmp_path: Path,
    capsys,
):
    from langparse.cli import main

    manifest_path, _ = _write_manifest(tmp_path)
    output_dir = tmp_path / "cli-reports"

    exit_code = main(
        [
            "benchmark-workbook-quality",
            str(manifest_path),
            "--output-dir",
            str(output_dir),
        ]
    )

    assert exit_code == 0
    assert "workbook quality benchmark completed" in capsys.readouterr().out


def test_report_digest_includes_artifact_options(tmp_path: Path):
    from langparse.services.workbook_quality_benchmark import WorkbookQualityBenchmarkService

    manifest_path, _ = _write_manifest(tmp_path)
    output_dir = tmp_path / "reports"
    service = WorkbookQualityBenchmarkService()

    with_markdown = service.run(manifest_path, output_dir=output_dir, markdown=True)
    without_markdown = service.run(manifest_path, output_dir=output_dir, markdown=False)

    assert with_markdown.run_digest != without_markdown.run_digest
    assert with_markdown.output_path.joinpath("workbook-quality-summary.md").is_file()
    assert not without_markdown.output_path.joinpath("workbook-quality-summary.md").exists()


def test_public_workbook_quality_seed_covers_structural_region_anchors(tmp_path: Path):
    from langparse.services.workbook_quality_benchmark import WorkbookQualityBenchmarkService
    from langparse.workbooks.quality.schema import load_workbook_quality_manifest

    manifest_path = Path("samples/workbook_quality/public-manifest.json")
    manifest = load_workbook_quality_manifest(manifest_path)

    report = WorkbookQualityBenchmarkService().run(
        manifest_path,
        output_dir=tmp_path / "public-seed-report",
    )

    assert {
        "adjacent-native-tables",
        "styled-adjacent-tables",
        "form-with-side-note",
    }.issubset({sample.sample_id for sample in manifest.samples})
    assert len(manifest.samples) >= 13
    assert report.summary["sample_count"] >= 13
    assert report.summary["status"] == "passed"
    assert report.summary["metrics"]["object_fact_recall"] == 1.0
    assert report.summary["metrics"]["object_semantic_recall"] == 0.0


def test_public_workbook_quality_fixtures_have_deterministic_zip_timestamps():
    fixture = Path("samples/workbook_quality/fixtures/simple_table.xlsx")

    with zipfile.ZipFile(fixture) as archive:
        timestamps = {entry.date_time for entry in archive.infolist()}

    assert timestamps == {(1980, 1, 1, 0, 0, 0)}


def test_public_workbook_quality_fixtures_have_deterministic_document_metadata():
    fixture = Path("samples/workbook_quality/fixtures/simple_table.xlsx")

    with zipfile.ZipFile(fixture) as archive:
        core_properties = archive.read("docProps/core.xml").decode("utf-8")

    assert (
        '<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        'xsi:type="dcterms:W3CDTF">2026-01-01T00:00:00Z</dcterms:modified>' in core_properties
    )


def test_public_workbook_quality_seed_contains_named_range():
    fixture = Path("samples/workbook_quality/fixtures/formula_hidden.xlsx")

    workbook = load_workbook(fixture, read_only=False, data_only=False)

    assert "RateInputs" in workbook.defined_names
    assert workbook.defined_names["RateInputs"].attr_text == "'Data'!$B$2:$B$3"


def test_cli_returns_nonzero_when_workbook_quality_gate_fails(tmp_path: Path):
    from langparse.cli import main

    manifest_path, _ = _write_manifest(tmp_path)
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["samples"][0]["expectation"]["sheets"][0]["blocks"][0]["kind"] = "form"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    exit_code = main(
        [
            "benchmark-workbook-quality",
            str(manifest_path),
            "--output-dir",
            str(tmp_path / "failed-report"),
        ]
    )

    assert exit_code == 1


def test_run_digest_changes_when_structural_truth_changes(tmp_path: Path):
    from langparse.services.workbook_quality_benchmark import WorkbookQualityBenchmarkService

    manifest_path, _ = _write_manifest(tmp_path)
    output_dir = tmp_path / "reports"
    service = WorkbookQualityBenchmarkService()
    first = service.run(manifest_path, output_dir=output_dir)

    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    headers = payload["samples"][0]["expectation"]["sheets"][0]["blocks"][0]["headers"]
    headers.reverse()
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")
    second = service.run(manifest_path, output_dir=output_dir)

    assert first.run_digest != second.run_digest


def test_report_replay_rejects_tampered_artifacts(tmp_path: Path):
    import pytest

    from langparse.services.workbook_quality_benchmark import (
        WorkbookQualityBenchmarkService,
        WorkbookQualityReportError,
    )

    manifest_path, _ = _write_manifest(tmp_path)
    output_dir = tmp_path / "reports"
    service = WorkbookQualityBenchmarkService()
    first = service.run(manifest_path, output_dir=output_dir)
    first.output_path.joinpath("workbook-quality-summary.json").write_text(
        "{}\n",
        encoding="utf-8",
    )

    with pytest.raises(WorkbookQualityReportError, match="Immutable report collision"):
        service.run(manifest_path, output_dir=output_dir)


def test_report_does_not_persist_cell_or_annotation_values(tmp_path: Path):
    from langparse.services.workbook_quality_benchmark import WorkbookQualityBenchmarkService

    manifest_path, _ = _write_manifest(tmp_path)
    report = WorkbookQualityBenchmarkService().run(
        manifest_path,
        output_dir=tmp_path / "reports",
    )
    report_text = "\n".join(
        path.read_text(encoding="utf-8") for path in report.output_path.iterdir() if path.is_file()
    )

    assert "Alpha" not in report_text
    assert '"Name"' not in report_text
    assert '"Value"' not in report_text


def test_rejects_sample_changed_after_manifest_load(tmp_path: Path):
    import pytest

    from langparse.services.parse_service import ParseService
    from langparse.services.workbook_quality_benchmark import (
        WorkbookQualityBenchmarkService,
        WorkbookQualityReportError,
    )

    manifest_path, first_path = _write_manifest(tmp_path)
    second_path = first_path.with_name("second.xlsx")
    _write_table(second_path)
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    second_sample = json.loads(json.dumps(payload["samples"][0]))
    second_sample["sample_id"] = "second-table"
    second_sample["path"] = second_path.name
    second_sample["sha256"] = "sha256:" + hashlib.sha256(second_path.read_bytes()).hexdigest()
    payload["samples"].append(second_sample)
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    class MutatingParseService:
        def __init__(self) -> None:
            self.delegate = ParseService()
            self.calls = 0

        def parse_result(self, path: Path):
            self.calls += 1
            if self.calls == 1:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Data"
                sheet.append(["Name", "Value"])
                sheet.append(["Changed", 999])
                workbook.save(second_path)
            return self.delegate.parse_result(path)

    service = WorkbookQualityBenchmarkService(parse_service=MutatingParseService())

    with pytest.raises(WorkbookQualityReportError, match="no longer matches manifest"):
        service.run(manifest_path, output_dir=tmp_path / "reports")
