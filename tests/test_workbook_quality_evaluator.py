import hashlib
import json
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook


def _write_table(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.save(path)


def _load_expectation(tmp_path: Path, workbook_path: Path):
    from langparse.workbooks.quality.schema import load_workbook_quality_manifest

    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "dataset_id": "quality-test",
                "dataset_version": "1",
                "split": "tuning",
                "source_root": ".",
                "quality_gate": {"minimum": {"block_recall": 0.0}, "maximum": {}},
                "samples": [
                    {
                        "sample_id": "table",
                        "path": workbook_path.name,
                        "sha256": (
                            "sha256:" + hashlib.sha256(workbook_path.read_bytes()).hexdigest()
                        ),
                        "expectation": {
                            "sheets": [
                                {
                                    "name": "Data",
                                    "blocks": [
                                        {
                                            "source_range": "A1:B2",
                                            "kind": "logical_table",
                                            "headers": [
                                                {"coordinate": "A", "path": ["Name"]},
                                                {"coordinate": "B", "path": ["Value"]},
                                            ],
                                            "rows": [
                                                {"source_range": "A1:B1", "role": "header"},
                                                {"source_range": "A2:B2", "role": "data"},
                                            ],
                                            "form_fields": [],
                                            "matrix_axes": {"rows": [], "columns": []},
                                        }
                                    ],
                                }
                            ],
                            "continuations": [],
                            "required_source_refs": ["Data!A1:B2"],
                            "objects": [],
                        },
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    return load_workbook_quality_manifest(manifest_path).samples[0].expectation


def _load_public_sample(sample_id: str):
    from langparse.parsers.excel_parser import ExcelParser
    from langparse.workbooks.quality.schema import load_workbook_quality_manifest

    manifest = load_workbook_quality_manifest(Path("samples/workbook_quality/public-manifest.json"))
    sample = next(item for item in manifest.samples if item.sample_id == sample_id)
    parsed = ExcelParser().parse_result(manifest.source_root / sample.path)
    return sample.expectation, parsed


def test_scores_end_to_end_logical_table_structure(tmp_path: Path):
    from langparse.parsers.excel_parser import ExcelParser
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    workbook_path = tmp_path / "table.xlsx"
    _write_table(workbook_path)
    expectation = _load_expectation(tmp_path, workbook_path)

    metrics = evaluate_workbook_result(
        expectation,
        ExcelParser().parse_result(workbook_path),
    )

    assert metrics.block_precision == 1.0
    assert metrics.block_recall == 1.0
    assert metrics.header_path_accuracy == 1.0
    assert metrics.row_role_f1 == 1.0
    assert metrics.source_ref_completeness == 1.0
    assert metrics.fallback_rate == 0.0
    assert metrics.object_fact_recall is None


def test_duplicate_observed_block_reduces_precision(tmp_path: Path):
    from langparse.parsers.excel_parser import ExcelParser
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    workbook_path = tmp_path / "table.xlsx"
    _write_table(workbook_path)
    expectation = _load_expectation(tmp_path, workbook_path)
    parsed = ExcelParser().parse_result(workbook_path)
    parsed.structure.sheets[0].blocks.append(deepcopy(parsed.structure.sheets[0].blocks[0]))

    metrics = evaluate_workbook_result(expectation, parsed)

    assert metrics.block_precision == 0.5
    assert metrics.block_recall == 1.0


def test_source_less_block_is_a_false_positive_and_fallback(tmp_path: Path):
    from langparse.parsers.excel_parser import ExcelParser
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    workbook_path = tmp_path / "table.xlsx"
    _write_table(workbook_path)
    expectation = _load_expectation(tmp_path, workbook_path)
    parsed = ExcelParser().parse_result(workbook_path)
    invalid = deepcopy(parsed.structure.sheets[0].blocks[0])
    invalid.block_id = "source-less"
    invalid.source_refs = []
    parsed.structure.sheets[0].blocks.append(invalid)

    metrics = evaluate_workbook_result(expectation, parsed)

    assert metrics.block_precision == 0.5
    assert metrics.block_recall == 1.0
    assert metrics.fallback_rate == 0.5


def test_form_and_matrix_partial_errors_reduce_accuracy():
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    form_expectation, form_parsed = _load_public_sample("form")
    form_parsed.structure.sheets[0].blocks[0].form.fields[0].value = "incorrect"
    matrix_expectation, matrix_parsed = _load_public_sample("matrix")
    matrix_parsed.structure.sheets[0].blocks[0].matrix.row_headers[0].value = "incorrect"

    form_metrics = evaluate_workbook_result(form_expectation, form_parsed)
    matrix_metrics = evaluate_workbook_result(matrix_expectation, matrix_parsed)

    assert form_metrics.form_field_exact_match == 0.5
    assert matrix_metrics.matrix_axis_accuracy == 0.75


def test_continuation_order_is_part_of_quality_score():
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    expectation, parsed = _load_public_sample("cross-sheet-continuation")
    parsed.structure.table_continuations[0].source_refs.reverse()

    metrics = evaluate_workbook_result(expectation, parsed)

    assert metrics.continuation_precision == 0.0
    assert metrics.continuation_recall == 0.0


def test_incorrect_object_fact_reduces_precision_and_recall():
    from langparse.workbooks.quality.evaluator import evaluate_workbook_result

    expectation, parsed = _load_public_sample("chart-facts")
    parsed.structure.snapshot.sheets[0].objects[0]["anchor"] = "E2"

    metrics = evaluate_workbook_result(expectation, parsed)

    assert metrics.object_fact_precision == 0.0
    assert metrics.object_fact_recall == 0.0
