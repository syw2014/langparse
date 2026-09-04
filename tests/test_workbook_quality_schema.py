import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    workbook.save(path)


def _sha256(path: Path) -> str:
    return f"sha256:{hashlib.sha256(path.read_bytes()).hexdigest()}"


def _empty_manifest() -> dict:
    return {
        "schema_version": 1,
        "dataset_id": "quality-test",
        "dataset_version": "1",
        "split": "tuning",
        "source_root": ".",
        "quality_gate": {"minimum": {}, "maximum": {}},
        "samples": [],
    }


def _empty_expectation() -> dict:
    return {
        "sheets": [],
        "continuations": [],
        "required_source_refs": [],
        "objects": [],
    }


def _valid_expectation() -> dict:
    return {
        "sheets": [
            {
                "name": "Data",
                "blocks": [
                    {
                        "source_range": "A1:B2",
                        "kind": "logical_table",
                        "headers": [{"coordinate": "A", "path": ["Name"]}],
                        "rows": [{"source_range": "A2:B2", "role": "data"}],
                        "form_fields": [{"label": "Name", "value": "Alpha"}],
                        "matrix_axes": {"rows": ["Alpha"], "columns": ["Value"]},
                    }
                ],
            }
        ],
        "continuations": [],
        "required_source_refs": ["Data!A1:B2"],
        "objects": [{"sheet_name": "Data", "kind": "chart", "anchor": "D2"}],
    }


def _write_manifest_with_expectation(tmp_path: Path, expectation: dict) -> Path:
    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    payload = _empty_manifest()
    payload["quality_gate"]["minimum"] = {"block_recall": 0.0}
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": expectation,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")
    return manifest_path


def _continuation_expectation(groups: list[list[str]]) -> dict:
    expectation = _valid_expectation()
    template = expectation["sheets"][0]["blocks"][0]
    blocks = []
    for source_range in ("A1:B2", "C1:D2", "E1:F2"):
        block = json.loads(json.dumps(template))
        block["source_range"] = source_range
        block["headers"] = []
        block["rows"] = []
        block["form_fields"] = []
        block["matrix_axes"] = {"rows": [], "columns": []}
        blocks.append(block)
    expectation["sheets"][0]["blocks"] = blocks
    expectation["continuations"] = groups
    expectation["required_source_refs"] = ["Data!A1:B2"]
    expectation["objects"] = []
    return expectation


def test_loads_versioned_workbook_quality_manifest(tmp_path: Path):
    from langparse.workbooks.quality.schema import load_workbook_quality_manifest

    fixtures = tmp_path / "fixtures"
    fixtures.mkdir()
    workbook_path = fixtures / "table.xlsx"
    _write_workbook(workbook_path)
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "dataset_id": "workbook-quality-seed",
                "dataset_version": "1",
                "split": "tuning",
                "source_root": "fixtures",
                "quality_gate": {
                    "minimum": {"block_recall": 1.0},
                    "maximum": {"fallback_rate": 0.0},
                },
                "samples": [
                    {
                        "sample_id": "simple-table",
                        "path": "table.xlsx",
                        "sha256": _sha256(workbook_path),
                        "expectation": {
                            "sheets": [
                                {
                                    "name": "Data",
                                    "blocks": [
                                        {
                                            "source_range": "A1:B2",
                                            "kind": "logical_table",
                                            "headers": [
                                                {"coordinate": "A", "path": ["Name"]},
                                                {"coordinate": "B", "path": ["Value"]},
                                            ],
                                            "rows": [
                                                {"source_range": "A1:B1", "role": "header"},
                                                {"source_range": "A2:B2", "role": "data"},
                                            ],
                                            "form_fields": [],
                                            "matrix_axes": {"rows": [], "columns": []},
                                        }
                                    ],
                                }
                            ],
                            "continuations": [],
                            "required_source_refs": ["Data!A1:B2"],
                            "objects": [],
                        },
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    manifest = load_workbook_quality_manifest(manifest_path)

    assert manifest.schema_version == 1
    assert manifest.dataset_id == "workbook-quality-seed"
    assert manifest.split == "tuning"
    assert manifest.quality_gate.minimum == {"block_recall": 1.0}
    assert manifest.quality_gate.maximum == {"fallback_rate": 0.0}
    assert manifest.samples[0].expectation.sheets[0].blocks[0].kind == "logical_table"
    assert manifest.samples[0].expectation.required_source_refs == ("Data!A1:B2",)


def test_rejects_unknown_manifest_fields(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "dataset_id": "quality-test",
                "dataset_version": "1",
                "split": "tuning",
                "source_root": ".",
                "quality_gate": {"minimum": {}, "maximum": {}},
                "samples": [],
                "unexpected": True,
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(WorkbookQualityManifestError, match="Manifest keys"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_sample_path_outside_source_root(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    source_root = tmp_path / "fixtures"
    source_root.mkdir()
    workbook_path = tmp_path / "outside.xlsx"
    _write_workbook(workbook_path)
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "dataset_id": "quality-test",
                "dataset_version": "1",
                "split": "tuning",
                "source_root": "fixtures",
                "quality_gate": {"minimum": {}, "maximum": {}},
                "samples": [
                    {
                        "sample_id": "outside",
                        "path": "../outside.xlsx",
                        "sha256": _sha256(workbook_path),
                        "expectation": {
                            "sheets": [],
                            "continuations": [],
                            "required_source_refs": [],
                            "objects": [],
                        },
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(WorkbookQualityManifestError, match="outside source_root"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unsupported_schema_version(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["schema_version"] = 2
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="schema_version"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_dataset_split(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["split"] = "development"
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="split"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_non_string_dataset_split(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["split"] = []
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="split"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_quality_gate_metric(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["quality_gate"]["minimum"] = {"imaginary_accuracy": 1.0}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Unknown quality metric"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_duplicate_json_keys(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        '{"schema_version":1,"schema_version":1,"dataset_id":"quality-test",'
        '"dataset_version":"1","split":"tuning","source_root":".",'
        '"quality_gate":{"minimum":{},"maximum":{}},"samples":[]}',
        encoding="utf-8",
    )

    with pytest.raises(WorkbookQualityManifestError, match="Duplicate JSON key"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_sample_fields(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": _empty_expectation(),
            "unexpected": True,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Sample keys"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_duplicate_sample_ids(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    sample = {
        "sample_id": "table",
        "path": workbook_path.name,
        "sha256": _sha256(workbook_path),
        "expectation": _empty_expectation(),
    }
    payload = _empty_manifest()
    payload["samples"] = [sample, dict(sample)]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Duplicate sample_id"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_expectation_fields(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    expectation = _empty_expectation()
    expectation["unexpected"] = True
    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": expectation,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Expectation keys"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_malformed_required_source_ref(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    expectation = _empty_expectation()
    expectation["required_source_refs"] = ["Data!not-a-range"]
    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": expectation,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="source ref"):
        load_workbook_quality_manifest(manifest_path)


@pytest.mark.parametrize("value", [True, -0.1, 1.1, "1.0"])
def test_rejects_invalid_quality_gate_threshold(tmp_path: Path, value):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["quality_gate"]["minimum"] = {"block_recall": value}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="threshold"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_block_fields(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    expectation = _empty_expectation()
    expectation["sheets"] = [
        {
            "name": "Data",
            "blocks": [
                {
                    "source_range": "A1:B2",
                    "kind": "logical_table",
                    "headers": [],
                    "rows": [],
                    "form_fields": [],
                    "matrix_axes": {"rows": [], "columns": []},
                    "unexpected": True,
                }
            ],
        }
    ]
    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": expectation,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Block keys"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_unknown_block_kind(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    workbook_path = tmp_path / "table.xlsx"
    _write_workbook(workbook_path)
    expectation = _empty_expectation()
    expectation["sheets"] = [
        {
            "name": "Data",
            "blocks": [
                {
                    "source_range": "A1:B2",
                    "kind": "magic_table",
                    "headers": [],
                    "rows": [],
                    "form_fields": [],
                    "matrix_axes": {"rows": [], "columns": []},
                }
            ],
        }
    ]
    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "table",
            "path": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "expectation": expectation,
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="block kind"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_absolute_source_root(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["source_root"] = str(tmp_path)
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="source_root"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_invalid_quality_gate_shape(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["quality_gate"] = {"minimum": {}, "unexpected": {}}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="Quality gate keys"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_invalid_dataset_identifier(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["dataset_id"] = "unsafe/id"
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="dataset_id"):
        load_workbook_quality_manifest(manifest_path)


def test_wraps_missing_sample_file_as_manifest_error(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["samples"] = [
        {
            "sample_id": "missing",
            "path": "missing.xlsx",
            "sha256": "sha256:" + "0" * 64,
            "expectation": _empty_expectation(),
        }
    ]
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="does not exist"):
        load_workbook_quality_manifest(manifest_path)


def test_wraps_invalid_json_as_manifest_error(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text("{not-json", encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="valid JSON"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_manifest_without_samples(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    payload = _empty_manifest()
    payload["quality_gate"]["minimum"] = {"block_recall": 0.0}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="at least one sample"):
        load_workbook_quality_manifest(manifest_path)


def test_rejects_quality_gate_without_thresholds(tmp_path: Path):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    manifest_path = _write_manifest_with_expectation(tmp_path, _empty_expectation())
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["quality_gate"] = {"minimum": {}, "maximum": {}}
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(WorkbookQualityManifestError, match="at least one threshold"):
        load_workbook_quality_manifest(manifest_path)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (lambda value: value["sheets"][0].update({"unexpected": True}), "Sheet keys"),
        (lambda value: value["sheets"][0].update({"name": 7}), "Sheet name"),
        (
            lambda value: value["sheets"][0]["blocks"][0].update({"source_range": 7}),
            "Block source_range",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["headers"][0].update(
                {"unexpected": True}
            ),
            "Header keys",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["headers"][0].update({"path": "Name"}),
            "Header path",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["rows"][0].update({"unexpected": True}),
            "Row keys",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["form_fields"][0].update(
                {"unexpected": True}
            ),
            "Form field keys",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0].update(
                {"matrix_axes": {"rows": [], "unexpected": []}}
            ),
            "Matrix axes keys",
        ),
        (lambda value: value["objects"][0].update({"unexpected": True}), "Object keys"),
    ],
)
def test_rejects_invalid_nested_truth_shapes(tmp_path: Path, mutation, message: str):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    expectation = _valid_expectation()
    mutation(expectation)
    manifest_path = _write_manifest_with_expectation(tmp_path, expectation)

    with pytest.raises(WorkbookQualityManifestError, match=message):
        load_workbook_quality_manifest(manifest_path)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (
            lambda value: value["sheets"][0]["blocks"][0]["rows"][0].update({"role": "invented"}),
            "Unknown row role",
        ),
        (
            lambda value: value["objects"][0].update({"kind": "diagram"}),
            "Unknown object kind",
        ),
        (
            lambda value: value.update({"continuations": [["Data!A1:B2", "Missing!A1:B2"]]}),
            "Continuation source ref",
        ),
        (
            lambda value: value["objects"][0].update({"sheet_name": "Missing"}),
            "Object sheet_name",
        ),
        (
            lambda value: value["sheets"].append(dict(value["sheets"][0])),
            "Duplicate sheet name",
        ),
        (
            lambda value: value.update({"required_source_refs": ["Missing!A1:B2"]}),
            "Required source ref",
        ),
        (
            lambda value: value.update({"required_source_refs": ["Data!A1:B2", "Data!A1:B2"]}),
            "Duplicate required_source_ref",
        ),
        (
            lambda value: value["sheets"][0]["blocks"].append(
                dict(value["sheets"][0]["blocks"][0])
            ),
            "Duplicate block truth",
        ),
        (
            lambda value: value["objects"].append(dict(value["objects"][0])),
            "Duplicate object truth",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["headers"].append(
                dict(value["sheets"][0]["blocks"][0]["headers"][0])
            ),
            "Duplicate header truth",
        ),
        (
            lambda value: value["sheets"][0]["blocks"][0]["rows"].append(
                dict(value["sheets"][0]["blocks"][0]["rows"][0])
            ),
            "Duplicate row truth",
        ),
    ],
)
def test_rejects_inconsistent_nested_truth(tmp_path: Path, mutation, message: str):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    expectation = _valid_expectation()
    mutation(expectation)
    manifest_path = _write_manifest_with_expectation(tmp_path, expectation)

    with pytest.raises(WorkbookQualityManifestError, match=message):
        load_workbook_quality_manifest(manifest_path)


@pytest.mark.parametrize(
    ("groups", "message"),
    [
        ([["Data!A1:B2", "Data!A1:B2"]], "duplicate source refs"),
        (
            [
                ["Data!A1:B2", "Data!C1:D2"],
                ["Data!A1:B2", "Data!C1:D2"],
            ],
            "Duplicate continuation group",
        ),
        (
            [
                ["Data!A1:B2", "Data!C1:D2"],
                ["Data!C1:D2", "Data!E1:F2"],
            ],
            "overlapping source refs",
        ),
    ],
)
def test_rejects_ambiguous_continuation_truth(tmp_path: Path, groups, message: str):
    from langparse.workbooks.quality.schema import (
        WorkbookQualityManifestError,
        load_workbook_quality_manifest,
    )

    manifest_path = _write_manifest_with_expectation(tmp_path, _continuation_expectation(groups))

    with pytest.raises(WorkbookQualityManifestError, match=message):
        load_workbook_quality_manifest(manifest_path)
