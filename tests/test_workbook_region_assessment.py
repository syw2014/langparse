from openpyxl.utils import get_column_letter, range_boundaries

from langparse.workbooks.classification import (
    assess_candidate_region,
    classify_candidate_region,
)
from langparse.workbooks.types import CandidateRegion, CellSnapshot, SheetSnapshot, SourceRef


def region(values: dict[str, object], source_range: str) -> tuple[SheetSnapshot, CandidateRegion]:
    sheet = SheetSnapshot(name="Data", index=0, used_range=source_range)
    for coordinate, value in values.items():
        sheet.cells[coordinate] = CellSnapshot(
            coordinate=coordinate,
            raw_value=value,
            display_value=str(value),
        )

    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    coordinates = [
        f"{get_column_letter(column)}{row}"
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
        if f"{get_column_letter(column)}{row}" in sheet.cells
    ]
    candidate = CandidateRegion(
        source_ref=SourceRef(sheet_name=sheet.name, range=source_range),
        cell_refs=coordinates,
    )
    return sheet, candidate


def test_clear_table_assessment_preserves_the_existing_winner():
    sheet, candidate = region(
        {
            "A1": "Name",
            "B1": "Value",
            "A2": "Alpha",
            "B2": "1",
            "A3": "Beta",
            "B3": "2",
        },
        "A1:B3",
    )

    assessment = assess_candidate_region(sheet, candidate)

    assert assessment.deterministic.kind == "logical_table"
    assert assessment.ambiguous is False
    assert classify_candidate_region(sheet, candidate) == assessment.deterministic


def test_sparse_text_region_registers_choice_only_ambiguity():
    sheet, candidate = region({"A1": "左上", "B2": "右下"}, "A1:B2")

    assessment = assess_candidate_region(sheet, candidate)

    assert assessment.deterministic.kind == "unclassified"
    assert assessment.ambiguous is True
    assert [choice.kind for choice in assessment.choices] == ["unclassified", "text"]
    assert assessment.ambiguity_codes == ("unclassified_with_compatible_choices",)
    assert len({choice.choice_id for choice in assessment.choices}) == 2


def test_unclassified_region_without_a_second_compatible_kind_is_not_ambiguous():
    sheet, candidate = region({"A1": ""}, "A1:A1")

    assessment = assess_candidate_region(sheet, candidate)

    assert [choice.kind for choice in assessment.choices] == ["unclassified"]
    assert assessment.ambiguous is False
    assert assessment.ambiguity_codes == ()


def test_choice_ids_are_stable_for_the_same_source_facts():
    first_sheet, first_candidate = region({"A1": "左上", "B2": "右下"}, "A1:B2")
    second_sheet, second_candidate = region({"A1": "左上", "B2": "右下"}, "A1:B2")

    first = assess_candidate_region(first_sheet, first_candidate)
    second = assess_candidate_region(second_sheet, second_candidate)

    assert [choice.choice_id for choice in first.choices] == [
        choice.choice_id for choice in second.choices
    ]


def test_choice_ids_change_when_local_structural_features_change():
    first_sheet, first_candidate = region({"A1": "left", "B2": "right"}, "A1:B2")
    second_sheet, second_candidate = region({"A1": "left", "A2": "right"}, "A1:B2")

    first = assess_candidate_region(first_sheet, first_candidate)
    second = assess_candidate_region(second_sheet, second_candidate)

    assert [(choice.kind, choice.local_score, choice.reason_codes) for choice in first.choices] == [
        (choice.kind, choice.local_score, choice.reason_codes) for choice in second.choices
    ]
    assert first.deterministic.features != second.deterministic.features
    assert [choice.choice_id for choice in first.choices] != [
        choice.choice_id for choice in second.choices
    ]
