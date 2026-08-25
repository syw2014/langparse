from openpyxl.utils import get_column_letter

from langparse.workbooks.assembly import assemble_baseline, assemble_workbook
from langparse.workbooks.rendering import compatibility_pages, render_workbook_markdown
from langparse.workbooks.types import CellSnapshot, SheetSnapshot, WorkbookSnapshot


def test_renderer_uses_coordinates_not_unnamed_headers():
    snapshot = WorkbookSnapshot(
        source="book.xlsx",
        filename="book.xlsx",
        sheets=[
            SheetSnapshot(
                name="Cover",
                index=0,
                used_range="A1:C2",
                cells={
                    "A1": CellSnapshot(coordinate="A1", raw_value="Title", display_value="Title"),
                    "C2": CellSnapshot(coordinate="C2", raw_value=3, display_value="3"),
                },
            )
        ],
    )
    ir, _ = assemble_baseline(snapshot)

    markdown = render_workbook_markdown(snapshot, ir)

    assert "## Sheet: Cover" in markdown
    assert "<!-- source_range: Cover!A1:C2 -->" in markdown
    assert "| A | B | C |" in markdown
    assert "Unnamed:" not in markdown

    pages = compatibility_pages(snapshot, ir)
    assert pages[0].metadata == {
        "part_kind": "sheet",
        "sheet_name": "Cover",
        "source_range": "A1:C2",
    }
    assert pages[0].tables[0]["rows"][0] == ["A", "B", "C"]
    assert pages[0].tables[0]["rows"][1:] == [["Title", "", ""], ["", "", "3"]]
    assert pages[0].tables[0]["row_numbers"] == [1, 2]


def test_renderer_escapes_markdown_and_leaves_merged_subordinates_empty():
    snapshot = WorkbookSnapshot(
        source="book.xlsx",
        filename="book.xlsx",
        sheets=[
            SheetSnapshot(
                name="Data",
                index=0,
                used_range="B3:C3",
                cells={
                    "B3": CellSnapshot(
                        coordinate="B3",
                        raw_value="A|B\nC",
                        display_value="A|B\nC",
                        colspan=2,
                    ),
                    "C3": CellSnapshot(coordinate="C3", merge_anchor="B3"),
                },
            )
        ],
    )
    ir, _ = assemble_baseline(snapshot)

    markdown = render_workbook_markdown(snapshot, ir)

    assert "| A\\|B<br>C |  |" in markdown


def test_empty_sheet_is_preserved_without_a_fake_table():
    snapshot = WorkbookSnapshot(
        source="book.xlsx",
        filename="book.xlsx",
        sheets=[SheetSnapshot(name="Empty", index=0)],
    )
    ir, _ = assemble_baseline(snapshot)

    pages = compatibility_pages(snapshot, ir)

    assert len(pages) == 1
    assert pages[0].tables == []
    assert "## Sheet: Empty" in pages[0].markdown_content


def test_semantic_renderer_deduplicates_print_fragments_and_preserves_snapshot():
    sheet = SheetSnapshot(name="Data", index=0, used_range="A1:D11")

    def put_row(row_number: int, values: list[object | None]) -> None:
        for column_number, value in enumerate(values, start=1):
            if value is None:
                continue
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            sheet.cells[coordinate] = CellSnapshot(
                coordinate=coordinate,
                raw_value=value,
                display_value=str(value),
            )

    put_row(1, ["清单"])
    put_row(2, ["单位工程", "第 1 页 共 2 页"])
    put_row(3, ["序号", "名称", "说明", "金额"])
    put_row(4, [0, "土方", "", 100])
    put_row(5, [1, "挖土", "", 40])
    put_row(6, ["清单"])
    put_row(7, ["单位工程", "第 2 页 共 2 页"])
    put_row(8, ["序号", "名称", "说明", "金额"])
    put_row(9, [2, "回填", "", 60])
    put_row(10, ["合计", "", "", 100])
    snapshot = WorkbookSnapshot(source="book.xlsx", filename="book.xlsx", sheets=[sheet])
    ir, _ = assemble_workbook(snapshot)

    markdown = render_workbook_markdown(snapshot, ir)

    assert markdown.count("### Table: 清单") == 1
    assert markdown.count("| 序号 | 名称 | 说明 | 金额 |") == 1
    assert "#### Section: 土方" in markdown
    assert "第 2 页 共 2 页" not in markdown
    assert "合计" in markdown
    assert ir.snapshot is snapshot
    assert ir.snapshot.sheets[0].cells["A6"].display_value == "清单"


def test_renderer_emits_every_mixed_block_and_keeps_full_compatibility_grid():
    sheet = SheetSnapshot(name="Mixed", index=0, used_range="A1:B8")

    def put(coordinate: str, value: object) -> None:
        sheet.cells[coordinate] = CellSnapshot(
            coordinate=coordinate,
            raw_value=value,
            display_value=str(value),
        )

    for coordinate, value in {
        "A1": "项目登记",
        "A2": "项目名称",
        "B2": "道路工程",
        "A3": "建设单位",
        "B3": "示例公司",
        "A6": "Name",
        "B6": "Value",
        "A7": "Alpha",
        "B7": 1,
        "A8": "Beta",
        "B8": 2,
    }.items():
        put(coordinate, value)
    snapshot = WorkbookSnapshot(source="book.xlsx", filename="book.xlsx", sheets=[sheet])
    ir, _ = assemble_workbook(snapshot)

    markdown = render_workbook_markdown(snapshot, ir)
    pages = compatibility_pages(snapshot, ir)

    assert markdown.index("### Form: 项目登记") < markdown.index("| Name | Value |")
    assert "### Table: Name" not in markdown
    assert "| Field | Value |" in markdown
    assert "| 项目名称 | 道路工程 |" in markdown
    assert pages[0].metadata["source_range"] == "A1:B8"
    assert pages[0].tables[0]["source_range"] == "A1:B8"
    assert len(pages[0].tables[0]["rows"]) == 9


def test_renderer_preserves_matrix_axes_and_text_line_order():
    matrix_sheet = SheetSnapshot(name="Matrix", index=0, used_range="A1:C3")
    for row_number, values in enumerate(
        [["指标", "1月", "2月"], ["收入", 10, 12], ["成本", 3, 4]], start=1
    ):
        for column_number, value in enumerate(values, start=1):
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            matrix_sheet.cells[coordinate] = CellSnapshot(
                coordinate=coordinate,
                raw_value=value,
                display_value=str(value),
            )
    text_sheet = SheetSnapshot(name="Notes", index=1, used_range="A1:A2")
    text_sheet.cells = {
        "A1": CellSnapshot(coordinate="A1", raw_value="第一行", display_value="第一行"),
        "A2": CellSnapshot(coordinate="A2", raw_value="第二行", display_value="第二行"),
    }
    snapshot = WorkbookSnapshot(
        source="book.xlsx", filename="book.xlsx", sheets=[matrix_sheet, text_sheet]
    )
    ir, _ = assemble_workbook(snapshot)

    markdown = render_workbook_markdown(snapshot, ir)

    assert "### Matrix: 指标" in markdown
    assert "|  | 1月 | 2月 |" in markdown
    assert "| 收入 | 10 | 12 |" in markdown
    assert markdown.index("第一行") < markdown.index("第二行")
    assert "Unnamed:" not in markdown


def test_renderer_marks_continuation_members_without_rendering_the_aggregate(tmp_path):
    from openpyxl import Workbook

    from langparse.parsers.excel_parser import ExcelParser

    path = tmp_path / "continuation.xlsx"
    workbook = Workbook()
    for index, sheet_name in enumerate(("清单1", "清单2"), start=1):
        sheet = workbook.active if index == 1 else workbook.create_sheet()
        sheet.title = sheet_name
        sheet.append(["清单"])
        sheet.append(["单位工程", f"第 {index} 页 共 2 页"])
        sheet.append(["序号", "名称", "说明", "金额"])
        sheet.append([index, "土方" if index == 1 else "回填", "", 100])
    workbook.save(path)

    parsed = ExcelParser().parse_result(path)
    group = parsed.structure.table_continuations[0]

    markdown = render_workbook_markdown(parsed.structure.snapshot, parsed.structure)

    assert markdown == (
        f"## Sheet: 清单1\n\n"
        "<!-- source_ranges: 清单1!A1:D4 -->\n\n"
        f"<!-- continuation_id: {group.continuation_id}; role: head -->\n\n"
        "### Table: 清单\n\n"
        "> 单位工程 | 第 1 页 共 2 页\n\n"
        "| 序号 | 名称 | 说明 | 金额 |\n"
        "| --- | --- | --- | --- |\n"
        "| 1 | 土方 |  | 100 |\n\n"
        "## Sheet: 清单2\n\n"
        "<!-- source_ranges: 清单2!A1:D4 -->\n\n"
        f"<!-- continuation_id: {group.continuation_id}; role: tail -->\n\n"
        "### Table: 清单\n\n"
        "> 单位工程 | 第 2 页 共 2 页\n\n"
        "| 序号 | 名称 | 说明 | 金额 |\n"
        "| --- | --- | --- | --- |\n"
        "| 2 | 回填 |  | 100 |"
    )
    assert markdown.count(group.continuation_id) == 2
    assert markdown.count("### Table: 清单") == 2
    assert group.logical_table.table_id not in markdown
