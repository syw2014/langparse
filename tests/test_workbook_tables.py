from openpyxl.utils import get_column_letter

from langparse.workbooks.tables import interpret_logical_table
from langparse.workbooks.types import CandidateRegion, CellSnapshot, SheetSnapshot, SourceRef


def _put_row(sheet: SheetSnapshot, row_number: int, values: list[object | None]) -> None:
    for column_number, value in enumerate(values, start=1):
        if value is None:
            continue
        coordinate = f"{get_column_letter(column_number)}{row_number}"
        sheet.cells[coordinate] = CellSnapshot(
            coordinate=coordinate,
            raw_value=value,
            display_value=str(value),
        )


def _two_fragment_sheet() -> tuple[SheetSnapshot, CandidateRegion]:
    sheet = SheetSnapshot(name="Data", index=0, used_range="A1:L12")
    header_1 = [
        "序号",
        "项目编码",
        "项目名称",
        "项目特征描述",
        "计量单位",
        "工程量",
        "综合单价(元)",
        "合价(元)",
        "其中",
        None,
        None,
        "备注",
    ]
    header_2 = [None, None, None, None, None, None, None, None, "人工费", "机械费", "管理费", None]
    for offset, page_number in ((0, 1), (6, 2)):
        _put_row(sheet, 1 + offset, ["表1-2 清单"])
        _put_row(
            sheet,
            2 + offset,
            [
                "单位工程名称",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                f"第 {page_number} 页 共 2 页",
            ],
        )
        _put_row(sheet, 3 + offset, header_1)
        _put_row(sheet, 4 + offset, header_2)
        _put_row(sheet, 5 + offset, [1 + offset, "code", "item"])
        _put_row(sheet, 6 + offset, [2 + offset, "code", "item"])
        for column in range(1, 9):
            anchor = f"{get_column_letter(column)}{3 + offset}"
            subordinate = f"{get_column_letter(column)}{4 + offset}"
            sheet.cells[subordinate] = CellSnapshot(
                coordinate=subordinate,
                merge_anchor=anchor,
            )
        for column in (10, 11):
            coordinate = f"{get_column_letter(column)}{3 + offset}"
            sheet.cells[coordinate] = CellSnapshot(
                coordinate=coordinate,
                merge_anchor=f"I{3 + offset}",
            )
        sheet.cells[f"L{4 + offset}"] = CellSnapshot(
            coordinate=f"L{4 + offset}",
            merge_anchor=f"L{3 + offset}",
        )
    candidate = CandidateRegion(
        source_ref=SourceRef(sheet_name="Data", range="A1:L12"),
        cell_refs=list(sheet.cells),
    )
    return sheet, candidate


def test_detects_repeated_print_fragments():
    sheet, candidate = _two_fragment_sheet()

    table = interpret_logical_table(sheet, candidate)

    assert [fragment.page_number for fragment in table.fragments] == [1, 2]
    assert [fragment.source_ref.range for fragment in table.fragments] == ["A1:L6", "A7:L12"]
    assert table.fragments[1].header_row_numbers == [9, 10]


def test_builds_multi_row_header_paths_from_merges():
    sheet, candidate = _two_fragment_sheet()

    table = interpret_logical_table(sheet, candidate)
    paths = {column.coordinate: column.path for column in table.columns}

    assert paths["A"] == ["序号"]
    assert paths["I"] == ["其中", "人工费"]
    assert paths["J"] == ["其中", "机械费"]
    assert paths["K"] == ["其中", "管理费"]
